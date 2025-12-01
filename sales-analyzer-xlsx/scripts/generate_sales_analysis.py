#!/usr/bin/env python3
"""
Sales Analyzer XLSX Generator
売上データからABC分析・パレート図・ダッシュボード付きExcelを生成
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from pathlib import Path
import argparse

# スタイル定義（Jony Iveデザイン - グレースケール）
HEADER_FILL = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="333333")
KPI_FONT = Font(bold=True, size=20, color="333333")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ABCランク色（グレースケール）
RANK_COLORS = {
    "A": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
    "B": PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
    "C": PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"),
}


def load_and_prepare_data(input_file, date_col, amount_col, category_col, item_col, staff_col):
    """CSVを読み込み、前処理"""
    df = pd.read_csv(input_file, encoding='utf-8')
    df[date_col] = pd.to_datetime(df[date_col])
    df['年月'] = df[date_col].dt.to_period('M').astype(str)
    return df


def create_dashboard_sheet(wb, df, date_col, amount_col, category_col, item_col, staff_col):
    """ダッシュボードシート作成"""
    ws = wb.create_sheet("ダッシュボード", 0)
    
    # タイトル
    ws['B2'] = "売上分析ダッシュボード"
    ws['B2'].font = Font(bold=True, size=24, color="2E5090")
    
    # KPIセクション
    ws['B4'] = "総売上"
    ws['B5'] = df[amount_col].sum()
    ws['B5'].number_format = '¥#,##0'
    ws['B5'].font = KPI_FONT
    
    ws['D4'] = "取引件数"
    ws['D5'] = len(df)
    ws['D5'].font = KPI_FONT
    
    ws['F4'] = "平均単価"
    ws['F5'] = df[amount_col].mean()
    ws['F5'].number_format = '¥#,##0'
    ws['F5'].font = KPI_FONT
    
    # 月別売上データ（グラフ用）
    monthly = df.groupby('年月')[amount_col].sum().reset_index()
    monthly = monthly.sort_values('年月')
    
    ws['B8'] = "月別売上推移"
    ws['B8'].font = TITLE_FONT
    
    ws['B9'] = "年月"
    ws['C9'] = "売上"
    ws['B9'].font = HEADER_FONT
    ws['C9'].font = HEADER_FONT
    ws['B9'].fill = HEADER_FILL
    ws['C9'].fill = HEADER_FILL
    
    for i, row in enumerate(monthly.values, start=10):
        ws[f'B{i}'] = row[0]
        ws[f'C{i}'] = row[1]
        ws[f'C{i}'].number_format = '¥#,##0'
    
    # 月別売上グラフ
    chart = BarChart()
    chart.type = "col"
    chart.title = "月別売上推移"
    chart.style = 10
    chart.y_axis.title = "売上（円）"
    chart.x_axis.title = "年月"
    
    data = Reference(ws, min_col=3, min_row=9, max_row=9+len(monthly))
    cats = Reference(ws, min_col=2, min_row=10, max_row=9+len(monthly))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 15
    chart.height = 8
    ws.add_chart(chart, "E8")
    
    # カテゴリ別構成比
    cat_sales = df.groupby(category_col)[amount_col].sum().reset_index()
    cat_sales = cat_sales.sort_values(amount_col, ascending=False)
    
    start_row = 10 + len(monthly) + 3
    ws[f'B{start_row}'] = "カテゴリ別売上構成"
    ws[f'B{start_row}'].font = TITLE_FONT
    
    ws[f'B{start_row+1}'] = "カテゴリ"
    ws[f'C{start_row+1}'] = "売上"
    ws[f'B{start_row+1}'].font = HEADER_FONT
    ws[f'C{start_row+1}'].font = HEADER_FONT
    ws[f'B{start_row+1}'].fill = HEADER_FILL
    ws[f'C{start_row+1}'].fill = HEADER_FILL
    
    for i, row in enumerate(cat_sales.values, start=start_row+2):
        ws[f'B{i}'] = row[0]
        ws[f'C{i}'] = row[1]
        ws[f'C{i}'].number_format = '¥#,##0'
    
    # 円グラフ
    pie = PieChart()
    pie.title = "カテゴリ別構成比"
    labels = Reference(ws, min_col=2, min_row=start_row+2, max_row=start_row+1+len(cat_sales))
    data = Reference(ws, min_col=3, min_row=start_row+1, max_row=start_row+1+len(cat_sales))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 10
    pie.height = 8
    ws.add_chart(pie, f"E{start_row}")
    
    # 列幅調整
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    return ws


def create_abc_analysis_sheet(wb, df, amount_col, item_col):
    """ABC分析シート作成"""
    ws = wb.create_sheet("ABC分析")
    
    # アイテム別集計
    item_sales = df.groupby(item_col)[amount_col].sum().reset_index()
    item_sales = item_sales.sort_values(amount_col, ascending=False).reset_index(drop=True)
    item_sales['順位'] = range(1, len(item_sales) + 1)
    
    # タイトル
    ws['B2'] = "ABC分析（パレート分析）"
    ws['B2'].font = Font(bold=True, size=18, color="2E5090")
    
    # ヘッダー
    headers = ['順位', item_col, '売上金額', '累積売上', '構成比', '累積構成比', 'ランク']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    # データ行
    total_sales = item_sales[amount_col].sum()
    for i, row in enumerate(item_sales.values, start=5):
        ws.cell(row=i, column=2, value=row[2])  # 順位
        ws.cell(row=i, column=3, value=row[0])  # アイテム名
        ws.cell(row=i, column=4, value=row[1])  # 売上
        ws.cell(row=i, column=4).number_format = '¥#,##0'
        
        # 累積売上（Excel関数）
        if i == 5:
            ws.cell(row=i, column=5, value=f'=D{i}')
        else:
            ws.cell(row=i, column=5, value=f'=E{i-1}+D{i}')
        ws.cell(row=i, column=5).number_format = '¥#,##0'
        
        # 構成比（Excel関数）
        ws.cell(row=i, column=6, value=f'=D{i}/SUM($D$5:$D${4+len(item_sales)})')
        ws.cell(row=i, column=6).number_format = '0.0%'
        
        # 累積構成比（Excel関数）
        ws.cell(row=i, column=7, value=f'=E{i}/SUM($D$5:$D${4+len(item_sales)})')
        ws.cell(row=i, column=7).number_format = '0.0%'
        
        # ABCランク（Excel関数）
        ws.cell(row=i, column=8, value=f'=IF(G{i}<=0.7,"A",IF(G{i}<=0.9,"B","C"))')
    
    # 条件付き書式（ABCランク色分け）
    ws.conditional_formatting.add(
        f'H5:H{4+len(item_sales)}',
        FormulaRule(formula=['$H5="A"'], fill=RANK_COLORS["A"])
    )
    ws.conditional_formatting.add(
        f'H5:H{4+len(item_sales)}',
        FormulaRule(formula=['$H5="B"'], fill=RANK_COLORS["B"])
    )
    ws.conditional_formatting.add(
        f'H5:H{4+len(item_sales)}',
        FormulaRule(formula=['$H5="C"'], fill=RANK_COLORS["C"])
    )
    
    # 売上金額にデータバー
    ws.conditional_formatting.add(
        f'D5:D{4+len(item_sales)}',
        DataBarRule(start_type='min', end_type='max',
                   color="5B9BD5", showValue=True, minLength=None, maxLength=None)
    )
    
    # パレート図（複合グラフ）
    # 棒グラフ（売上）
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "パレート図（ABC分析）"
    chart1.y_axis.title = "売上（円）"
    
    data1 = Reference(ws, min_col=4, min_row=4, max_row=4+len(item_sales))
    cats = Reference(ws, min_col=3, min_row=5, max_row=4+len(item_sales))
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    
    # 折れ線グラフ（累積構成比）
    chart2 = LineChart()
    chart2.y_axis.axId = 200
    chart2.y_axis.title = "累積構成比"
    
    data2 = Reference(ws, min_col=7, min_row=4, max_row=4+len(item_sales))
    chart2.add_data(data2, titles_from_data=True)
    chart2.style = 10
    
    # 累積構成比を右軸に
    chart2.y_axis.crosses = "max"
    chart1.y_axis.crosses = "min"
    
    # グラフ結合
    chart1 += chart2
    chart1.width = 20
    chart1.height = 12
    
    ws.add_chart(chart1, "J4")
    
    # 列幅調整
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    
    return ws


def create_staff_analysis_sheet(wb, df, amount_col, staff_col):
    """担当者別分析シート作成"""
    ws = wb.create_sheet("担当者別分析")
    
    # 担当者リスト（元データシート参照用）
    staff_list = df[staff_col].unique().tolist()
    
    # タイトル
    ws['B2'] = "担当者別売上分析"
    ws['B2'].font = Font(bold=True, size=18, color="2E5090")
    
    # ヘッダー
    headers = ['担当者', '売上合計', '件数', '平均単価', '構成比']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    # 元データシートを参照してSUMIF/COUNTIF
    data_sheet_name = "元データ"
    staff_col_letter = "D"  # 元データでの担当者列
    amount_col_letter = "H"  # 元データでの売上金額列
    
    for i, staff in enumerate(staff_list, start=5):
        ws.cell(row=i, column=2, value=staff)
        
        # 売上合計（SUMIF関数）
        ws.cell(row=i, column=3, value=f'=SUMIF({data_sheet_name}!$D:$D,B{i},{data_sheet_name}!$H:$H)')
        ws.cell(row=i, column=3).number_format = '¥#,##0'
        
        # 件数（COUNTIF関数）
        ws.cell(row=i, column=4, value=f'=COUNTIF({data_sheet_name}!$D:$D,B{i})')
        
        # 平均単価
        ws.cell(row=i, column=5, value=f'=IF(D{i}>0,C{i}/D{i},0)')
        ws.cell(row=i, column=5).number_format = '¥#,##0'
        
        # 構成比
        ws.cell(row=i, column=6, value=f'=C{i}/SUM($C$5:$C${4+len(staff_list)})')
        ws.cell(row=i, column=6).number_format = '0.0%'
    
    # 合計行
    total_row = 5 + len(staff_list)
    ws.cell(row=total_row, column=2, value="合計")
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=f'=SUM(C5:C{total_row-1})')
    ws.cell(row=total_row, column=3).number_format = '¥#,##0'
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=f'=SUM(D5:D{total_row-1})')
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    
    # 担当者別棒グラフ
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "担当者別売上"
    chart.y_axis.title = "売上（円）"
    
    data = Reference(ws, min_col=3, min_row=4, max_row=4+len(staff_list))
    cats = Reference(ws, min_col=2, min_row=5, max_row=4+len(staff_list))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10
    
    ws.add_chart(chart, "H4")
    
    # データバー
    ws.conditional_formatting.add(
        f'C5:C{4+len(staff_list)}',
        DataBarRule(start_type='min', end_type='max',
                   color="666666", showValue=True, minLength=None, maxLength=None)
    )
    
    # 列幅
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    
    return ws


def create_monthly_trend_sheet(wb, df, date_col, amount_col):
    """月次トレンドシート作成"""
    ws = wb.create_sheet("月次トレンド")
    
    # 月別集計
    monthly = df.groupby('年月')[amount_col].agg(['sum', 'count', 'mean']).reset_index()
    monthly.columns = ['年月', '売上合計', '件数', '平均単価']
    monthly = monthly.sort_values('年月')
    
    # タイトル
    ws['B2'] = "月次売上トレンド分析"
    ws['B2'].font = Font(bold=True, size=18, color="2E5090")
    
    # ヘッダー
    headers = ['年月', '売上合計', '件数', '平均単価', '前月比', '前月差額']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    # データ行
    for i, row in enumerate(monthly.values, start=5):
        ws.cell(row=i, column=2, value=row[0])  # 年月
        ws.cell(row=i, column=3, value=row[1])  # 売上合計
        ws.cell(row=i, column=3).number_format = '¥#,##0'
        ws.cell(row=i, column=4, value=row[2])  # 件数
        ws.cell(row=i, column=5, value=row[3])  # 平均単価
        ws.cell(row=i, column=5).number_format = '¥#,##0'
        
        # 前月比（Excel関数）
        if i == 5:
            ws.cell(row=i, column=6, value="-")
        else:
            ws.cell(row=i, column=6, value=f'=IF(C{i-1}>0,C{i}/C{i-1}-1,"-")')
            ws.cell(row=i, column=6).number_format = '+0.0%;-0.0%;0.0%'
        
        # 前月差額
        if i == 5:
            ws.cell(row=i, column=7, value="-")
        else:
            ws.cell(row=i, column=7, value=f'=C{i}-C{i-1}')
            ws.cell(row=i, column=7).number_format = '¥#,##0;[Red]-¥#,##0'
    
    # 条件付き書式（前月比の色分け）
    ws.conditional_formatting.add(
        f'F6:F{4+len(monthly)}',
        FormulaRule(formula=['$F6>0'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f'F6:F{4+len(monthly)}',
        FormulaRule(formula=['$F6<0'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    
    # トレンドグラフ
    chart = LineChart()
    chart.title = "月次売上推移"
    chart.style = 10
    chart.y_axis.title = "売上（円）"
    chart.x_axis.title = "年月"
    
    data = Reference(ws, min_col=3, min_row=4, max_row=4+len(monthly))
    cats = Reference(ws, min_col=2, min_row=5, max_row=4+len(monthly))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 10
    
    ws.add_chart(chart, "I4")
    
    # 列幅
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    
    return ws


def create_raw_data_sheet(wb, df):
    """元データシート作成"""
    ws = wb.create_sheet("元データ")
    
    # DataFrameをシートに書き込み
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
    
    # オートフィルター設定
    ws.auto_filter.ref = ws.dimensions
    
    # 列幅自動調整
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width
    
    return ws


def generate_sales_analysis_excel(input_file, output_file, date_col, amount_col, category_col, item_col, staff_col):
    """メイン処理"""
    # データ読み込み
    df = load_and_prepare_data(input_file, date_col, amount_col, category_col, item_col, staff_col)
    
    # ワークブック作成
    wb = Workbook()
    
    # デフォルトシート削除
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 元データシート（他シートから参照されるため先に作成）
    create_raw_data_sheet(wb, df)
    
    # 各分析シート作成
    create_dashboard_sheet(wb, df, date_col, amount_col, category_col, item_col, staff_col)
    create_abc_analysis_sheet(wb, df, amount_col, item_col)
    create_staff_analysis_sheet(wb, df, amount_col, staff_col)
    create_monthly_trend_sheet(wb, df, date_col, amount_col)
    
    # シート順序調整
    wb._sheets = [wb['ダッシュボード'], wb['ABC分析'], wb['担当者別分析'], wb['月次トレンド'], wb['元データ']]
    
    # 保存
    wb.save(output_file)
    print(f"Generated: {output_file}")
    
    return output_file


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="売上分析Excel生成")
    parser.add_argument("--input", required=True, help="入力CSVファイル")
    parser.add_argument("--output", required=True, help="出力Excelファイル")
    parser.add_argument("--date-col", default="日付", help="日付列名")
    parser.add_argument("--amount-col", default="売上金額", help="売上金額列名")
    parser.add_argument("--category-col", default="カテゴリ", help="カテゴリ列名")
    parser.add_argument("--item-col", default="サービス名", help="アイテム列名")
    parser.add_argument("--staff-col", default="担当者", help="担当者列名")
    
    args = parser.parse_args()
    
    generate_sales_analysis_excel(
        args.input,
        args.output,
        args.date_col,
        args.amount_col,
        args.category_col,
        args.item_col,
        args.staff_col
    )
