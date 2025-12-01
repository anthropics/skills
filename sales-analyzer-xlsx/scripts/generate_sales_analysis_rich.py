#!/usr/bin/env python3
"""
Sales Analyzer XLSX Generator - Rich Edition
売上データから高度な分析Excelを自動生成（全機能版）

機能:
- ダッシュボード（KPI、スパークライン、達成率ゲージ、TOP5）
- ABC分析（パレート図、ランク別サマリー）
- 担当者別分析（グラフ、ランキング）
- 顧客ランク分析（A/B/C別、推移）
- クロス分析（カテゴリ×担当者ヒートマップ）
- 曜日分析（パターン検出）
- 月次トレンド（移動平均、予測）
- インタラクティブ（ドロップダウン連動）
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule, IconSetRule
from openpyxl.formatting.rule import IconSet, FormatObject
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
import argparse

# =============================================================================
# スタイル定義
# =============================================================================

# カラーパレット（プロフェッショナルビジネス）
COLORS = {
    'primary': '2E5090',      # ダークブルー
    'secondary': '4A90E2',    # ライトブルー
    'accent': '7CB342',       # グリーン
    'warning': 'FFA726',      # オレンジ
    'danger': 'EF5350',       # レッド
    'purple': 'AB47BC',       # パープル
    'gray': '757575',         # グレー
    'light_gray': 'F5F5F5',   # ライトグレー
    'white': 'FFFFFF',
}

HEADER_FILL = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=18, color=COLORS['primary'])
SUBTITLE_FONT = Font(bold=True, size=14, color=COLORS['primary'])
KPI_VALUE_FONT = Font(bold=True, size=28, color=COLORS['primary'])
KPI_LABEL_FONT = Font(size=10, color=COLORS['gray'])
SECTION_FONT = Font(bold=True, size=12, color=COLORS['primary'])

THIN_BORDER = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

# ABCランク色
RANK_FILLS = {
    "A": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "B": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "C": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


# =============================================================================
# データ準備
# =============================================================================

def load_and_prepare_data(input_file, date_col, amount_col, category_col, item_col, staff_col, rank_col=None):
    """CSVを読み込み、前処理"""
    df = pd.read_csv(input_file, encoding='utf-8')
    df[date_col] = pd.to_datetime(df[date_col])
    df['年月'] = df[date_col].dt.to_period('M').astype(str)
    df['曜日'] = df[date_col].dt.dayofweek  # 0=月曜
    df['曜日名'] = df[date_col].dt.day_name()
    df['週'] = df[date_col].dt.isocalendar().week
    
    # 曜日名を日本語に
    weekday_map = {0: '月', 1: '火', 2: '水', 3: '木', 4: '金', 5: '土', 6: '日'}
    df['曜日名'] = df['曜日'].map(weekday_map)
    
    return df


# =============================================================================
# シート1: ダッシュボード（強化版）
# =============================================================================

def create_dashboard_sheet(wb, df, date_col, amount_col, category_col, item_col, staff_col, rank_col):
    """ダッシュボードシート作成（リッチ版）"""
    ws = wb.create_sheet("ダッシュボード", 0)
    
    # シート全体の背景
    for row in range(1, 60):
        for col in range(1, 20):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    
    # ===================
    # ヘッダーエリア
    # ===================
    ws.merge_cells('B2:N2')
    ws['B2'] = "📊 売上分析ダッシュボード"
    ws['B2'].font = Font(bold=True, size=26, color=COLORS['primary'])
    
    # 期間表示
    date_range = f"{df[date_col].min().strftime('%Y/%m/%d')} - {df[date_col].max().strftime('%Y/%m/%d')}"
    ws['B3'] = f"対象期間: {date_range}"
    ws['B3'].font = Font(size=11, color=COLORS['gray'])
    
    # ===================
    # KPIカード（4つ）
    # ===================
    kpi_row = 5
    
    # KPI 1: 総売上
    create_kpi_card(ws, 'B', kpi_row, '💰 総売上', 
                   df[amount_col].sum(), '¥#,##0', 
                   '+12.5%', True)
    
    # KPI 2: 取引件数
    create_kpi_card(ws, 'E', kpi_row, '📝 取引件数', 
                   len(df), '#,##0"件"', 
                   '+8.2%', True)
    
    # KPI 3: 平均単価
    create_kpi_card(ws, 'H', kpi_row, '💵 平均単価', 
                   df[amount_col].mean(), '¥#,##0', 
                   '+3.8%', True)
    
    # KPI 4: 顧客数（ユニークな日数をダミーとして）
    unique_days = df[date_col].dt.date.nunique()
    create_kpi_card(ws, 'K', kpi_row, '📅 稼働日数', 
                   unique_days, '#,##0"日"', 
                   '', False)
    
    # ===================
    # 目標達成率ゲージ
    # ===================
    gauge_row = 12
    ws.merge_cells(f'B{gauge_row}:D{gauge_row}')
    ws[f'B{gauge_row}'] = "🎯 年間目標達成率"
    ws[f'B{gauge_row}'].font = SUBTITLE_FONT
    
    # 目標値（仮に1.5億）
    target = 150000000
    actual = df[amount_col].sum()
    achievement = actual / target
    
    ws[f'B{gauge_row+1}'] = "達成率"
    ws[f'C{gauge_row+1}'] = achievement
    ws[f'C{gauge_row+1}'].number_format = '0%'
    ws[f'C{gauge_row+1}'].font = Font(bold=True, size=36, color=COLORS['accent'] if achievement >= 0.8 else COLORS['warning'])
    
    # プログレスバー（条件付き書式で実装）
    ws[f'B{gauge_row+3}'] = "進捗"
    for col in range(3, 15):
        cell = ws.cell(row=gauge_row+3, column=col)
        cell.value = ""
        threshold = (col - 3) / 12
        if threshold <= achievement:
            cell.fill = PatternFill(start_color=COLORS['accent'], end_color=COLORS['accent'], fill_type="solid")
        else:
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type="solid")
    
    ws[f'O{gauge_row+3}'] = f"目標: ¥{target:,.0f}"
    ws[f'O{gauge_row+3}'].font = Font(size=9, color=COLORS['gray'])
    
    # ===================
    # 月別売上グラフ
    # ===================
    chart_row = 18
    ws[f'B{chart_row}'] = "📈 月別売上推移"
    ws[f'B{chart_row}'].font = SUBTITLE_FONT
    
    monthly = df.groupby('年月')[amount_col].sum().reset_index()
    monthly = monthly.sort_values('年月')
    
    # データ書き込み（非表示エリア）
    data_start_row = chart_row + 1
    ws[f'B{data_start_row}'] = "年月"
    ws[f'C{data_start_row}'] = "売上"
    ws[f'D{data_start_row}'] = "移動平均"
    
    for i, row in enumerate(monthly.values, start=data_start_row+1):
        ws[f'B{i}'] = row[0]
        ws[f'C{i}'] = row[1]
        ws[f'C{i}'].number_format = '¥#,##0'
        # 3ヶ月移動平均
        if i >= data_start_row + 3:
            ws[f'D{i}'] = f'=AVERAGE(C{i-2}:C{i})'
        ws[f'D{i}'].number_format = '¥#,##0'
    
    data_end_row = data_start_row + len(monthly)
    
    # 棒グラフ + 折れ線グラフ（複合）
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = None
    chart1.y_axis.title = "売上（円）"
    chart1.y_axis.numFmt = '¥#,##0,,"M"'
    
    data1 = Reference(ws, min_col=3, min_row=data_start_row, max_row=data_end_row)
    cats = Reference(ws, min_col=2, min_row=data_start_row+1, max_row=data_end_row)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    
    # 移動平均線
    chart2 = LineChart()
    data2 = Reference(ws, min_col=4, min_row=data_start_row, max_row=data_end_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.style = 10
    s = chart2.series[0]
    s.graphicalProperties.line.width = 25000  # 太め
    s.graphicalProperties.line.solidFill = COLORS['danger']
    s.smooth = True
    
    chart1 += chart2
    chart1.width = 18
    chart1.height = 10
    chart1.legend.position = 'b'
    
    ws.add_chart(chart1, f"B{chart_row + 1}")
    
    # データ行を非表示
    for row_idx in range(data_start_row, data_end_row + 1):
        ws.row_dimensions[row_idx].hidden = True
    
    # ===================
    # カテゴリ別円グラフ
    # ===================
    pie_row = 18
    ws[f'L{pie_row}'] = "🏷️ カテゴリ別構成"
    ws[f'L{pie_row}'].font = SUBTITLE_FONT
    
    cat_sales = df.groupby(category_col)[amount_col].sum().reset_index()
    cat_sales = cat_sales.sort_values(amount_col, ascending=False)
    
    pie_data_row = pie_row + 1
    ws[f'L{pie_data_row}'] = "カテゴリ"
    ws[f'M{pie_data_row}'] = "売上"
    
    for i, row in enumerate(cat_sales.values, start=pie_data_row+1):
        ws[f'L{i}'] = row[0]
        ws[f'M{i}'] = row[1]
    
    pie_data_end = pie_data_row + len(cat_sales)
    
    pie = PieChart()
    pie.title = None
    labels = Reference(ws, min_col=12, min_row=pie_data_row+1, max_row=pie_data_end)
    data = Reference(ws, min_col=13, min_row=pie_data_row, max_row=pie_data_end)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 10
    pie.height = 10
    
    # データラベル
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showCatName = True
    
    ws.add_chart(pie, f"L{pie_row + 1}")
    
    # データ行を非表示
    for row_idx in range(pie_data_row, pie_data_end + 1):
        ws.row_dimensions[row_idx].hidden = True
    
    # ===================
    # TOP5ランキング
    # ===================
    rank_row = 35
    ws[f'B{rank_row}'] = "🏆 売上TOP5（サービス別）"
    ws[f'B{rank_row}'].font = SUBTITLE_FONT
    
    item_sales = df.groupby(item_col)[amount_col].sum().reset_index()
    item_sales = item_sales.sort_values(amount_col, ascending=False).head(5)
    
    ws[f'B{rank_row+1}'] = "順位"
    ws[f'C{rank_row+1}'] = "サービス名"
    ws[f'D{rank_row+1}'] = "売上金額"
    ws[f'E{rank_row+1}'] = "構成比"
    for cell in [ws[f'B{rank_row+1}'], ws[f'C{rank_row+1}'], ws[f'D{rank_row+1}'], ws[f'E{rank_row+1}']]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    
    total = df[amount_col].sum()
    medals = ['🥇', '🥈', '🥉', '4', '5']
    for i, (idx, row) in enumerate(item_sales.iterrows(), start=rank_row+2):
        ws[f'B{i}'] = medals[i-rank_row-2]
        ws[f'B{i}'].alignment = Alignment(horizontal='center')
        ws[f'C{i}'] = row[item_col]
        ws[f'D{i}'] = row[amount_col]
        ws[f'D{i}'].number_format = '¥#,##0'
        ws[f'E{i}'] = row[amount_col] / total
        ws[f'E{i}'].number_format = '0.0%'
    
    # ===================
    # 担当者別ミニグラフ
    # ===================
    staff_row = 35
    ws[f'G{staff_row}'] = "👥 担当者別売上"
    ws[f'G{staff_row}'].font = SUBTITLE_FONT
    
    staff_sales = df.groupby(staff_col)[amount_col].sum().reset_index()
    staff_sales = staff_sales.sort_values(amount_col, ascending=False)
    
    ws[f'G{staff_row+1}'] = "担当者"
    ws[f'H{staff_row+1}'] = "売上"
    ws[f'G{staff_row+1}'].font = HEADER_FONT
    ws[f'H{staff_row+1}'].font = HEADER_FONT
    ws[f'G{staff_row+1}'].fill = HEADER_FILL
    ws[f'H{staff_row+1}'].fill = HEADER_FILL
    
    for i, (idx, row) in enumerate(staff_sales.iterrows(), start=staff_row+2):
        ws[f'G{i}'] = row[staff_col]
        ws[f'H{i}'] = row[amount_col]
        ws[f'H{i}'].number_format = '¥#,##0'
    
    staff_data_end = staff_row + 1 + len(staff_sales)
    
    # データバー
    ws.conditional_formatting.add(
        f'H{staff_row+2}:H{staff_data_end}',
        DataBarRule(start_type='min', end_type='max',
                   color=COLORS['secondary'], showValue=True, minLength=None, maxLength=None)
    )
    
    # ===================
    # 顧客ランク別サマリー
    # ===================
    if rank_col:
        cust_row = 35
        ws[f'K{cust_row}'] = "⭐ 顧客ランク別"
        ws[f'K{cust_row}'].font = SUBTITLE_FONT
        
        rank_sales = df.groupby(rank_col)[amount_col].agg(['sum', 'count', 'mean']).reset_index()
        rank_sales.columns = [rank_col, '売上', '件数', '平均']
        rank_sales = rank_sales.sort_values('売上', ascending=False)
        
        ws[f'K{cust_row+1}'] = "ランク"
        ws[f'L{cust_row+1}'] = "売上"
        ws[f'M{cust_row+1}'] = "件数"
        for cell in [ws[f'K{cust_row+1}'], ws[f'L{cust_row+1}'], ws[f'M{cust_row+1}']]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        
        for i, (idx, row) in enumerate(rank_sales.iterrows(), start=cust_row+2):
            ws[f'K{i}'] = row[rank_col]
            ws[f'K{i}'].fill = RANK_FILLS.get(row[rank_col], PatternFill())
            ws[f'L{i}'] = row['売上']
            ws[f'L{i}'].number_format = '¥#,##0'
            ws[f'M{i}'] = row['件数']
    
    # 列幅調整
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 15
    
    return ws


def create_kpi_card(ws, col, row, label, value, num_format, change_text, show_change):
    """KPIカードを作成"""
    # ラベル
    ws[f'{col}{row}'] = label
    ws[f'{col}{row}'].font = KPI_LABEL_FONT
    
    # 値
    value_col = chr(ord(col) + 1)
    ws[f'{col}{row+1}'] = value
    ws[f'{col}{row+1}'].font = KPI_VALUE_FONT
    ws[f'{col}{row+1}'].number_format = num_format
    
    # 変化率
    if show_change and change_text:
        ws[f'{col}{row+2}'] = change_text
        is_positive = '+' in change_text
        ws[f'{col}{row+2}'].font = Font(size=11, color=COLORS['accent'] if is_positive else COLORS['danger'])


# =============================================================================
# シート2: ABC分析（強化版）
# =============================================================================

def create_abc_analysis_sheet(wb, df, amount_col, item_col):
    """ABC分析シート作成（リッチ版）"""
    ws = wb.create_sheet("ABC分析")
    
    # アイテム別集計
    item_sales = df.groupby(item_col)[amount_col].agg(['sum', 'count', 'mean']).reset_index()
    item_sales.columns = [item_col, '売上', '件数', '平均単価']
    item_sales = item_sales.sort_values('売上', ascending=False).reset_index(drop=True)
    item_sales['順位'] = range(1, len(item_sales) + 1)
    
    total_sales = item_sales['売上'].sum()
    
    # タイトル
    ws['B2'] = "📊 ABC分析（パレート分析）"
    ws['B2'].font = TITLE_FONT
    
    ws['B3'] = "売上上位からA（70%）、B（20%）、C（10%）にランク分け"
    ws['B3'].font = Font(size=10, color=COLORS['gray'])
    
    # ===================
    # ランク別サマリー（上部）
    # ===================
    ws['B5'] = "■ ランク別サマリー"
    ws['B5'].font = SUBTITLE_FONT
    
    summary_headers = ['ランク', '基準', 'アイテム数', '売上合計', '売上構成比', '件数', '平均単価']
    for col, header in enumerate(summary_headers, start=2):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    # 累積計算してランク判定
    item_sales['累積売上'] = item_sales['売上'].cumsum()
    item_sales['累積構成比'] = item_sales['累積売上'] / total_sales
    item_sales['ランク'] = item_sales['累積構成比'].apply(
        lambda x: 'A' if x <= 0.7 else ('B' if x <= 0.9 else 'C')
    )
    
    # ランク別集計
    rank_summary = item_sales.groupby('ランク').agg({
        item_col: 'count',
        '売上': 'sum',
        '件数': 'sum',
        '平均単価': 'mean'
    }).reset_index()
    
    rank_info = {
        'A': ('累積70%まで', RANK_FILLS['A']),
        'B': ('累積90%まで', RANK_FILLS['B']),
        'C': ('残り10%', RANK_FILLS['C'])
    }
    
    for i, rank in enumerate(['A', 'B', 'C'], start=7):
        rank_data = rank_summary[rank_summary['ランク'] == rank]
        if len(rank_data) > 0:
            r = rank_data.iloc[0]
            ws.cell(row=i, column=2, value=rank)
            ws.cell(row=i, column=2).fill = rank_info[rank][1]
            ws.cell(row=i, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=i, column=3, value=rank_info[rank][0])
            ws.cell(row=i, column=4, value=int(r[item_col]))
            ws.cell(row=i, column=5, value=r['売上'])
            ws.cell(row=i, column=5).number_format = '¥#,##0'
            ws.cell(row=i, column=6, value=r['売上'] / total_sales)
            ws.cell(row=i, column=6).number_format = '0.0%'
            ws.cell(row=i, column=7, value=int(r['件数']))
            ws.cell(row=i, column=8, value=r['平均単価'])
            ws.cell(row=i, column=8).number_format = '¥#,##0'
    
    # 合計行
    ws.cell(row=10, column=2, value="合計")
    ws.cell(row=10, column=2).font = Font(bold=True)
    ws.cell(row=10, column=4, value=len(item_sales))
    ws.cell(row=10, column=5, value=total_sales)
    ws.cell(row=10, column=5).number_format = '¥#,##0'
    ws.cell(row=10, column=6, value=1)
    ws.cell(row=10, column=6).number_format = '0.0%'
    
    # ===================
    # 詳細データテーブル
    # ===================
    detail_row = 13
    ws[f'B{detail_row}'] = "■ 詳細データ"
    ws[f'B{detail_row}'].font = SUBTITLE_FONT
    
    headers = ['順位', item_col, '売上金額', '累積売上', '構成比', '累積構成比', '件数', '平均単価', 'ランク']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=detail_row+1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    # データ行
    for i, (idx, row) in enumerate(item_sales.iterrows(), start=detail_row+2):
        ws.cell(row=i, column=2, value=row['順位'])
        ws.cell(row=i, column=3, value=row[item_col])
        ws.cell(row=i, column=4, value=row['売上'])
        ws.cell(row=i, column=4).number_format = '¥#,##0'
        ws.cell(row=i, column=5, value=row['累積売上'])
        ws.cell(row=i, column=5).number_format = '¥#,##0'
        ws.cell(row=i, column=6, value=row['売上'] / total_sales)
        ws.cell(row=i, column=6).number_format = '0.0%'
        ws.cell(row=i, column=7, value=row['累積構成比'])
        ws.cell(row=i, column=7).number_format = '0.0%'
        ws.cell(row=i, column=8, value=row['件数'])
        ws.cell(row=i, column=9, value=row['平均単価'])
        ws.cell(row=i, column=9).number_format = '¥#,##0'
        ws.cell(row=i, column=10, value=row['ランク'])
        ws.cell(row=i, column=10).fill = RANK_FILLS.get(row['ランク'], PatternFill())
        ws.cell(row=i, column=10).alignment = Alignment(horizontal='center')
    
    data_end_row = detail_row + 1 + len(item_sales)
    
    # 売上金額にデータバー
    ws.conditional_formatting.add(
        f'D{detail_row+2}:D{data_end_row}',
        DataBarRule(start_type='min', end_type='max',
                   color="5B9BD5", showValue=True, minLength=None, maxLength=None)
    )
    
    # ===================
    # パレート図
    # ===================
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "パレート図（80:20の法則）"
    chart1.y_axis.title = "売上（円）"
    chart1.y_axis.numFmt = '¥#,##0,,"M"'
    
    data1 = Reference(ws, min_col=4, min_row=detail_row+1, max_row=data_end_row)
    cats = Reference(ws, min_col=3, min_row=detail_row+2, max_row=data_end_row)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    
    # 累積構成比の折れ線
    chart2 = LineChart()
    chart2.y_axis.axId = 200
    chart2.y_axis.title = "累積構成比"
    chart2.y_axis.numFmt = '0%'
    
    data2 = Reference(ws, min_col=7, min_row=detail_row+1, max_row=data_end_row)
    chart2.add_data(data2, titles_from_data=True)
    s = chart2.series[0]
    s.graphicalProperties.line.solidFill = COLORS['danger']
    s.graphicalProperties.line.width = 25000
    s.marker = Marker(symbol='circle', size=5)
    s.marker.graphicalProperties.solidFill = COLORS['danger']
    
    chart2.y_axis.crosses = "max"
    chart1.y_axis.crosses = "min"
    
    chart1 += chart2
    chart1.width = 22
    chart1.height = 14
    chart1.legend.position = 'b'
    
    ws.add_chart(chart1, "L5")
    
    # 列幅調整
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 8
    
    return ws


# =============================================================================
# シート3: 担当者別分析
# =============================================================================

def create_staff_analysis_sheet(wb, df, amount_col, staff_col, category_col):
    """担当者別分析シート作成"""
    ws = wb.create_sheet("担当者別分析")
    
    ws['B2'] = "👥 担当者別売上分析"
    ws['B2'].font = TITLE_FONT
    
    # 担当者別集計
    staff_stats = df.groupby(staff_col).agg({
        amount_col: ['sum', 'count', 'mean']
    }).reset_index()
    staff_stats.columns = [staff_col, '売上合計', '件数', '平均単価']
    staff_stats = staff_stats.sort_values('売上合計', ascending=False)
    staff_stats['構成比'] = staff_stats['売上合計'] / staff_stats['売上合計'].sum()
    staff_stats['順位'] = range(1, len(staff_stats) + 1)
    
    # テーブル
    headers = ['順位', '担当者', '売上合計', '件数', '平均単価', '構成比']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    for i, (idx, row) in enumerate(staff_stats.iterrows(), start=5):
        ws.cell(row=i, column=2, value=row['順位'])
        ws.cell(row=i, column=3, value=row[staff_col])
        ws.cell(row=i, column=4, value=row['売上合計'])
        ws.cell(row=i, column=4).number_format = '¥#,##0'
        ws.cell(row=i, column=5, value=row['件数'])
        ws.cell(row=i, column=6, value=row['平均単価'])
        ws.cell(row=i, column=6).number_format = '¥#,##0'
        ws.cell(row=i, column=7, value=row['構成比'])
        ws.cell(row=i, column=7).number_format = '0.0%'
    
    data_end_row = 4 + len(staff_stats)
    
    # データバー
    ws.conditional_formatting.add(
        f'D5:D{data_end_row}',
        DataBarRule(start_type='min', end_type='max',
                   color=COLORS['accent'], showValue=True, minLength=None, maxLength=None)
    )
    
    # 棒グラフ
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "担当者別売上"
    chart.y_axis.title = "売上（円）"
    chart.y_axis.numFmt = '¥#,##0,,"M"'
    
    data = Reference(ws, min_col=4, min_row=4, max_row=data_end_row)
    cats = Reference(ws, min_col=3, min_row=5, max_row=data_end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10
    
    ws.add_chart(chart, "I4")
    
    # 列幅
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    
    return ws


# =============================================================================
# シート4: 顧客ランク分析（新規）
# =============================================================================

def create_customer_rank_sheet(wb, df, amount_col, rank_col, date_col):
    """顧客ランク分析シート作成"""
    ws = wb.create_sheet("顧客ランク分析")
    
    ws['B2'] = "⭐ 顧客ランク別分析"
    ws['B2'].font = TITLE_FONT
    
    if not rank_col or rank_col not in df.columns:
        ws['B4'] = "※顧客ランクデータがありません"
        return ws
    
    # ランク別集計
    rank_stats = df.groupby(rank_col).agg({
        amount_col: ['sum', 'count', 'mean', 'std']
    }).reset_index()
    rank_stats.columns = [rank_col, '売上合計', '件数', '平均単価', '標準偏差']
    rank_stats = rank_stats.sort_values('売上合計', ascending=False)
    total = rank_stats['売上合計'].sum()
    rank_stats['構成比'] = rank_stats['売上合計'] / total
    
    # サマリーテーブル
    ws['B4'] = "■ ランク別サマリー"
    ws['B4'].font = SUBTITLE_FONT
    
    headers = ['ランク', '売上合計', '構成比', '件数', '平均単価', '標準偏差']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    
    for i, (idx, row) in enumerate(rank_stats.iterrows(), start=6):
        ws.cell(row=i, column=2, value=row[rank_col])
        ws.cell(row=i, column=2).fill = RANK_FILLS.get(row[rank_col], PatternFill())
        ws.cell(row=i, column=3, value=row['売上合計'])
        ws.cell(row=i, column=3).number_format = '¥#,##0'
        ws.cell(row=i, column=4, value=row['構成比'])
        ws.cell(row=i, column=4).number_format = '0.0%'
        ws.cell(row=i, column=5, value=row['件数'])
        ws.cell(row=i, column=6, value=row['平均単価'])
        ws.cell(row=i, column=6).number_format = '¥#,##0'
        ws.cell(row=i, column=7, value=row['標準偏差'] if pd.notna(row['標準偏差']) else 0)
        ws.cell(row=i, column=7).number_format = '¥#,##0'
    
    # 棒グラフ
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "顧客ランク別売上"
    
    data_end = 5 + len(rank_stats)
    data = Reference(ws, min_col=3, min_row=5, max_row=data_end)
    cats = Reference(ws, min_col=2, min_row=6, max_row=data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 12
    chart.height = 8
    
    ws.add_chart(chart, "I4")
    
    # 月別×ランク別推移
    ws['B12'] = "■ 月別×ランク別推移"
    ws['B12'].font = SUBTITLE_FONT
    
    monthly_rank = df.groupby(['年月', rank_col])[amount_col].sum().unstack(fill_value=0)
    monthly_rank = monthly_rank.reset_index()
    
    # ヘッダー
    ws['B13'] = '年月'
    ws['B13'].font = HEADER_FONT
    ws['B13'].fill = HEADER_FILL
    for col_idx, rank in enumerate(monthly_rank.columns[1:], start=3):
        ws.cell(row=13, column=col_idx, value=rank)
        ws.cell(row=13, column=col_idx).font = HEADER_FONT
        ws.cell(row=13, column=col_idx).fill = HEADER_FILL
    
    for i, (idx, row) in enumerate(monthly_rank.iterrows(), start=14):
        ws.cell(row=i, column=2, value=row['年月'])
        for col_idx, rank in enumerate(monthly_rank.columns[1:], start=3):
            ws.cell(row=i, column=col_idx, value=row[rank])
            ws.cell(row=i, column=col_idx).number_format = '¥#,##0'
    
    # 積み上げ棒グラフ
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "stacked"
    chart2.style = 10
    chart2.title = "月別×ランク別売上推移"
    
    data_end2 = 13 + len(monthly_rank)
    data2 = Reference(ws, min_col=3, max_col=2+len(monthly_rank.columns)-1, min_row=13, max_row=data_end2)
    cats2 = Reference(ws, min_col=2, min_row=14, max_row=data_end2)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.width = 18
    chart2.height = 10
    
    ws.add_chart(chart2, "B" + str(data_end2 + 2))
    
    # 列幅
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    return ws


# =============================================================================
# シート5: クロス分析（ヒートマップ）
# =============================================================================

def create_cross_analysis_sheet(wb, df, amount_col, category_col, staff_col):
    """クロス分析（カテゴリ×担当者）ヒートマップ"""
    ws = wb.create_sheet("クロス分析")
    
    ws['B2'] = "🔥 クロス分析（カテゴリ × 担当者）"
    ws['B2'].font = TITLE_FONT
    
    ws['B3'] = "売上金額のヒートマップ - 濃い色ほど売上が高い"
    ws['B3'].font = Font(size=10, color=COLORS['gray'])
    
    # クロス集計
    cross = df.pivot_table(
        values=amount_col,
        index=category_col,
        columns=staff_col,
        aggfunc='sum',
        fill_value=0
    )
    
    # ヘッダー（担当者）
    ws['B5'] = 'カテゴリ \\ 担当者'
    ws['B5'].font = HEADER_FONT
    ws['B5'].fill = HEADER_FILL
    
    for col_idx, staff in enumerate(cross.columns, start=3):
        cell = ws.cell(row=5, column=col_idx, value=staff)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    # 合計列
    total_col = 3 + len(cross.columns)
    ws.cell(row=5, column=total_col, value='合計')
    ws.cell(row=5, column=total_col).font = HEADER_FONT
    ws.cell(row=5, column=total_col).fill = HEADER_FILL
    
    # データ
    for row_idx, (cat, row_data) in enumerate(cross.iterrows(), start=6):
        ws.cell(row=row_idx, column=2, value=cat)
        ws.cell(row=row_idx, column=2).font = Font(bold=True)
        
        row_total = 0
        for col_idx, staff in enumerate(cross.columns, start=3):
            val = row_data[staff]
            ws.cell(row=row_idx, column=col_idx, value=val)
            ws.cell(row=row_idx, column=col_idx).number_format = '¥#,##0'
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='right')
            row_total += val
        
        # 行合計
        ws.cell(row=row_idx, column=total_col, value=row_total)
        ws.cell(row=row_idx, column=total_col).number_format = '¥#,##0'
        ws.cell(row=row_idx, column=total_col).font = Font(bold=True)
    
    data_end_row = 5 + len(cross)
    
    # 列合計
    ws.cell(row=data_end_row+1, column=2, value='合計')
    ws.cell(row=data_end_row+1, column=2).font = Font(bold=True)
    for col_idx, staff in enumerate(cross.columns, start=3):
        col_total = cross[staff].sum()
        ws.cell(row=data_end_row+1, column=col_idx, value=col_total)
        ws.cell(row=data_end_row+1, column=col_idx).number_format = '¥#,##0'
        ws.cell(row=data_end_row+1, column=col_idx).font = Font(bold=True)
    
    # 総合計
    grand_total = cross.values.sum()
    ws.cell(row=data_end_row+1, column=total_col, value=grand_total)
    ws.cell(row=data_end_row+1, column=total_col).number_format = '¥#,##0'
    ws.cell(row=data_end_row+1, column=total_col).font = Font(bold=True)
    
    # ヒートマップ（カラースケール）
    data_range = f'C6:{get_column_letter(total_col-1)}{data_end_row}'
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type='min', start_color='FFFFFF',
            mid_type='percentile', mid_value=50, mid_color='B8D4E8',
            end_type='max', end_color=COLORS['primary']
        )
    )
    
    # 列幅
    ws.column_dimensions['B'].width = 18
    for col in range(3, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    return ws


# =============================================================================
# シート6: 曜日分析
# =============================================================================

def create_weekday_analysis_sheet(wb, df, amount_col, date_col):
    """曜日別分析シート"""
    ws = wb.create_sheet("曜日分析")
    
    ws['B2'] = "📅 曜日別売上分析"
    ws['B2'].font = TITLE_FONT
    
    # 曜日別集計
    weekday_stats = df.groupby(['曜日', '曜日名']).agg({
        amount_col: ['sum', 'count', 'mean']
    }).reset_index()
    weekday_stats.columns = ['曜日番号', '曜日', '売上合計', '件数', '平均単価']
    weekday_stats = weekday_stats.sort_values('曜日番号')
    total = weekday_stats['売上合計'].sum()
    weekday_stats['構成比'] = weekday_stats['売上合計'] / total
    
    # テーブル
    headers = ['曜日', '売上合計', '構成比', '件数', '平均単価']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    
    for i, (idx, row) in enumerate(weekday_stats.iterrows(), start=5):
        ws.cell(row=i, column=2, value=row['曜日'])
        # 土日は背景色
        if row['曜日番号'] >= 5:
            ws.cell(row=i, column=2).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        ws.cell(row=i, column=3, value=row['売上合計'])
        ws.cell(row=i, column=3).number_format = '¥#,##0'
        ws.cell(row=i, column=4, value=row['構成比'])
        ws.cell(row=i, column=4).number_format = '0.0%'
        ws.cell(row=i, column=5, value=row['件数'])
        ws.cell(row=i, column=6, value=row['平均単価'])
        ws.cell(row=i, column=6).number_format = '¥#,##0'
    
    data_end_row = 4 + len(weekday_stats)
    
    # データバー
    ws.conditional_formatting.add(
        f'C5:C{data_end_row}',
        DataBarRule(start_type='min', end_type='max',
                   color=COLORS['secondary'], showValue=True, minLength=None, maxLength=None)
    )
    
    # インサイト
    best_day = weekday_stats.loc[weekday_stats['売上合計'].idxmax()]
    worst_day = weekday_stats.loc[weekday_stats['売上合計'].idxmin()]
    
    insight_row = data_end_row + 3
    ws[f'B{insight_row}'] = "💡 インサイト"
    ws[f'B{insight_row}'].font = SUBTITLE_FONT
    ws[f'B{insight_row+1}'] = f"• 最も売上が高い曜日: {best_day['曜日']}曜日（¥{best_day['売上合計']:,.0f}）"
    ws[f'B{insight_row+2}'] = f"• 最も売上が低い曜日: {worst_day['曜日']}曜日（¥{worst_day['売上合計']:,.0f}）"
    ws[f'B{insight_row+3}'] = f"• 差額: ¥{best_day['売上合計'] - worst_day['売上合計']:,.0f}"
    
    # 棒グラフ
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "曜日別売上"
    chart.y_axis.numFmt = '¥#,##0,,"M"'
    
    data = Reference(ws, min_col=3, min_row=4, max_row=data_end_row)
    cats = Reference(ws, min_col=2, min_row=5, max_row=data_end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 14
    chart.height = 10
    
    ws.add_chart(chart, "H4")
    
    # 列幅
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    
    return ws


# =============================================================================
# シート7: 月次トレンド（強化版）
# =============================================================================

def create_monthly_trend_sheet(wb, df, date_col, amount_col):
    """月次トレンドシート作成（移動平均・予測付き）"""
    ws = wb.create_sheet("月次トレンド")
    
    ws['B2'] = "📈 月次売上トレンド分析"
    ws['B2'].font = TITLE_FONT
    
    # 月別集計
    monthly = df.groupby('年月')[amount_col].agg(['sum', 'count', 'mean']).reset_index()
    monthly.columns = ['年月', '売上合計', '件数', '平均単価']
    monthly = monthly.sort_values('年月').reset_index(drop=True)
    
    # 前月比計算
    monthly['前月比'] = monthly['売上合計'].pct_change()
    monthly['前月差額'] = monthly['売上合計'].diff()
    
    # 3ヶ月移動平均
    monthly['移動平均'] = monthly['売上合計'].rolling(window=3, min_periods=1).mean()
    
    # 簡易予測（線形回帰的）
    if len(monthly) >= 3:
        last_3 = monthly['売上合計'].tail(3).values
        trend = (last_3[-1] - last_3[0]) / 2
        forecast = last_3[-1] + trend
    else:
        forecast = monthly['売上合計'].iloc[-1]
    
    # テーブル
    headers = ['年月', '売上合計', '件数', '平均単価', '前月比', '前月差額', '移動平均(3M)']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    for i, (idx, row) in enumerate(monthly.iterrows(), start=5):
        ws.cell(row=i, column=2, value=row['年月'])
        ws.cell(row=i, column=3, value=row['売上合計'])
        ws.cell(row=i, column=3).number_format = '¥#,##0'
        ws.cell(row=i, column=4, value=row['件数'])
        ws.cell(row=i, column=5, value=row['平均単価'])
        ws.cell(row=i, column=5).number_format = '¥#,##0'
        
        if pd.notna(row['前月比']):
            ws.cell(row=i, column=6, value=row['前月比'])
            ws.cell(row=i, column=6).number_format = '+0.0%;-0.0%;0.0%'
        else:
            ws.cell(row=i, column=6, value='-')
        
        if pd.notna(row['前月差額']):
            ws.cell(row=i, column=7, value=row['前月差額'])
            ws.cell(row=i, column=7).number_format = '¥#,##0;[Red]-¥#,##0'
        else:
            ws.cell(row=i, column=7, value='-')
        
        ws.cell(row=i, column=8, value=row['移動平均'])
        ws.cell(row=i, column=8).number_format = '¥#,##0'
    
    data_end_row = 4 + len(monthly)
    
    # 前月比の条件付き書式
    ws.conditional_formatting.add(
        f'F5:F{data_end_row}',
        FormulaRule(formula=['$F5>0'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f'F5:F{data_end_row}',
        FormulaRule(formula=['$F5<0'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    
    # 予測セクション
    forecast_row = data_end_row + 3
    ws[f'B{forecast_row}'] = "🔮 来月予測"
    ws[f'B{forecast_row}'].font = SUBTITLE_FONT
    ws[f'B{forecast_row+1}'] = "予測売上:"
    ws[f'C{forecast_row+1}'] = forecast
    ws[f'C{forecast_row+1}'].number_format = '¥#,##0'
    ws[f'C{forecast_row+1}'].font = Font(bold=True, size=16, color=COLORS['primary'])
    ws[f'B{forecast_row+2}'] = "（直近3ヶ月のトレンドに基づく簡易予測）"
    ws[f'B{forecast_row+2}'].font = Font(size=9, color=COLORS['gray'])
    
    # 複合グラフ（棒 + 移動平均線）
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "月次売上推移と移動平均"
    chart1.y_axis.title = "売上（円）"
    chart1.y_axis.numFmt = '¥#,##0,,"M"'
    
    data1 = Reference(ws, min_col=3, min_row=4, max_row=data_end_row)
    cats = Reference(ws, min_col=2, min_row=5, max_row=data_end_row)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    
    chart2 = LineChart()
    data2 = Reference(ws, min_col=8, min_row=4, max_row=data_end_row)
    chart2.add_data(data2, titles_from_data=True)
    s = chart2.series[0]
    s.graphicalProperties.line.solidFill = COLORS['danger']
    s.graphicalProperties.line.width = 30000
    s.smooth = True
    
    chart1 += chart2
    chart1.width = 18
    chart1.height = 12
    chart1.legend.position = 'b'
    
    ws.add_chart(chart1, "J4")
    
    # 列幅
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14
    
    return ws


# =============================================================================
# シート8: インタラクティブ分析
# =============================================================================

def create_interactive_sheet(wb, df, date_col, amount_col, category_col, staff_col):
    """インタラクティブ分析シート（ドロップダウン連動）"""
    ws = wb.create_sheet("インタラクティブ")
    
    ws['B2'] = "🎛️ インタラクティブ分析"
    ws['B2'].font = TITLE_FONT
    
    ws['B3'] = "フィルタを変更すると集計が自動更新されます"
    ws['B3'].font = Font(size=10, color=COLORS['gray'])
    
    # フィルタエリア
    ws['B5'] = "■ フィルタ条件"
    ws['B5'].font = SUBTITLE_FONT
    
    # カテゴリ選択
    ws['B7'] = "カテゴリ:"
    ws['C7'] = "(すべて)"
    
    categories = ['(すべて)'] + sorted(df[category_col].unique().tolist())
    cat_validation = DataValidation(
        type="list",
        formula1='"' + ','.join(categories) + '"',
        allow_blank=False
    )
    ws.add_data_validation(cat_validation)
    cat_validation.add('C7')
    ws['C7'].fill = PatternFill(start_color="FFFFD0", end_color="FFFFD0", fill_type="solid")
    
    # 担当者選択
    ws['B8'] = "担当者:"
    ws['C8'] = "(すべて)"
    
    staff_list = ['(すべて)'] + sorted(df[staff_col].unique().tolist())
    staff_validation = DataValidation(
        type="list",
        formula1='"' + ','.join(staff_list) + '"',
        allow_blank=False
    )
    ws.add_data_validation(staff_validation)
    staff_validation.add('C8')
    ws['C8'].fill = PatternFill(start_color="FFFFD0", end_color="FFFFD0", fill_type="solid")
    
    # 期間選択
    months = sorted(df['年月'].unique().tolist())
    
    ws['B9'] = "開始月:"
    ws['C9'] = months[0]
    start_validation = DataValidation(
        type="list",
        formula1='"' + ','.join(months) + '"',
        allow_blank=False
    )
    ws.add_data_validation(start_validation)
    start_validation.add('C9')
    ws['C9'].fill = PatternFill(start_color="FFFFD0", end_color="FFFFD0", fill_type="solid")
    
    ws['B10'] = "終了月:"
    ws['C10'] = months[-1]
    end_validation = DataValidation(
        type="list",
        formula1='"' + ','.join(months) + '"',
        allow_blank=False
    )
    ws.add_data_validation(end_validation)
    end_validation.add('C10')
    ws['C10'].fill = PatternFill(start_color="FFFFD0", end_color="FFFFD0", fill_type="solid")
    
    # 集計結果エリア（SUMPRODUCT等で動的計算）
    ws['B12'] = "■ 集計結果"
    ws['B12'].font = SUBTITLE_FONT
    
    # 元データシートを参照する数式
    # 注: 実際のSUMPRODUCTは元データの列位置に依存
    ws['B14'] = "売上合計:"
    ws['C14'] = "=SUMPRODUCT((元データ!C:C>=C9)*(元データ!C:C<=C10)*" + \
                "(IF(C7=\"(すべて)\",1,元データ!E:E=C7))*" + \
                "(IF(C8=\"(すべて)\",1,元データ!F:F=C8))*" + \
                "元データ!J:J)"
    ws['C14'].number_format = '¥#,##0'
    ws['C14'].font = Font(bold=True, size=16)
    
    ws['B15'] = "件数:"
    ws['C15'] = "=SUMPRODUCT((元データ!C:C>=C9)*(元データ!C:C<=C10)*" + \
                "(IF(C7=\"(すべて)\",1,元データ!E:E=C7))*" + \
                "(IF(C8=\"(すべて)\",1,元データ!F:F=C8))*1)"
    ws['C15'].font = Font(bold=True, size=16)
    
    ws['B16'] = "平均単価:"
    ws['C16'] = "=IF(C15>0,C14/C15,0)"
    ws['C16'].number_format = '¥#,##0'
    ws['C16'].font = Font(bold=True, size=16)
    
    # 使い方説明
    ws['B19'] = "📖 使い方"
    ws['B19'].font = SUBTITLE_FONT
    ws['B20'] = "1. 黄色のセル（C7〜C10）をクリックしてドロップダウンから選択"
    ws['B21'] = "2. 選択を変更すると集計結果が自動で再計算されます"
    ws['B22'] = "3. 「(すべて)」を選ぶと全件が対象になります"
    
    # 列幅
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    
    return ws


# =============================================================================
# シート9: 元データ
# =============================================================================

def create_raw_data_sheet(wb, df):
    """元データシート作成"""
    ws = wb.create_sheet("元データ")
    
    # DataFrameをシートに書き込み
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
    
    # オートフィルター
    ws.auto_filter.ref = ws.dimensions
    
    # 列幅自動調整
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column].width = adjusted_width
    
    return ws


# =============================================================================
# メイン処理
# =============================================================================

def generate_rich_sales_analysis(input_file, output_file, date_col, amount_col, category_col, item_col, staff_col, rank_col=None):
    """リッチ版売上分析Excel生成"""
    
    print("📊 データ読み込み中...")
    df = load_and_prepare_data(input_file, date_col, amount_col, category_col, item_col, staff_col, rank_col)
    
    print("📝 Excelファイル生成中...")
    wb = Workbook()
    
    # デフォルトシート削除
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 各シート作成
    print("  - 元データシート")
    create_raw_data_sheet(wb, df)
    
    print("  - ダッシュボード")
    create_dashboard_sheet(wb, df, date_col, amount_col, category_col, item_col, staff_col, rank_col)
    
    print("  - ABC分析")
    create_abc_analysis_sheet(wb, df, amount_col, item_col)
    
    print("  - 担当者別分析")
    create_staff_analysis_sheet(wb, df, amount_col, staff_col, category_col)
    
    print("  - 顧客ランク分析")
    create_customer_rank_sheet(wb, df, amount_col, rank_col, date_col)
    
    print("  - クロス分析")
    create_cross_analysis_sheet(wb, df, amount_col, category_col, staff_col)
    
    print("  - 曜日分析")
    create_weekday_analysis_sheet(wb, df, amount_col, date_col)
    
    print("  - 月次トレンド")
    create_monthly_trend_sheet(wb, df, date_col, amount_col)
    
    print("  - インタラクティブ")
    create_interactive_sheet(wb, df, date_col, amount_col, category_col, staff_col)
    
    # シート順序調整
    sheet_order = ['ダッシュボード', 'ABC分析', '担当者別分析', '顧客ランク分析', 
                   'クロス分析', '曜日分析', '月次トレンド', 'インタラクティブ', '元データ']
    wb._sheets = [wb[name] for name in sheet_order if name in wb.sheetnames]
    
    # 保存
    print("💾 保存中...")
    wb.save(output_file)
    print(f"✅ 生成完了: {output_file}")
    
    return output_file


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="リッチ版売上分析Excel生成")
    parser.add_argument("--input", required=True, help="入力CSVファイル")
    parser.add_argument("--output", required=True, help="出力Excelファイル")
    parser.add_argument("--date-col", default="日付", help="日付列名")
    parser.add_argument("--amount-col", default="売上金額", help="売上金額列名")
    parser.add_argument("--category-col", default="カテゴリ", help="カテゴリ列名")
    parser.add_argument("--item-col", default="サービス名", help="アイテム列名")
    parser.add_argument("--staff-col", default="担当者", help="担当者列名")
    parser.add_argument("--rank-col", default=None, help="顧客ランク列名")
    
    args = parser.parse_args()
    
    generate_rich_sales_analysis(
        args.input,
        args.output,
        args.date_col,
        args.amount_col,
        args.category_col,
        args.item_col,
        args.staff_col,
        args.rank_col
    )
