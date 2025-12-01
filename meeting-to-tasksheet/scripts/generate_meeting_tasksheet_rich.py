#!/usr/bin/env python3
"""
Meeting to Tasksheet Generator - Rich Edition with Dynamic Formulas v2
会議内容から議事録・タスク一覧・ガントチャート付きExcelを生成（動的関数版）

特徴:
- 進捗サマリーの集計はすべてExcel関数（COUNTIF/COUNTIFS/SUMIF/SUMPRODUCT）で実装
- タスク一覧のステータス変更で自動再計算
- ガントチャートは条件付き書式で動的色変更
- 期限状況（超過・今週期限・来週以降）もTODAY()関数で動的判定
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import argparse
import json

# =============================================================================
# スタイル定義
# =============================================================================

FONT_NAME = "メイリオ"

COLORS = {
    'primary': '2E5090',
    'secondary': '4A90E2',
    'accent': '7CB342',
    'warning': 'FFA726',
    'danger': 'EF5350',
    'purple': 'AB47BC',
    'gray': '757575',
    'light_gray': 'F5F5F5',
    'dark': '333333',
    'white': 'FFFFFF',
    'green_bg': 'C8E6C9',
    'orange_bg': 'FFE0B2',
    'red_bg': 'FFCDD2',
    'bar': '666666',
    'gray_bar': 'BDBDBD',
    'gold': 'FFD700',
}

HEADER_FILL = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=20, color=COLORS['primary'])
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=12, color=COLORS['primary'])
NORMAL_FONT = Font(name=FONT_NAME, size=10)
SMALL_FONT = Font(name=FONT_NAME, size=9)

THIN_BORDER = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

STATUS_OPTIONS = "未着手,進行中,完了,保留"
PRIORITY_OPTIONS = "高,中,低"
RISK_LEVEL_OPTIONS = "高,中,低"
RISK_STATUS_OPTIONS = "未対応,対応中,解決済,保留"


# =============================================================================
# シート1: 議事録
# =============================================================================

def create_minutes_sheet(wb, meeting_info, summary, decisions, next_meeting, prev_action_items=None):
    """議事録シート作成"""
    ws = wb.create_sheet("議事録", 0)
    ws.sheet_properties.tabColor = COLORS['primary']
    
    for row in range(1, 60):
        for col in range(1, 15):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    
    ws.merge_cells('B2:H2')
    ws['B2'] = "📋 会議議事録"
    ws['B2'].font = Font(name=FONT_NAME, bold=True, size=26, color=COLORS['primary'])
    
    meeting_no = meeting_info.get("meeting_no", 1)
    ws['I2'] = f"第{meeting_no}回"
    ws['I2'].font = Font(name=FONT_NAME, size=12, color=COLORS['gray'])
    
    ws['B4'] = "■ 会議基本情報"
    ws['B4'].font = SECTION_FONT
    
    info_items = [
        ("会議名", meeting_info.get("title", "")),
        ("日時", meeting_info.get("datetime", "")),
        ("場所", meeting_info.get("location", "")),
        ("参加者", meeting_info.get("attendees", "")),
        ("記録者", meeting_info.get("recorder", "Claude")),
    ]
    
    row = 5
    for label, value in info_items:
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(name=FONT_NAME, bold=True, size=10)
        ws[f'B{row}'].fill = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
        ws[f'C{row}'] = value
        ws[f'C{row}'].font = NORMAL_FONT
        ws.merge_cells(f'C{row}:H{row}')
        row += 1
    
    if prev_action_items:
        row += 1
        ws[f'B{row}'] = "■ 前回からの宿題"
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        for col, header in zip(['B', 'C', 'D', 'E'], ["No.", "内容", "担当", "状況"]):
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].font = HEADER_FONT
            ws[f'{col}{row}'].fill = HEADER_FILL
        row += 1
        
        for i, item in enumerate(prev_action_items, start=1):
            ws[f'B{row}'] = i
            ws[f'C{row}'] = item.get('content', '')
            ws[f'D{row}'] = item.get('assignee', '')
            status = item.get('status', '未完了')
            ws[f'E{row}'] = f"{'✅' if status == '完了' else '⬜'} {status}"
            for col in ['B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].font = NORMAL_FONT
                ws[f'{col}{row}'].border = THIN_BORDER
            row += 1
    
    row += 1
    ws[f'B{row}'] = "■ 議論内容サマリー"
    ws[f'B{row}'].font = SECTION_FONT
    row += 1
    ws[f'B{row}'] = summary
    ws.merge_cells(f'B{row}:H{row+4}')
    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws[f'B{row}'].font = NORMAL_FONT
    ws.row_dimensions[row].height = 80
    
    row += 6
    ws[f'B{row}'] = "■ 決定事項"
    ws[f'B{row}'].font = SECTION_FONT
    row += 1
    
    ws[f'B{row}'] = "No."
    ws[f'C{row}'] = "内容"
    ws[f'B{row}'].font = HEADER_FONT
    ws[f'C{row}'].font = HEADER_FONT
    ws[f'B{row}'].fill = HEADER_FILL
    ws[f'C{row}'].fill = HEADER_FILL
    ws.merge_cells(f'C{row}:H{row}')
    row += 1
    
    for i, decision in enumerate(decisions, start=1):
        ws[f'B{row}'] = i
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        ws[f'C{row}'] = f"✓ {decision}"
        ws[f'B{row}'].font = NORMAL_FONT
        ws[f'C{row}'].font = NORMAL_FONT
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'C{row}'].border = THIN_BORDER
        ws.merge_cells(f'C{row}:H{row}')
        row += 1
    
    row += 1
    ws[f'B{row}'] = "■ 次回会議"
    ws[f'B{row}'].font = SECTION_FONT
    row += 1
    ws[f'B{row}'] = "日時"
    ws[f'B{row}'].font = Font(name=FONT_NAME, bold=True)
    ws[f'C{row}'] = next_meeting.get("datetime", "未定")
    row += 1
    ws[f'B{row}'] = "議題"
    ws[f'B{row}'].font = Font(name=FONT_NAME, bold=True)
    ws[f'C{row}'] = next_meeting.get("agenda", "")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = [2, 12, 15, 12, 12, 12, 12, 15][ord(col) - ord('A')]
    
    return ws


# =============================================================================
# シート2: タスク一覧（動的関数版）
# =============================================================================

def create_task_sheet(wb, tasks, start_date):
    """タスク一覧シート作成（Excel関数で動的計算）"""
    ws = wb.create_sheet("タスク一覧")
    ws.sheet_properties.tabColor = COLORS['secondary']
    
    ws.merge_cells('B2:L2')
    ws['B2'] = "📝 タスク一覧"
    ws['B2'].font = Font(name=FONT_NAME, bold=True, size=20, color=COLORS['primary'])
    
    # 凡例
    ws['B3'] = "【凡例】ステータス: 未着手/進行中/完了/保留 ｜ 優先度: 高/中/低 ｜ ◆マイルストーン ｜ 完了→緑, 保留→オレンジ, 期限超過→赤"
    ws['B3'].font = Font(name=FONT_NAME, size=9, color=COLORS['gray'])
    
    # ヘッダー
    headers = ["No.", "タスク名", "担当者", "開始日", "期限", "残日数", "ステータス", "優先度", "工数(h)", "依存", "備考"]
    data_start_row = 5
    
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=data_start_row - 1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    
    # タスクデータ入力
    for i, task in enumerate(tasks):
        row = data_start_row + i
        
        # No.
        ws.cell(row=row, column=2, value=i + 1).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        # タスク名（マイルストーンは◆付き）
        name = task.get('name', '')
        if task.get('is_milestone'):
            name = f"◆{name}" if not name.startswith('◆') else name
        ws.cell(row=row, column=3, value=name).border = THIN_BORDER
        if task.get('is_milestone'):
            ws.cell(row=row, column=3).font = Font(name=FONT_NAME, bold=True, color=COLORS['primary'])
        else:
            ws.cell(row=row, column=3).font = NORMAL_FONT
        
        # 担当者
        ws.cell(row=row, column=4, value=task.get('assignee', '')).border = THIN_BORDER
        ws.cell(row=row, column=4).font = NORMAL_FONT
        
        # 開始日（DATE関数）
        start_str = task.get('start_date', '')
        if start_str:
            if isinstance(start_str, str):
                parts = start_str.split('-')
                ws.cell(row=row, column=5, value=f"=DATE({parts[0]},{int(parts[1])},{int(parts[2])})")
            else:
                ws.cell(row=row, column=5, value=start_str)
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=5).number_format = 'YYYY-MM-DD'
        ws.cell(row=row, column=5).font = NORMAL_FONT
        
        # 期限（DATE関数）
        due_str = task.get('due_date', '')
        if due_str:
            if isinstance(due_str, str):
                parts = due_str.split('-')
                ws.cell(row=row, column=6, value=f"=DATE({parts[0]},{int(parts[1])},{int(parts[2])})")
            else:
                ws.cell(row=row, column=6, value=due_str)
        ws.cell(row=row, column=6).border = THIN_BORDER
        ws.cell(row=row, column=6).number_format = 'YYYY-MM-DD'
        ws.cell(row=row, column=6).font = NORMAL_FONT
        
        # 残日数（期限 - TODAY()、完了の場合は"-"）
        ws.cell(row=row, column=7, value=f'=IF(H{row}="完了","-",F{row}-TODAY())')
        ws.cell(row=row, column=7).border = THIN_BORDER
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=7).font = NORMAL_FONT
        
        # ステータス
        ws.cell(row=row, column=8, value=task.get('status', '未着手')).border = THIN_BORDER
        ws.cell(row=row, column=8).font = NORMAL_FONT
        ws.cell(row=row, column=8).alignment = Alignment(horizontal='center')
        
        # 優先度
        ws.cell(row=row, column=9, value=task.get('priority', '中')).border = THIN_BORDER
        ws.cell(row=row, column=9).font = NORMAL_FONT
        ws.cell(row=row, column=9).alignment = Alignment(horizontal='center')
        
        # 工数
        ws.cell(row=row, column=10, value=task.get('estimated_hours', 0)).border = THIN_BORDER
        ws.cell(row=row, column=10).font = NORMAL_FONT
        ws.cell(row=row, column=10).alignment = Alignment(horizontal='center')
        
        # 依存
        depends = task.get('depends_on', '')
        ws.cell(row=row, column=11, value=f"←{depends}" if depends else "").border = THIN_BORDER
        ws.cell(row=row, column=11).font = NORMAL_FONT
        
        # 備考
        ws.cell(row=row, column=12, value=task.get('notes', '')).border = THIN_BORDER
        ws.cell(row=row, column=12).font = NORMAL_FONT
        ws.cell(row=row, column=12).alignment = Alignment(wrap_text=True)
    
    data_end_row = data_start_row + len(tasks) - 1
    
    # ドロップダウン（ステータス）
    status_dv = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'H{data_start_row}:H{data_end_row}')
    
    # ドロップダウン（優先度）
    priority_dv = DataValidation(type="list", formula1=f'"{PRIORITY_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(priority_dv)
    priority_dv.add(f'I{data_start_row}:I{data_end_row}')
    
    # 条件付き書式：完了は緑背景
    green_fill = PatternFill(start_color=COLORS['green_bg'], end_color=COLORS['green_bg'], fill_type='solid')
    ws.conditional_formatting.add(
        f'B{data_start_row}:L{data_end_row}',
        FormulaRule(formula=[f'$H{data_start_row}="完了"'], fill=green_fill)
    )
    
    # 条件付き書式：保留はオレンジ背景
    orange_fill = PatternFill(start_color=COLORS['orange_bg'], end_color=COLORS['orange_bg'], fill_type='solid')
    ws.conditional_formatting.add(
        f'B{data_start_row}:L{data_end_row}',
        FormulaRule(formula=[f'$H{data_start_row}="保留"'], fill=orange_fill)
    )
    
    # 条件付き書式：期限超過は期限セル赤背景（未完了のみ）
    red_fill = PatternFill(start_color=COLORS['red_bg'], end_color=COLORS['red_bg'], fill_type='solid')
    ws.conditional_formatting.add(
        f'F{data_start_row}:F{data_end_row}',
        FormulaRule(formula=[f'AND($F{data_start_row}<TODAY(),$H{data_start_row}<>"完了")'], fill=red_fill)
    )
    
    # 残日数が負の場合も赤背景
    ws.conditional_formatting.add(
        f'G{data_start_row}:G{data_end_row}',
        FormulaRule(formula=[f'AND(ISNUMBER($G{data_start_row}),$G{data_start_row}<0)'], fill=red_fill)
    )
    
    # 列幅
    col_widths = [2, 5, 30, 18, 12, 12, 8, 10, 8, 8, 8, 40]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    return ws, data_end_row, data_start_row


# =============================================================================
# シート3: ガントチャート（動的関数版）
# =============================================================================

def create_gantt_sheet(wb, tasks, start_date, end_date, task_data_start, task_data_end):
    """ガントチャートシート作成（条件付き書式で動的色変更）"""
    ws = wb.create_sheet("ガントチャート")
    ws.sheet_properties.tabColor = COLORS['accent']
    
    ws.merge_cells('B1:F1')
    ws['B1'] = "📊 ガントチャート"
    ws['B1'].font = Font(name=FONT_NAME, bold=True, size=20, color=COLORS['primary'])
    
    # 基準日（変更可能）
    ws['B2'] = "基準日:"
    ws['B2'].font = Font(name=FONT_NAME, size=9)
    ws['C2'] = f"=DATE({start_date.year},{start_date.month},{start_date.day})"
    ws['C2'].number_format = 'YYYY-MM-DD'
    ws['C2'].font = Font(name=FONT_NAME, size=9)
    
    # ヘッダー
    headers = ["No.", "タスク名", "担当", "開始", "期限", "状態"]
    header_row = 3
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(name=FONT_NAME, bold=True, color=COLORS['white'], size=9)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    
    # 週ヘッダー（24週分 = 約6ヶ月）
    num_weeks = 24
    base_date_col = 8
    
    for week in range(num_weeks):
        col = base_date_col + week
        # 週の開始日を数式で計算
        cell = ws.cell(row=header_row, column=col, value=f"=$C$2+{week*7}")
        cell.number_format = 'MM/DD'
        cell.font = Font(name=FONT_NAME, size=8)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type='solid')
        cell.border = THIN_BORDER
    
    # タスクデータ（タスク一覧シートを参照）
    data_start_row = 4
    for i, task in enumerate(tasks):
        row = data_start_row + i
        task_sheet_row = task_data_start + i
        
        # タスク一覧シートを参照
        ws.cell(row=row, column=2, value=f"=タスク一覧!B{task_sheet_row}").border = THIN_BORDER
        ws.cell(row=row, column=2).font = Font(name=FONT_NAME, size=9)
        
        ws.cell(row=row, column=3, value=f"=タスク一覧!C{task_sheet_row}").border = THIN_BORDER
        ws.cell(row=row, column=3).font = Font(name=FONT_NAME, size=9)
        
        ws.cell(row=row, column=4, value=f"=タスク一覧!D{task_sheet_row}").border = THIN_BORDER
        ws.cell(row=row, column=4).font = Font(name=FONT_NAME, size=9)
        
        ws.cell(row=row, column=5, value=f"=タスク一覧!E{task_sheet_row}").border = THIN_BORDER
        ws.cell(row=row, column=5).number_format = 'MM/DD'
        ws.cell(row=row, column=5).font = Font(name=FONT_NAME, size=9)
        
        ws.cell(row=row, column=6, value=f"=タスク一覧!F{task_sheet_row}").border = THIN_BORDER
        ws.cell(row=row, column=6).number_format = 'MM/DD'
        ws.cell(row=row, column=6).font = Font(name=FONT_NAME, size=9)
        
        ws.cell(row=row, column=7, value=f"=タスク一覧!H{task_sheet_row}").border = THIN_BORDER
        ws.cell(row=row, column=7).font = Font(name=FONT_NAME, size=9)
        
        # ガントバー：各週のセルに条件式
        for week in range(num_weeks):
            col = base_date_col + week
            cell = ws.cell(row=row, column=col)
            week_start_ref = f"${get_column_letter(col)}$3"
            task_start_ref = f"$E{row}"
            task_end_ref = f"$F{row}"
            
            # この週がタスク期間内なら"■"を表示
            cell.value = f'=IF(AND({week_start_ref}>={task_start_ref},{week_start_ref}<={task_end_ref}+6),"■","")'
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
    
    data_end_row = data_start_row + len(tasks) - 1
    gantt_range = f"{get_column_letter(base_date_col)}{data_start_row}:{get_column_letter(base_date_col + num_weeks - 1)}{data_end_row}"
    
    # 条件付き書式：ステータスに応じた色
    # 完了: 緑
    green_fill = PatternFill(start_color=COLORS['accent'], end_color=COLORS['accent'], fill_type='solid')
    ws.conditional_formatting.add(
        gantt_range,
        FormulaRule(formula=[f'AND($G{data_start_row}="完了",{get_column_letter(base_date_col)}{data_start_row}="■")'], fill=green_fill)
    )
    
    # 進行中: 青
    blue_fill = PatternFill(start_color=COLORS['bar'], end_color=COLORS['bar'], fill_type='solid')
    ws.conditional_formatting.add(
        gantt_range,
        FormulaRule(formula=[f'AND($G{data_start_row}="進行中",{get_column_letter(base_date_col)}{data_start_row}="■")'], fill=blue_fill)
    )
    
    # 保留: オレンジ
    orange_fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
    ws.conditional_formatting.add(
        gantt_range,
        FormulaRule(formula=[f'AND($G{data_start_row}="保留",{get_column_letter(base_date_col)}{data_start_row}="■")'], fill=orange_fill)
    )
    
    # 未着手: グレー
    gray_fill = PatternFill(start_color=COLORS['gray_bar'], end_color=COLORS['gray_bar'], fill_type='solid')
    ws.conditional_formatting.add(
        gantt_range,
        FormulaRule(formula=[f'AND($G{data_start_row}="未着手",{get_column_letter(base_date_col)}{data_start_row}="■")'], fill=gray_fill)
    )
    
    # 今週の列をハイライト
    for week in range(num_weeks):
        col = base_date_col + week
        col_letter = get_column_letter(col)
        today_range = f"{col_letter}{header_row}:{col_letter}{data_end_row}"
        red_fill = PatternFill(start_color=COLORS['red_bg'], end_color=COLORS['red_bg'], fill_type='solid')
        ws.conditional_formatting.add(
            today_range,
            FormulaRule(formula=[f'AND({col_letter}$3<=TODAY(),{col_letter}$3+6>=TODAY())'], fill=red_fill)
        )
    
    # 列幅
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 4
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 8
    for c in range(base_date_col, base_date_col + num_weeks):
        ws.column_dimensions[get_column_letter(c)].width = 4
    
    # 凡例
    legend_row = data_end_row + 2
    ws[f'B{legend_row}'] = "【凡例】■ 未着手=グレー / 進行中=青 / 完了=緑 / 保留=オレンジ ｜ 今週=赤背景 ｜ タスク一覧でステータスを変更するとバーの色が自動で変わります"
    ws[f'B{legend_row}'].font = Font(name=FONT_NAME, size=9, color=COLORS['gray'])
    
    return ws


# =============================================================================
# シート4: 進捗サマリー（完全動的関数版）
# =============================================================================

def create_summary_sheet(wb, tasks, task_data_start, task_data_end):
    """進捗サマリーシート作成（全てExcel関数で動的計算）"""
    ws = wb.create_sheet("進捗サマリー")
    ws.sheet_properties.tabColor = COLORS['warning']
    
    task_count = task_data_end - task_data_start + 1
    status_range = f"タスク一覧!H{task_data_start}:H{task_data_end}"
    priority_range = f"タスク一覧!I{task_data_start}:I{task_data_end}"
    hours_range = f"タスク一覧!J{task_data_start}:J{task_data_end}"
    assignee_range = f"タスク一覧!D{task_data_start}:D{task_data_end}"
    due_range = f"タスク一覧!F{task_data_start}:F{task_data_end}"
    
    # タイトル
    ws.merge_cells('A1:F1')
    ws['A1'] = "📊 進捗サマリー"
    ws['A1'].font = Font(name=FONT_NAME, size=20, bold=True, color=COLORS['primary'])
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35
    
    # 完了率（大きく表示）
    ws.merge_cells('A3:B5')
    ws['A3'] = f'=TEXT(COUNTIF({status_range},"完了")/{task_count},"0%")'
    ws['A3'].font = Font(name=FONT_NAME, size=48, bold=True, color=COLORS['primary'])
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['C3'] = "完了率"
    ws['C3'].font = Font(name=FONT_NAME, size=14, bold=True)
    
    ws['C4'] = f'=CONCATENATE("完了: ",COUNTIF({status_range},"完了")," / 全",{task_count},"タスク")'
    ws['C4'].font = Font(name=FONT_NAME, size=12)
    
    ws['C5'] = f'=CONCATENATE("総工数: ",SUM({hours_range}),"h")'
    ws['C5'].font = Font(name=FONT_NAME, size=11, color=COLORS['gray'])
    
    # ステータス別集計
    ws['A7'] = "■ ステータス別集計"
    ws['A7'].font = Font(name=FONT_NAME, size=12, bold=True)
    
    status_headers = ["ステータス", "件数", "割合"]
    for col, header in enumerate(status_headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    
    statuses = ["未着手", "進行中", "完了", "保留"]
    for i, status in enumerate(statuses):
        row = 9 + i
        ws.cell(row=row, column=1, value=status).border = THIN_BORDER
        ws.cell(row=row, column=1).font = NORMAL_FONT
        
        # COUNTIF関数で動的にカウント
        ws.cell(row=row, column=2, value=f'=COUNTIF({status_range},"{status}")').border = THIN_BORDER
        ws.cell(row=row, column=2).font = NORMAL_FONT
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        # 割合
        ws.cell(row=row, column=3, value=f'=IF({task_count}>0,B{row}/{task_count},0)').border = THIN_BORDER
        ws.cell(row=row, column=3).font = NORMAL_FONT
        ws.cell(row=row, column=3).number_format = '0%'
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    
    # 合計行
    ws.cell(row=13, column=1, value="合計").border = THIN_BORDER
    ws.cell(row=13, column=1).font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=13, column=2, value="=SUM(B9:B12)").border = THIN_BORDER
    ws.cell(row=13, column=2).font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=13, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=13, column=3, value="=SUM(C9:C12)").border = THIN_BORDER
    ws.cell(row=13, column=3).font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=13, column=3).number_format = '0%'
    ws.cell(row=13, column=3).alignment = Alignment(horizontal='center')
    
    # 優先度別集計
    ws['A15'] = "■ 優先度別集計"
    ws['A15'].font = Font(name=FONT_NAME, size=12, bold=True)
    
    priority_headers = ["優先度", "件数", "完了数", "完了率"]
    for col, header in enumerate(priority_headers, 1):
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    
    priorities = ["高", "中", "低"]
    for i, priority in enumerate(priorities):
        row = 17 + i
        ws.cell(row=row, column=1, value=priority).border = THIN_BORDER
        ws.cell(row=row, column=1).font = NORMAL_FONT
        
        ws.cell(row=row, column=2, value=f'=COUNTIF({priority_range},"{priority}")').border = THIN_BORDER
        ws.cell(row=row, column=2).font = NORMAL_FONT
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=3, value=f'=COUNTIFS({priority_range},"{priority}",{status_range},"完了")').border = THIN_BORDER
        ws.cell(row=row, column=3).font = NORMAL_FONT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=4, value=f'=IF(B{row}>0,C{row}/B{row},0)').border = THIN_BORDER
        ws.cell(row=row, column=4).font = NORMAL_FONT
        ws.cell(row=row, column=4).number_format = '0%'
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    
    # 担当者別集計
    ws['A22'] = "■ 担当者別タスク数・工数"
    ws['A22'].font = Font(name=FONT_NAME, size=12, bold=True)
    
    assignee_headers = ["担当者", "タスク数", "総工数(h)", "完了数", "完了率"]
    for col, header in enumerate(assignee_headers, 1):
        cell = ws.cell(row=23, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    
    assignees = list(set(t.get('assignee', '') for t in tasks if t.get('assignee')))
    for i, assignee in enumerate(assignees):
        row = 24 + i
        ws.cell(row=row, column=1, value=assignee).border = THIN_BORDER
        ws.cell(row=row, column=1).font = NORMAL_FONT
        
        ws.cell(row=row, column=2, value=f'=COUNTIF({assignee_range},A{row})').border = THIN_BORDER
        ws.cell(row=row, column=2).font = NORMAL_FONT
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=3, value=f'=SUMIF({assignee_range},A{row},{hours_range})').border = THIN_BORDER
        ws.cell(row=row, column=3).font = NORMAL_FONT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=4, value=f'=COUNTIFS({assignee_range},A{row},{status_range},"完了")').border = THIN_BORDER
        ws.cell(row=row, column=4).font = NORMAL_FONT
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=5, value=f'=IF(B{row}>0,D{row}/B{row},0)').border = THIN_BORDER
        ws.cell(row=row, column=5).font = NORMAL_FONT
        ws.cell(row=row, column=5).number_format = '0%'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
    
    # 期限状況
    deadline_start_row = 24 + len(assignees) + 2
    ws[f'A{deadline_start_row}'] = "■ 期限状況"
    ws[f'A{deadline_start_row}'].font = Font(name=FONT_NAME, size=12, bold=True)
    
    deadline_headers = ["状況", "件数"]
    for col, header in enumerate(deadline_headers, 1):
        cell = ws.cell(row=deadline_start_row + 1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    
    # 期限超過（未完了のみ）
    row = deadline_start_row + 2
    ws.cell(row=row, column=1, value="🚨 期限超過").border = THIN_BORDER
    ws.cell(row=row, column=1).font = NORMAL_FONT
    ws.cell(row=row, column=2, value=f'=SUMPRODUCT(({due_range}<TODAY())*({status_range}<>"完了"))').border = THIN_BORDER
    ws.cell(row=row, column=2).font = Font(name=FONT_NAME, color=COLORS['danger'])
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    
    # 今週期限
    row += 1
    ws.cell(row=row, column=1, value="⚠️ 今週期限").border = THIN_BORDER
    ws.cell(row=row, column=1).font = NORMAL_FONT
    ws.cell(row=row, column=2, value=f'=SUMPRODUCT(({due_range}>=TODAY())*({due_range}<=TODAY()+7)*({status_range}<>"完了"))').border = THIN_BORDER
    ws.cell(row=row, column=2).font = Font(name=FONT_NAME, color=COLORS['warning'])
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    
    # 来週以降
    row += 1
    ws.cell(row=row, column=1, value="📅 来週以降").border = THIN_BORDER
    ws.cell(row=row, column=1).font = NORMAL_FONT
    ws.cell(row=row, column=2, value=f'=SUMPRODUCT(({due_range}>TODAY()+7)*({status_range}<>"完了"))').border = THIN_BORDER
    ws.cell(row=row, column=2).font = NORMAL_FONT
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    
    # 列幅
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    
    # 注記
    note_row = row + 3
    ws[f'A{note_row}'] = "※ タスク一覧シートのステータス・優先度を変更すると、この集計も自動的に更新されます"
    ws[f'A{note_row}'].font = Font(name=FONT_NAME, size=9, color=COLORS['gray'], italic=True)
    
    return ws


# =============================================================================
# シート5: リスク・課題トラッキング
# =============================================================================

def create_risk_sheet(wb, risks):
    """リスク・課題トラッキングシート"""
    ws = wb.create_sheet("リスク・課題")
    ws.sheet_properties.tabColor = COLORS['danger']
    
    ws.merge_cells('B2:J2')
    ws['B2'] = "⚠️ リスク・課題トラッキング"
    ws['B2'].font = Font(name=FONT_NAME, bold=True, size=20, color=COLORS['danger'])
    
    # 凡例
    ws['B3'] = "【凡例】レベル: 高/中/低 ｜ 状態: 未対応/対応中/解決済/保留 ｜ 高リスク→赤背景"
    ws['B3'].font = Font(name=FONT_NAME, size=9, color=COLORS['gray'])
    
    headers = ["No.", "種別", "内容", "レベル", "影響", "対策", "担当", "期限", "状態"]
    data_start_row = 6
    
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=data_start_row - 1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color=COLORS['danger'], end_color=COLORS['danger'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    for i, risk in enumerate(risks, start=data_start_row):
        ws.cell(row=i, column=2, value=i - data_start_row + 1).alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=3, value=risk.get('type', 'リスク'))
        ws.cell(row=i, column=4, value=risk.get('content', ''))
        ws.cell(row=i, column=5, value=risk.get('level', '中'))
        ws.cell(row=i, column=6, value=risk.get('impact', ''))
        ws.cell(row=i, column=7, value=risk.get('countermeasure', ''))
        ws.cell(row=i, column=8, value=risk.get('assignee', ''))
        ws.cell(row=i, column=9, value=risk.get('due_date', ''))
        ws.cell(row=i, column=10, value=risk.get('status', '未対応'))
        
        for col in range(2, 11):
            ws.cell(row=i, column=col).font = NORMAL_FONT
            ws.cell(row=i, column=col).border = THIN_BORDER
            ws.cell(row=i, column=col).alignment = Alignment(vertical='center', wrap_text=True)
    
    data_end_row = data_start_row + len(risks) - 1 if risks else data_start_row
    
    if risks:
        # レベルドロップダウン
        level_dv = DataValidation(type="list", formula1=f'"{RISK_LEVEL_OPTIONS}"', allow_blank=True)
        ws.add_data_validation(level_dv)
        level_dv.add(f'E{data_start_row}:E{data_end_row}')
        
        # ステータスドロップダウン
        status_dv = DataValidation(type="list", formula1=f'"{RISK_STATUS_OPTIONS}"', allow_blank=True)
        ws.add_data_validation(status_dv)
        status_dv.add(f'J{data_start_row}:J{data_end_row}')
        
        # 条件付き書式：高リスクは赤背景
        red_fill = PatternFill(start_color=COLORS['red_bg'], end_color=COLORS['red_bg'], fill_type='solid')
        ws.conditional_formatting.add(
            f'B{data_start_row}:J{data_end_row}',
            FormulaRule(formula=[f'$E{data_start_row}="高"'], fill=red_fill)
        )
    
    # 集計（動的関数）
    summary_row = data_end_row + 3 if risks else 8
    ws[f'B{summary_row}'] = "■ 集計"
    ws[f'B{summary_row}'].font = SECTION_FONT
    
    if risks:
        level_range = f"E{data_start_row}:E{data_end_row}"
        status_range = f"J{data_start_row}:J{data_end_row}"
        
        ws[f'B{summary_row+1}'] = "総件数"
        ws[f'C{summary_row+1}'] = f'=COUNTA(D{data_start_row}:D{data_end_row})'
        ws[f'C{summary_row+1}'].font = NORMAL_FONT
        
        ws[f'B{summary_row+2}'] = "🚨 高リスク"
        ws[f'C{summary_row+2}'] = f'=COUNTIF({level_range},"高")'
        ws[f'C{summary_row+2}'].font = Font(name=FONT_NAME, bold=True, color=COLORS['danger'])
        
        ws[f'B{summary_row+3}'] = "未対応"
        ws[f'C{summary_row+3}'] = f'=COUNTIF({status_range},"未対応")'
        ws[f'C{summary_row+3}'].font = NORMAL_FONT
        
        ws[f'B{summary_row+4}'] = "解決済"
        ws[f'C{summary_row+4}'] = f'=COUNTIF({status_range},"解決済")'
        ws[f'C{summary_row+4}'].font = NORMAL_FONT
    
    for col, width in zip(['B','C','D','E','F','G','H','I','J'], [5, 10, 28, 8, 18, 25, 15, 12, 10]):
        ws.column_dimensions[col].width = width
    
    return ws


# =============================================================================
# シート6: Slack投稿用サマリー
# =============================================================================

def create_slack_summary_sheet(wb, meeting_info, decisions, tasks, risks, task_data_start, task_data_end):
    """Slack/Teams投稿用サマリー（動的関数で進捗表示）"""
    ws = wb.create_sheet("Slack投稿用")
    ws.sheet_properties.tabColor = '4A154B'
    
    ws['B2'] = "💬 Slack/Teams投稿用サマリー"
    ws['B2'].font = Font(name=FONT_NAME, bold=True, size=20, color=COLORS['primary'])
    
    ws['B3'] = "下記テキストをコピーしてSlack/Teamsに貼り付けてください"
    ws['B3'].font = Font(name=FONT_NAME, size=10, color=COLORS['gray'])
    
    # 動的進捗表示
    task_count = task_data_end - task_data_start + 1
    status_range = f"タスク一覧!H{task_data_start}:H{task_data_end}"
    due_range = f"タスク一覧!F{task_data_start}:F{task_data_end}"
    
    ws['B5'] = "■ 現在の進捗状況（自動更新）"
    ws['B5'].font = SECTION_FONT
    
    ws['B6'] = "完了率:"
    ws['C6'] = f'=COUNTIF({status_range},"完了")/{task_count}'
    ws['C6'].number_format = '0%'
    ws['C6'].font = Font(name=FONT_NAME, bold=True, size=14)
    
    ws['B7'] = "完了タスク:"
    ws['C7'] = f'=COUNTIF({status_range},"完了")&"/"&{task_count}&"件"'
    ws['C7'].font = NORMAL_FONT
    
    ws['B8'] = "期限超過:"
    ws['C8'] = f'=SUMPRODUCT(({due_range}<TODAY())*({status_range}<>"完了"))&"件"'
    ws['C8'].font = Font(name=FONT_NAME, color=COLORS['danger'])
    
    ws['B9'] = "今週期限:"
    ws['C9'] = f'=SUMPRODUCT(({due_range}>=TODAY())*({due_range}<=TODAY()+7)*({status_range}<>"完了"))&"件"'
    ws['C9'].font = Font(name=FONT_NAME, color=COLORS['warning'])
    
    # マークダウン形式
    row = 12
    ws[f'B{row}'] = "■ マークダウン形式"
    ws[f'B{row}'].font = SECTION_FONT
    row += 1
    
    md_lines = [
        f"# 📋 {meeting_info.get('title', '会議')} 議事録",
        f"**日時**: {meeting_info.get('datetime', '')}",
        f"**参加者**: {meeting_info.get('attendees', '')}",
        "",
        "## ✅ 決定事項",
    ]
    for i, d in enumerate(decisions, 1):
        md_lines.append(f"{i}. {d}")
    
    md_lines.extend(["", "## 📝 タスク一覧", "| タスク | 担当 | 期限 | 優先度 |", "|--------|------|------|--------|"])
    for t in tasks:
        name = t.get('name', '').replace('◆', '')
        md_lines.append(f"| {name} | {t.get('assignee','')} | {t.get('due_date','')} | {t.get('priority', '中')} |")
    
    if risks:
        md_lines.extend(["", "## ⚠️ 主要リスク"])
        for r in risks[:3]:
            md_lines.append(f"- [{r.get('level', '中')}] {r.get('content', '')}")
    
    for line in md_lines:
        ws[f'B{row}'] = line
        ws[f'B{row}'].font = Font(name="Consolas", size=10)
        row += 1
    
    # プレーンテキスト形式
    row += 2
    ws[f'B{row}'] = "■ プレーンテキスト形式"
    ws[f'B{row}'].font = SECTION_FONT
    row += 1
    
    plain_lines = [
        f"【{meeting_info.get('title', '会議')} 議事録】",
        f"日時: {meeting_info.get('datetime', '')}",
        f"参加者: {meeting_info.get('attendees', '')}",
        "",
        "▼ 決定事項",
    ]
    for i, d in enumerate(decisions, 1):
        plain_lines.append(f"  {i}. {d}")
    
    plain_lines.extend(["", "▼ タスク一覧"])
    for t in tasks:
        name = t.get('name', '').replace('◆', '')
        plain_lines.append(f"  ・{name} / {t.get('assignee','')} / {t.get('due_date','')}")
    
    for line in plain_lines:
        ws[f'B{row}'] = line
        ws[f'B{row}'].font = Font(name=FONT_NAME, size=10)
        row += 1
    
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 15
    
    return ws


# =============================================================================
# メイン処理
# =============================================================================

def generate_meeting_tasksheet(meeting_info, summary, decisions, tasks, risks=None, next_meeting=None, prev_action_items=None, output_file="meeting_tasksheet.xlsx"):
    """メイン生成関数"""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 日付範囲計算
    all_dates = []
    for t in tasks:
        for key in ['start_date', 'due_date']:
            d = t.get(key)
            if d:
                if isinstance(d, str):
                    try:
                        d = datetime.strptime(d, '%Y-%m-%d').date()
                    except:
                        continue
                all_dates.append(d)
    
    start_date = min(all_dates) if all_dates else datetime.now().date()
    end_date = (max(all_dates) if all_dates else start_date) + timedelta(days=5)
    
    # date型に変換
    if hasattr(start_date, 'date'):
        start_date = start_date
    
    next_meeting = next_meeting or {}
    risks = risks or []
    
    # シート作成
    create_minutes_sheet(wb, meeting_info, summary, decisions, next_meeting, prev_action_items)
    _, task_data_end, task_data_start = create_task_sheet(wb, tasks, start_date)
    create_gantt_sheet(wb, tasks, start_date, end_date, task_data_start, task_data_end)
    create_summary_sheet(wb, tasks, task_data_start, task_data_end)
    if risks:
        create_risk_sheet(wb, risks)
    create_slack_summary_sheet(wb, meeting_info, decisions, tasks, risks, task_data_start, task_data_end)
    
    wb.save(output_file)
    print(f"Generated: {output_file}")
    return output_file


def get_demo_data():
    """デモ用データ"""
    return {
        "meeting_info": {
            "title": "Claude Agent Skills セミナー打ち合わせ",
            "datetime": "2025年1月15日（水）14:00-15:30",
            "location": "オンライン（Zoom）",
            "attendees": "MIRAIさん、ダイチさん、Ken",
            "recorder": "Claude",
            "meeting_no": 1,
        },
        "summary": "Claude Agent Skillsセミナーの企画・準備に関する打ち合わせを実施。開催日は2月20日（木）に決定。会場は商工会議所の大会議室を使用。定員30名で参加費無料。告知は複数チャネルで実施予定。",
        "decisions": [
            "開催日: 2025年2月20日（木）13:00-16:00",
            "会場: 商工会議所 大会議室",
            "定員: 30名（先着順）",
            "参加費: 無料",
        ],
        "tasks": [
            {"name": "会場予約確定", "assignee": "MIRAIさん", "start_date": "2025-01-15", "due_date": "2025-01-20", "status": "進行中", "priority": "高", "estimated_hours": 1},
            {"name": "告知チラシ作成", "assignee": "MIRAIさん", "start_date": "2025-01-16", "due_date": "2025-01-25", "status": "未着手", "priority": "高", "estimated_hours": 4, "depends_on": "1"},
            {"name": "募集ページ作成", "assignee": "ダイチさん", "start_date": "2025-01-16", "due_date": "2025-01-25", "status": "未着手", "priority": "高", "estimated_hours": 3},
            {"name": "デモ用サンプルデータ準備", "assignee": "Ken", "start_date": "2025-01-15", "due_date": "2025-01-25", "status": "完了", "priority": "中", "estimated_hours": 2},
            {"name": "スライド資料作成", "assignee": "Ken", "start_date": "2025-01-20", "due_date": "2025-02-10", "status": "未着手", "priority": "中", "estimated_hours": 8, "depends_on": "4"},
            {"name": "デモ環境準備", "assignee": "Ken", "start_date": "2025-02-01", "due_date": "2025-02-10", "status": "未着手", "priority": "高", "estimated_hours": 4},
            {"name": "リハーサル", "assignee": "全員", "start_date": "2025-02-15", "due_date": "2025-02-15", "status": "未着手", "priority": "高", "estimated_hours": 3, "is_milestone": True, "depends_on": "5,6"},
            {"name": "セミナー本番", "assignee": "全員", "start_date": "2025-02-20", "due_date": "2025-02-20", "status": "未着手", "priority": "高", "estimated_hours": 4, "is_milestone": True, "depends_on": "7"},
            {"name": "フォローアップ", "assignee": "MIRAIさん", "start_date": "2025-02-21", "due_date": "2025-02-25", "status": "未着手", "priority": "中", "estimated_hours": 2, "depends_on": "8"},
        ],
        "risks": [
            {"type": "リスク", "content": "ネットワーク環境が不安定", "level": "高", "impact": "デモ失敗", "countermeasure": "モバイルルーター準備", "assignee": "Ken", "due_date": "2025-02-10", "status": "対応中"},
            {"type": "リスク", "content": "参加者が集まらない", "level": "中", "impact": "効果減少", "countermeasure": "複数チャネルで告知", "assignee": "MIRAIさん", "due_date": "2025-02-15", "status": "未対応"},
        ],
        "next_meeting": {"datetime": "2025年1月29日（水）14:00-15:00", "agenda": "進捗確認・告知状況共有"},
        "prev_action_items": [
            {"content": "会場候補リストアップ", "assignee": "MIRAIさん", "status": "完了"},
            {"content": "デモ素案作成", "assignee": "Ken", "status": "完了"},
        ],
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Meeting to Tasksheet Generator - Rich Edition with Dynamic Formulas")
    parser.add_argument("--output", default="meeting_tasksheet.xlsx", help="出力ファイル名")
    parser.add_argument("--demo", action="store_true", help="デモデータで生成")
    parser.add_argument("--json", help="JSONファイルからデータを読み込み")
    args = parser.parse_args()
    
    if args.json:
        with open(args.json, encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = get_demo_data()
    
    generate_meeting_tasksheet(
        meeting_info=data.get("meeting_info", {}),
        summary=data.get("summary", ""),
        decisions=data.get("decisions", []),
        tasks=data.get("tasks", []),
        risks=data.get("risks"),
        next_meeting=data.get("next_meeting"),
        prev_action_items=data.get("prev_action_items"),
        output_file=args.output
    )
