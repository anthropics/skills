#!/usr/bin/env python3
"""
Meeting to Tasksheet Generator - Simple Version
会議内容から議事録・タスク一覧・ガントチャート付きExcelを生成（シンプル版）

リッチ版（generate_meeting_tasksheet_rich.py）の使用を推奨します。
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import argparse

# スタイル定義
FONT_NAME = "メイリオ"
HEADER_FILL = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="2E5090")
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=12, color="2E5090")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ガントチャートの色
GANTT_FILL = PatternFill(start_color="666666", end_color="666666", fill_type="solid")
TODAY_FILL = PatternFill(start_color="5B7B94", end_color="5B7B94", fill_type="solid")
MILESTONE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")


def create_minutes_sheet(wb, meeting_info, summary, decisions, next_meeting):
    """議事録シート作成"""
    ws = wb.create_sheet("議事録", 0)
    
    # タイトル
    ws['B2'] = "会議議事録"
    ws['B2'].font = Font(name=FONT_NAME, bold=True, size=24, color="2E5090")
    
    # 基本情報
    ws['B4'] = "会議基本情報"
    ws['B4'].font = SECTION_FONT
    
    info_items = [
        ("会議名", meeting_info.get("title", "")),
        ("日時", meeting_info.get("datetime", "")),
        ("場所", meeting_info.get("location", "")),
        ("参加者", meeting_info.get("attendees", "")),
    ]
    
    row = 5
    for label, value in info_items:
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(name=FONT_NAME, bold=True)
        ws[f'C{row}'] = value
        ws[f'C{row}'].font = NORMAL_FONT
        row += 1
    
    # 議論サマリー
    row += 1
    ws[f'B{row}'] = "議論内容サマリー"
    ws[f'B{row}'].font = SECTION_FONT
    row += 1
    ws[f'B{row}'] = summary
    ws.merge_cells(f'B{row}:F{row+3}')
    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # 決定事項
    row += 5
    ws[f'B{row}'] = "決定事項"
    ws[f'B{row}'].font = SECTION_FONT
    row += 1
    
    ws[f'B{row}'] = "No."
    ws[f'C{row}'] = "内容"
    ws[f'B{row}'].font = HEADER_FONT
    ws[f'C{row}'].font = HEADER_FONT
    ws[f'B{row}'].fill = HEADER_FILL
    ws[f'C{row}'].fill = HEADER_FILL
    
    for i, decision in enumerate(decisions, start=1):
        row += 1
        ws[f'B{row}'] = i
        ws[f'C{row}'] = decision
    
    # 次回予定
    row += 2
    ws[f'B{row}'] = "次回会議"
    ws[f'B{row}'].font = SECTION_FONT
    row += 1
    ws[f'B{row}'] = next_meeting
    
    # 列幅
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    
    return ws


def create_task_list_sheet(wb, tasks):
    """タスク一覧シート作成"""
    ws = wb.create_sheet("タスク一覧")
    
    # タイトル
    ws['B2'] = "タスク一覧"
    ws['B2'].font = Font(name=FONT_NAME, bold=True, size=20, color="2E5090")
    
    # ヘッダー
    headers = ['No.', 'タスク名', '担当者', '開始日', '期限', 'ステータス', '優先度', '備考']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    # データ行
    for i, task in enumerate(tasks, start=5):
        ws.cell(row=i, column=2, value=i-4)
        ws.cell(row=i, column=3, value=task['name'])
        ws.cell(row=i, column=4, value=task['assignee'])
        ws.cell(row=i, column=5, value=task['start_date'])
        ws.cell(row=i, column=5).number_format = 'YYYY/MM/DD'
        ws.cell(row=i, column=6, value=task['due_date'])
        ws.cell(row=i, column=6).number_format = 'YYYY/MM/DD'
        ws.cell(row=i, column=7, value=task['status'])
        ws.cell(row=i, column=8, value=task['priority'])
        ws.cell(row=i, column=9, value=task.get('note', ''))
        
        # 罫線
        for col in range(2, 10):
            ws.cell(row=i, column=col).border = THIN_BORDER
    
    # ステータス選択リスト
    status_validation = DataValidation(
        type="list",
        formula1='"未着手,進行中,完了,保留"',
        allow_blank=True
    )
    ws.add_data_validation(status_validation)
    status_validation.add(f'G5:G{4+len(tasks)}')
    
    # 優先度選択リスト
    priority_validation = DataValidation(
        type="list",
        formula1='"高,中,低"',
        allow_blank=True
    )
    ws.add_data_validation(priority_validation)
    priority_validation.add(f'H5:H{4+len(tasks)}')
    
    # 条件付き書式（ステータス色分け）
    ws.conditional_formatting.add(
        f'G5:G{4+len(tasks)}',
        FormulaRule(formula=['$G5="完了"'], 
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f'G5:G{4+len(tasks)}',
        FormulaRule(formula=['$G5="進行中"'], 
                   fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f'G5:G{4+len(tasks)}',
        FormulaRule(formula=['$G5="保留"'], 
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    
    # 期限アラート（期限超過で赤）
    ws.conditional_formatting.add(
        f'F5:F{4+len(tasks)}',
        FormulaRule(formula=['AND($F5<TODAY(),$G5<>"完了")'], 
                   fill=PatternFill(start_color="5B7B94", end_color="5B7B94", fill_type="solid"))
    )
    # 期限3日以内で黄
    ws.conditional_formatting.add(
        f'F5:F{4+len(tasks)}',
        FormulaRule(formula=['AND($F5<=TODAY()+3,$F5>=TODAY(),$G5<>"完了")'], 
                   fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    
    # 列幅
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 25
    
    return ws


def create_gantt_chart_sheet(wb, tasks, start_date, end_date):
    """ガントチャートシート作成"""
    ws = wb.create_sheet("ガントチャート")
    
    # タイトル
    ws['B2'] = "ガントチャート"
    ws['B2'].font = Font(name=FONT_NAME, bold=True, size=20, color="2E5090")
    
    # 日付範囲を計算
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    days = (end - start).days + 1
    
    # ヘッダー行（タスク情報）
    ws['B4'] = "タスク"
    ws['C4'] = "担当"
    ws['D4'] = "開始"
    ws['E4'] = "期限"
    for cell in [ws['B4'], ws['C4'], ws['D4'], ws['E4']]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    # 日付ヘッダー
    date_start_col = 6
    for i in range(days):
        current_date = start + timedelta(days=i)
        col = date_start_col + i
        cell = ws.cell(row=3, column=col, value=current_date.day)
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(name=FONT_NAME, size=9)
        
        # 月が変わったら月名を上の行に
        if i == 0 or current_date.day == 1:
            month_cell = ws.cell(row=2, column=col, value=f"{current_date.month}月")
            month_cell.font = Font(name=FONT_NAME, bold=True, size=10)
        
        # 土日は背景色
        if current_date.weekday() >= 5:
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        ws.cell(row=4, column=col, value=current_date)
        ws.cell(row=4, column=col).number_format = 'D'
        ws.cell(row=4, column=col).font = Font(name=FONT_NAME, size=8)
        
        # 列幅を狭く
        ws.column_dimensions[get_column_letter(col)].width = 3
    
    # タスク行
    for row_idx, task in enumerate(tasks, start=5):
        ws.cell(row=row_idx, column=2, value=task['name'])
        ws.cell(row=row_idx, column=3, value=task['assignee'])
        
        task_start = task['start_date']
        task_end = task['due_date']
        
        if isinstance(task_start, str):
            task_start = datetime.strptime(task_start, "%Y-%m-%d")
        if isinstance(task_end, str):
            task_end = datetime.strptime(task_end, "%Y-%m-%d")
        
        ws.cell(row=row_idx, column=4, value=task_start)
        ws.cell(row=row_idx, column=4).number_format = 'MM/DD'
        ws.cell(row=row_idx, column=5, value=task_end)
        ws.cell(row=row_idx, column=5).number_format = 'MM/DD'
        
        # ガントバー
        for i in range(days):
            current_date = start + timedelta(days=i)
            col = date_start_col + i
            
            if task_start <= current_date <= task_end:
                if task.get('is_milestone'):
                    ws.cell(row=row_idx, column=col).fill = MILESTONE_FILL
                else:
                    ws.cell(row=row_idx, column=col).fill = GANTT_FILL
            
            # 土日背景
            elif current_date.weekday() >= 5:
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # 今日のラインを強調
    today = datetime.now()
    if start <= today <= end:
        today_col = date_start_col + (today - start).days
        for row in range(2, 5 + len(tasks)):
            cell = ws.cell(row=row, column=today_col)
            cell.border = Border(left=Side(style='thick', color='FF0000'))
    
    # 列幅
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    
    # 行の高さ
    for row in range(5, 5 + len(tasks)):
        ws.row_dimensions[row].height = 22
    
    return ws


def create_progress_sheet(wb, tasks):
    """進捗サマリーシート作成"""
    ws = wb.create_sheet("進捗サマリー")
    
    # タイトル
    ws['B2'] = "進捗サマリー"
    ws['B2'].font = Font(name=FONT_NAME, bold=True, size=20, color="2E5090")
    
    # ステータス別集計（Excel関数で実装）
    ws['B4'] = "ステータス別集計"
    ws['B4'].font = SECTION_FONT
    
    statuses = ['未着手', '進行中', '完了', '保留']
    ws['B5'] = "ステータス"
    ws['C5'] = "件数"
    ws['B5'].font = HEADER_FONT
    ws['C5'].font = HEADER_FONT
    ws['B5'].fill = HEADER_FILL
    ws['C5'].fill = HEADER_FILL
    
    task_count = len(tasks)
    for i, status in enumerate(statuses, start=6):
        ws[f'B{i}'] = status
        ws[f'C{i}'] = f'=COUNTIF(タスク一覧!$G$5:$G${4+task_count},B{i})'
    
    # 合計行
    ws['B10'] = "合計"
    ws['B10'].font = Font(name=FONT_NAME, bold=True)
    ws['C10'] = f'=SUM(C6:C9)'
    ws['C10'].font = Font(name=FONT_NAME, bold=True)
    
    # 完了率
    ws['B12'] = "完了率"
    ws['B12'].font = SECTION_FONT
    ws['C12'] = '=IF(C10>0,C8/C10,0)'
    ws['C12'].number_format = '0%'
    ws['C12'].font = Font(name=FONT_NAME, bold=True, size=24, color="2E5090")
    
    # 円グラフ
    pie = PieChart()
    pie.title = "ステータス別構成"
    labels = Reference(ws, min_col=2, min_row=6, max_row=9)
    data = Reference(ws, min_col=3, min_row=5, max_row=9)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 12
    pie.height = 10
    ws.add_chart(pie, "E4")
    
    # 担当者別タスク数
    ws['B15'] = "担当者別タスク数"
    ws['B15'].font = SECTION_FONT
    
    assignees = list(set([t['assignee'] for t in tasks]))
    ws['B16'] = "担当者"
    ws['C16'] = "件数"
    ws['B16'].font = HEADER_FONT
    ws['C16'].font = HEADER_FONT
    ws['B16'].fill = HEADER_FILL
    ws['C16'].fill = HEADER_FILL
    
    for i, assignee in enumerate(assignees, start=17):
        ws[f'B{i}'] = assignee
        ws[f'C{i}'] = f'=COUNTIF(タスク一覧!$D$5:$D${4+task_count},B{i})'
    
    # 担当者別棒グラフ
    bar = BarChart()
    bar.type = "col"
    bar.title = "担当者別タスク数"
    data = Reference(ws, min_col=3, min_row=16, max_row=16+len(assignees))
    cats = Reference(ws, min_col=2, min_row=17, max_row=16+len(assignees))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.width = 12
    bar.height = 8
    ws.add_chart(bar, "E15")
    
    # 列幅
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    
    return ws


def generate_meeting_tasksheet(meeting_info, summary, decisions, tasks, next_meeting, output_file, start_date, end_date):
    """メイン処理"""
    wb = Workbook()
    
    # デフォルトシート削除
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 各シート作成
    create_minutes_sheet(wb, meeting_info, summary, decisions, next_meeting)
    create_task_list_sheet(wb, tasks)
    create_gantt_chart_sheet(wb, tasks, start_date, end_date)
    create_progress_sheet(wb, tasks)
    
    # 保存
    wb.save(output_file)
    print(f"Generated: {output_file}")
    
    return output_file


# デモ用サンプルデータ
def create_demo_data():
    """MIRAICHIセミナー打ち合わせのデモデータ"""
    
    meeting_info = {
        "title": "Claude Agent Skillsセミナー 企画打ち合わせ",
        "datetime": "2025年1月15日（水）14:00-15:00",
        "location": "MIRAICHIオフィス会議室",
        "attendees": "MIRAIさん、ダイチさん、Ken"
    }
    
    summary = """2月20日開催のClaudeAgentSkillsセミナーについて企画を協議。
ターゲットは地方企業のDX担当者・経営者。Excel使用経験はあるが生成AI未経験の層を想定。
構成は基礎説明1時間＋デモ・ワークショップ2時間の計3時間。
デモは売上分析Excel自動生成、議事録→タスク管理表変換、シフト表自動生成の3本立て。
定員30名、最低15名で開催。参加費無料で個別相談へつなげる方針。"""
    
    decisions = [
        "開催日: 2025年2月20日（木）14:00-17:00",
        "ターゲット: 地方企業のDX担当者・経営者",
        "構成: 基礎説明1時間 + デモ・ワークショップ2時間",
        "デモ内容: ①売上分析Excel ②議事録→タスク管理 ③シフト表自動生成",
        "定員: 30名（最低15名で開催決定）",
        "参加費: 無料（後日個別相談へ誘導）",
        "リハーサル: 2月15日 14:00-16:00"
    ]
    
    next_meeting = "2025年1月29日（水）14:00- 進捗確認MTG"
    
    tasks = [
        {
            "name": "会場予約（2/20）",
            "assignee": "MIRAIさん",
            "start_date": datetime(2025, 1, 15),
            "due_date": datetime(2025, 1, 20),
            "status": "未着手",
            "priority": "高",
            "note": "午後14:00-17:00で確保"
        },
        {
            "name": "告知チラシ作成",
            "assignee": "MIRAIさん",
            "start_date": datetime(2025, 1, 16),
            "due_date": datetime(2025, 1, 25),
            "status": "未着手",
            "priority": "中",
            "note": "A4片面"
        },
        {
            "name": "参加者募集ページ作成",
            "assignee": "ダイチさん",
            "start_date": datetime(2025, 1, 16),
            "due_date": datetime(2025, 1, 25),
            "status": "未着手",
            "priority": "高",
            "note": "Googleフォーム使用"
        },
        {
            "name": "デモ用サンプルデータ準備",
            "assignee": "ダイチさん",
            "start_date": datetime(2025, 1, 20),
            "due_date": datetime(2025, 1, 25),
            "status": "未着手",
            "priority": "中",
            "note": "売上CSV、会議メモ等"
        },
        {
            "name": "スライド資料作成",
            "assignee": "Ken",
            "start_date": datetime(2025, 1, 20),
            "due_date": datetime(2025, 2, 10),
            "status": "未着手",
            "priority": "高",
            "note": "基礎説明パート用"
        },
        {
            "name": "デモ環境準備",
            "assignee": "Ken",
            "start_date": datetime(2025, 1, 25),
            "due_date": datetime(2025, 2, 10),
            "status": "未着手",
            "priority": "高",
            "note": "Agent Skills設定確認"
        },
        {
            "name": "リハーサル実施",
            "assignee": "全員",
            "start_date": datetime(2025, 2, 15),
            "due_date": datetime(2025, 2, 15),
            "status": "未着手",
            "priority": "高",
            "note": "本番同様の流れで",
            "is_milestone": True
        },
        {
            "name": "セミナー本番",
            "assignee": "全員",
            "start_date": datetime(2025, 2, 20),
            "due_date": datetime(2025, 2, 20),
            "status": "未着手",
            "priority": "高",
            "note": "マイルストーン",
            "is_milestone": True
        }
    ]
    
    return meeting_info, summary, decisions, tasks, next_meeting


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="会議→タスク管理表Excel生成（シンプル版）")
    parser.add_argument("--output", default="meeting_tasks.xlsx", help="出力Excelファイル")
    parser.add_argument("--start-date", default="2025-01-15", help="ガントチャート開始日")
    parser.add_argument("--end-date", default="2025-02-25", help="ガントチャート終了日")
    parser.add_argument("--demo", action="store_true", help="デモデータで生成")
    
    args = parser.parse_args()
    
    if args.demo:
        meeting_info, summary, decisions, tasks, next_meeting = create_demo_data()
        generate_meeting_tasksheet(
            meeting_info, summary, decisions, tasks, next_meeting,
            args.output, args.start_date, args.end_date
        )
