#!/usr/bin/env python3
"""Convert Ondotori (T&D) CSV exports to a formatted Excel workbook."""

import argparse
import io
import re
import sys
from pathlib import Path
from typing import Dict, Set

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


DATETIME_VARIANTS = ["日時", "測定日時", "Date/Time", "DateTime"]
TEMP_VARIANTS = ["温度", "温度(℃)", "温度(°C)", "CH1", "Temperature"]
HUMIDITY_VARIANTS = ["湿度", "湿度(%RH)", "湿度(%)", "CH2", "Humidity"]

ENCODINGS = ["utf-8-sig", "shift-jis", "cp932", "utf-8"]

# Values starting with these characters are treated as Excel formulas — sanitize them.
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def read_ondotori_csv(path: Path) -> pd.DataFrame:
    """Read an Ondotori CSV, skipping device info header rows."""
    raw = None
    for encoding in ENCODINGS:
        try:
            candidate = path.read_text(encoding=encoding)
            raw = candidate
            break
        except (UnicodeDecodeError, LookupError):
            continue

    if raw is None:
        raise ValueError(
            f"Could not read {path.name}. "
            "Supported encodings: Shift-JIS, UTF-8, CP932."
        )

    lines = raw.splitlines()

    # Find the data header row (contains a known datetime column name)
    header_idx = None
    for i, line in enumerate(lines):
        if any(col in line for col in DATETIME_VARIANTS):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(
            f"Could not find data header row in {path.name}. "
            f"Expected one of these column names: {DATETIME_VARIANTS}"
        )

    data_text = "\n".join(lines[header_idx:])
    df = pd.read_csv(io.StringIO(data_text), on_bad_lines="warn")
    return df


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns to standard names."""
    rename_map = {}
    for col in df.columns:
        if col in DATETIME_VARIANTS:
            rename_map[col] = "日時"
        elif col in TEMP_VARIANTS:
            rename_map[col] = "温度(℃)"
        elif col in HUMIDITY_VARIANTS:
            rename_map[col] = "湿度(%)"
    return df.rename(columns=rename_map)


def select_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only the relevant columns in order, sorted by datetime."""
    wanted = ["日時", "温度(℃)", "湿度(%)"]
    available = [c for c in wanted if c in df.columns]
    missing = [c for c in wanted if c not in df.columns]
    if missing:
        print(f"  Warning: column(s) not found and skipped: {missing}")
    df = df[available].dropna(how="all")
    if "日時" in df.columns:
        df["日時"] = pd.to_datetime(df["日時"], errors="coerce")
        df = df.sort_values("日時").reset_index(drop=True)
        df["日時"] = df["日時"].dt.strftime("%Y/%m/%d %H:%M")
    return df


def sanitize_value(value: object) -> object:
    """Prevent Excel formula injection by prefixing dangerous strings with a single quote."""
    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


def safe_sheet_name(name: str, existing: Set[str]) -> str:
    """Return a valid, unique Excel sheet name (max 31 chars)."""
    name = re.sub(r'[\\/*?:\[\]]', "_", name)[:31]
    candidate = name
    suffix = 2
    while candidate in existing:
        candidate = f"{name[:28]}_{suffix}"
        suffix += 1
    return candidate


def column_width(df: pd.DataFrame, col_name: str) -> int:
    """Calculate display width based on the widest value in the column."""
    max_data_len = df[col_name].astype(str).str.len().max() if len(df) > 0 else 0
    return max(len(col_name), int(max_data_len), 12) + 2


def write_excel(sheets: Dict[str, pd.DataFrame], output_path: Path) -> None:
    """Write one sheet per sensor into a formatted Excel workbook."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    header_font = Font(bold=True, name="Arial")
    header_fill = PatternFill("solid", start_color="D9E1F2")
    center = Alignment(horizontal="center", vertical="center")

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name)

        # Write header
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Write data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=sanitize_value(value))

        # Fit column widths to actual data content
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = column_width(df, col_name)

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert Ondotori CSV files to a formatted Excel workbook."
    )
    parser.add_argument("csv_files", nargs="+", type=Path, help="Ondotori CSV file(s)")
    parser.add_argument(
        "--output", "-o", type=Path, default=Path("ondotori_data.xlsx"),
        help="Output Excel file path (default: ondotori_data.xlsx)",
    )
    args = parser.parse_args()

    sheets: Dict[str, pd.DataFrame] = {}
    used_names: Set[str] = set()

    for csv_path in args.csv_files:
        if not csv_path.exists():
            print(f"Error: file not found: {csv_path}")
            sys.exit(1)

        print(f"Reading {csv_path.name}...")
        try:
            df = read_ondotori_csv(csv_path)
        except ValueError as e:
            print(f"Error: {e}")
            sys.exit(1)
        df = normalize_columns(df)
        df = select_columns(df)

        sheet_name = safe_sheet_name(csv_path.stem, used_names)
        used_names.add(sheet_name)
        sheets[sheet_name] = df
        print(f"  -> Sheet '{sheet_name}': {len(df)} rows")

    print(f"Writing {args.output}...")
    write_excel(sheets, args.output)
    print(f"Done. {len(sheets)} sheet(s) created in {args.output}")


if __name__ == "__main__":
    main()
