#!/usr/bin/env python3
"""
ヒアリング文字起こし → Excel自動記入スクリプト

文字起こしテキストを解析し、ヒアリングシートに自動記入する。
確信度に応じてセル色分け、低確信項目はコメントで補足。
"""

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment

# 既存スクリプトからインポート
from create_hearing_sheet import (
    create_hearing_sheet,
    INDUSTRIES,
    DEPARTMENTS,
    STYLES
)


# =============================================================================
# 設定
# =============================================================================

# 確信度に応じたセル背景色
CONFIDENCE_FILLS = {
    'high': None,  # 白（デフォルト）
    'medium': PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid'),  # 薄黄
    'low': PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid'),  # 薄オレンジ
    'none': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),  # 薄グレー
}

# フィールドマッピングのパス
FIELD_MAPPINGS_PATH = Path(__file__).parent.parent / 'references' / 'field_mappings.json'


# =============================================================================
# 解析用プロンプト生成
# =============================================================================

def generate_extraction_prompt(
    transcript: str,
    industry: str,
    departments: list,
    field_mappings: dict
) -> str:
    """
    文字起こし解析用のプロンプトを生成
    
    Args:
        transcript: 文字起こしテキスト
        industry: 業種
        departments: 対象部門リスト
        field_mappings: フィールド定義
    
    Returns:
        解析用プロンプト文字列
    """
    
    # 対象シートのフィールド情報を抽出
    target_sheets = ['基本情報', 'IT環境', '優先度・ゴール整理'] + departments
    
    fields_description = []
    for sheet_name in target_sheets:
        if sheet_name in field_mappings.get('sheets', {}):
            sheet_info = field_mappings['sheets'][sheet_name]
            fields = sheet_info.get('fields', {})
            issues = sheet_info.get('issues', {})
            
            field_list = []
            for field_id, field_def in fields.items():
                field_desc = f"  - {field_id}: {field_def['label']}"
                if 'options' in field_def:
                    field_desc += f" (選択肢: {', '.join(field_def['options'])})"
                field_list.append(field_desc)
            
            if field_list:
                fields_description.append(f"\n### {sheet_name}\n" + "\n".join(field_list))
            
            if issues:
                issue_list = list(issues.get('keywords_map', {}).keys())
                if issue_list:
                    fields_description.append(f"  - issues: 課題 (選択肢: {', '.join(issue_list[:5])}...)")
    
    prompt = f"""あなたはヒアリング内容を分析し、構造化データに変換するアシスタントです。

以下の文字起こしテキストを分析し、ヒアリングシートの各項目に該当する情報を抽出してください。

## 対象企業情報
- 業種: {industry}
- 対象部門: {', '.join(departments)}

## 抽出対象フィールド
{''.join(fields_description)}

## 文字起こしテキスト
---
{transcript}
---

## 抽出ルール

1. **確信度の判定基準**
   - high: 明確な固有名詞、具体的な数値、断定的な発言がある
   - medium: 文脈から推測可能だが、曖昧な表現や不確かな発言
   - low: 複数の解釈が可能、または断片的な情報のみ
   - none: 該当する発言が見つからない

2. **選択肢のマッチング**
   - 発言内容を最も近い選択肢にマッチングする
   - 表記ゆれを吸収する（例: 「フリー」→「freee」、「エクセル」→「Excel」）

3. **課題（issues）の抽出**
   - 困っている、大変、時間がかかる、などのネガティブな発言を検出
   - 該当する課題項目を配列で返す

4. **エピソード（episode）の抽出**
   - 具体的なエピソード、数値を含む発言を抽出
   - 要約せず、できるだけ原文に近い形で記録

## 出力形式

以下のJSON形式で出力してください。JSONのみを出力し、他の説明は不要です。

```json
{{
  "basic_info": {{
    "company_name": {{"value": "値またはnull", "confidence": "high|medium|low|none", "note": "補足またはnull"}},
    "employees": {{"value": "45", "confidence": "high", "note": null}}
  }},
  "it_environment": {{
    "accounting_software": {{"value": "freee", "confidence": "high", "note": null}}
  }},
  "departments": {{
    "経理": {{
      "staff_count": {{"value": "2", "confidence": "high", "note": null}},
      "issues": {{"selected": ["課題1", "課題2"], "confidence": "high", "note": null}},
      "episode": {{"value": "具体的なエピソード", "confidence": "high", "note": null}}
    }}
  }},
  "priority": {{
    "priority_1": {{"value": "最優先課題", "confidence": "high", "note": null}},
    "budget": {{"value": "100〜300万円", "confidence": "medium", "note": "補足説明"}}
  }}
}}
```

該当する発言がない項目は省略してください。
"""
    return prompt


# =============================================================================
# 解析結果の適用
# =============================================================================

def apply_extraction_to_excel(
    wb: Workbook,
    extraction_result: dict,
    field_mappings: dict
) -> Workbook:
    """
    解析結果をExcelに適用
    
    Args:
        wb: 対象のWorkbook
        extraction_result: 解析結果JSON
        field_mappings: フィールド定義
    
    Returns:
        更新されたWorkbook
    """
    
    # 基本情報シート
    if '基本情報' in wb.sheetnames and 'basic_info' in extraction_result:
        ws = wb['基本情報']
        apply_fields_to_sheet(ws, extraction_result['basic_info'], 
                              field_mappings['sheets'].get('基本情報', {}).get('fields', {}))
    
    # IT環境シート
    if 'IT環境' in wb.sheetnames and 'it_environment' in extraction_result:
        ws = wb['IT環境']
        apply_fields_to_sheet(ws, extraction_result['it_environment'],
                              field_mappings['sheets'].get('IT環境', {}).get('fields', {}))
    
    # 部門別シート
    if 'departments' in extraction_result:
        for dept_name, dept_data in extraction_result['departments'].items():
            if dept_name in wb.sheetnames:
                ws = wb[dept_name]
                dept_fields = field_mappings['sheets'].get(dept_name, {}).get('fields', {})
                apply_fields_to_sheet(ws, dept_data, dept_fields)
                
                # 課題の適用
                if 'issues' in dept_data:
                    apply_issues_to_sheet(ws, dept_data['issues'], dept_name)
                
                # エピソードの適用
                if 'episode' in dept_data:
                    apply_episode_to_sheet(ws, dept_data['episode'])
    
    # 優先度・ゴール整理シート
    if '優先度・ゴール整理' in wb.sheetnames and 'priority' in extraction_result:
        ws = wb['優先度・ゴール整理']
        apply_fields_to_sheet(ws, extraction_result['priority'],
                              field_mappings['sheets'].get('優先度・ゴール整理', {}).get('fields', {}))
    
    return wb


def apply_fields_to_sheet(ws, data: dict, field_defs: dict):
    """
    フィールドデータをシートに適用
    """
    for field_id, field_data in data.items():
        if field_id in ['issues', 'episode']:
            continue  # 別処理
        
        if field_id not in field_defs:
            continue
        
        field_def = field_defs[field_id]
        cell_ref = field_def.get('cell')
        
        if not cell_ref or not isinstance(field_data, dict):
            continue
        
        value = field_data.get('value')
        confidence = field_data.get('confidence', 'none')
        note = field_data.get('note')
        
        if value is None:
            continue
        
        # セルに値を設定（マージセルの場合はスキップ）
        cell = ws[cell_ref]
        try:
            cell.value = value
            
            # 確信度に応じた背景色
            fill = CONFIDENCE_FILLS.get(confidence)
            if fill:
                cell.fill = fill
            
            # 補足コメント
            if note and confidence in ['medium', 'low']:
                cell.comment = Comment(note, 'Auto-Fill')
        except AttributeError:
            # マージセルの場合は動的検索で再試行
            found_cell = find_cell_by_label(ws, field_def.get('label', ''))
            if found_cell:
                target_cell = ws[found_cell]
                try:
                    target_cell.value = value
                    fill = CONFIDENCE_FILLS.get(confidence)
                    if fill:
                        target_cell.fill = fill
                    if note and confidence in ['medium', 'low']:
                        target_cell.comment = Comment(note, 'Auto-Fill')
                except AttributeError:
                    pass  # スキップ


def apply_issues_to_sheet(ws, issues_data: dict, dept_name: str):
    """
    課題チェックボックスをシートに適用
    
    実際のシートでは課題がチェックボックス形式のため、
    該当する課題の行を見つけて「■」に置換する。
    """
    selected = issues_data.get('selected', [])
    confidence = issues_data.get('confidence', 'none')
    note = issues_data.get('note')
    
    if not selected:
        return
    
    # 困っていることセクションを探す
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and '□' in str(cell_value):
            # チェックボックス行
            for issue in selected:
                if issue in str(cell_value):
                    # □を■に置換
                    new_value = str(cell_value).replace('□', '■', 1)
                    ws.cell(row=row, column=1).value = new_value
                    
                    # 確信度に応じた背景色
                    fill = CONFIDENCE_FILLS.get(confidence)
                    if fill:
                        ws.cell(row=row, column=1).fill = fill
                    
                    break


def apply_episode_to_sheet(ws, episode_data: dict):
    """
    エピソードをシートに適用
    
    「具体的なエピソード」セクションを探して記入
    """
    value = episode_data.get('value')
    confidence = episode_data.get('confidence', 'none')
    note = episode_data.get('note')
    
    if not value:
        return
    
    # エピソードセクションを探す
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and '具体的なエピソード' in str(cell_value):
            # 次の行に記入
            target_row = row + 1
            ws.cell(row=target_row, column=1).value = value
            
            # 確信度に応じた背景色
            fill = CONFIDENCE_FILLS.get(confidence)
            if fill:
                for col in range(1, 5):
                    ws.cell(row=target_row, column=col).fill = fill
            
            # 補足コメント
            if note:
                ws.cell(row=target_row, column=1).comment = Comment(note, 'Auto-Fill')
            
            break


# =============================================================================
# セル位置の動的検索
# =============================================================================

def find_cell_by_label(ws, label: str) -> Optional[str]:
    """
    ラベルに対応する入力セルを検索
    
    Args:
        ws: ワークシート
        label: 検索するラベル
    
    Returns:
        セル参照（例: "B5"）またはNone
    """
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and label in str(cell_value):
            # 隣のセル（入力欄）を返す
            return f"B{row}"
    return None


def update_field_cell_refs(ws, field_defs: dict) -> dict:
    """
    シートの実際の構造に基づいてセル参照を更新
    
    Args:
        ws: ワークシート
        field_defs: フィールド定義
    
    Returns:
        更新されたフィールド定義
    """
    updated_defs = field_defs.copy()
    
    for field_id, field_def in updated_defs.items():
        label = field_def.get('label')
        if label:
            found_cell = find_cell_by_label(ws, label)
            if found_cell:
                updated_defs[field_id]['cell'] = found_cell
    
    return updated_defs


# =============================================================================
# メイン処理
# =============================================================================

def fill_hearing_sheet(
    company_name: str,
    industry: str,
    departments: list,
    transcript: str,
    input_file: str = None,
    output_path: str = None,
    extraction_result: dict = None
) -> str:
    """
    文字起こしからヒアリングシートに自動記入
    
    Args:
        company_name: 企業名
        industry: 業種
        departments: 対象部門リスト
        transcript: 文字起こしテキスト
        input_file: 既存のExcelファイル（省略時は新規生成）
        output_path: 出力パス
        extraction_result: 解析済み結果（省略時はプロンプトを出力）
    
    Returns:
        生成されたファイルパス
    """
    
    # フィールドマッピングを読み込み
    with open(FIELD_MAPPINGS_PATH, 'r', encoding='utf-8') as f:
        field_mappings = json.load(f)
    
    # Excelファイルの準備
    if input_file:
        wb = load_workbook(input_file)
    else:
        # 新規シート生成
        temp_path = f"/tmp/hearing_sheet_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        create_hearing_sheet(
            company_name=company_name,
            industry=industry,
            departments=departments,
            output_path=temp_path
        )
        wb = load_workbook(temp_path)
    
    # 解析結果がない場合はプロンプトを出力
    if extraction_result is None:
        prompt = generate_extraction_prompt(
            transcript=transcript,
            industry=industry,
            departments=departments,
            field_mappings=field_mappings
        )
        print("=" * 60)
        print("以下のプロンプトをClaudeに送信して、JSON結果を取得してください:")
        print("=" * 60)
        print(prompt)
        print("=" * 60)
        print("\n取得したJSONを --extraction-json オプションで指定して再実行してください。")
        return None
    
    # 各シートのセル参照を動的に更新
    for sheet_name in field_mappings['sheets']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            field_mappings['sheets'][sheet_name]['fields'] = update_field_cell_refs(
                ws, field_mappings['sheets'][sheet_name].get('fields', {})
            )
    
    # 解析結果をExcelに適用
    wb = apply_extraction_to_excel(wb, extraction_result, field_mappings)
    
    # 出力パスの決定
    if output_path is None:
        date_str = datetime.now().strftime("%Y%m%d")
        output_path = f"/mnt/user-data/outputs/ヒアリングシート_{company_name}_{date_str}_記入済.xlsx"
    
    # 保存
    output_dir = Path(output_path).parent
    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ 記入済みヒアリングシートを生成しました: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description='ヒアリング文字起こし → Excel自動記入')
    parser.add_argument('--company', required=True, help='ヒアリング先企業名')
    parser.add_argument('--industry', required=True, choices=INDUSTRIES, help='業種')
    parser.add_argument('--departments', nargs='+', required=True, help='対象部門')
    parser.add_argument('--transcript', required=True, help='文字起こしファイルパス')
    parser.add_argument('--input', help='既存のExcelファイル（省略時は新規生成）')
    parser.add_argument('--output', help='出力ファイルパス')
    parser.add_argument('--extraction-json', help='解析結果JSONファイルパス')
    
    args = parser.parse_args()
    
    # 文字起こしを読み込み
    with open(args.transcript, 'r', encoding='utf-8') as f:
        transcript = f.read()
    
    # 解析結果を読み込み（指定されている場合）
    extraction_result = None
    if args.extraction_json:
        with open(args.extraction_json, 'r', encoding='utf-8') as f:
            extraction_result = json.load(f)
    
    fill_hearing_sheet(
        company_name=args.company,
        industry=args.industry,
        departments=args.departments,
        transcript=transcript,
        input_file=args.input,
        output_path=args.output,
        extraction_result=extraction_result
    )


if __name__ == '__main__':
    main()
