#!/usr/bin/env python3
"""
ヒアリングシート生成スクリプト
業種・部門に応じた動的なヒアリングシートをExcel形式で生成
"""

import argparse
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# =============================================================================
# 設定
# =============================================================================

# 業種リスト
INDUSTRIES = [
    "製造業",
    "小売業",
    "卸売業",
    "建設業",
    "福祉・介護",
    "医療",
    "金融・保険",
    "不動産",
    "運輸・物流",
    "飲食・宿泊",
    "情報通信・IT",
    "その他サービス業",
    "農林水産業",
    "公共・団体"
]

# 部門リスト
DEPARTMENTS = [
    "経理",
    "人事・総務",
    "営業",
    "現場・製造",
    "情報システム"
]

# スタイル定義
STYLES = {
    'title_font': Font(name='メイリオ', size=16, bold=True, color='000000'),
    'section_font': Font(name='メイリオ', size=12, bold=True, color='666666'),
    'label_font': Font(name='メイリオ', size=10, color='333333'),
    'input_font': Font(name='メイリオ', size=10, color='000000'),
    'header_fill': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
    'input_fill': PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid'),
    'thin_border': Border(
        bottom=Side(style='thin', color='E0E0E0')
    ),
    'box_border': Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
}


# =============================================================================
# 業種別追加質問
# =============================================================================

INDUSTRY_QUESTIONS = {
    "製造業": {
        "現場・製造": [
            ("生産管理システム", ["なし", "Excel", "専用システム", "ERP"]),
            ("生産方式", ["見込生産", "受注生産", "混合"]),
            ("設備保全", ["事後保全", "定期保全", "予知保全"]),
            ("品質管理方式", ["抜き取り", "全数検査", "自動検査"]),
            ("IoT導入状況", ["未導入", "一部導入", "全面導入"]),
            ("図面管理", ["紙", "ファイルサーバー", "PDM/PLM"]),
        ],
        "経理": [
            ("原価計算", ["していない", "簡易", "標準原価", "実際原価"]),
            ("在庫評価", ["総平均法", "移動平均法", "先入先出法"]),
        ],
        "common_issues": [
            "生産計画と実績の乖離が大きい",
            "設備の突発故障で生産が止まる",
            "原価管理・原価計算ができていない",
            "多品種少量生産への対応が難しい",
            "熟練工のノウハウが継承できない",
            "在庫の過不足が発生する",
            "品質トラブルの原因追跡が難しい",
            "紙の帳票が多く二重入力が発生",
        ]
    },
    "小売業": {
        "営業": [
            ("POSシステム", "text"),
            ("POS連携", ["なし", "売上データのみ", "在庫連動"]),
            ("EC展開", ["なし", "検討中", "運用中"]),
            ("顧客分析", ["していない", "Excel", "BIツール"]),
            ("ポイントカード", ["なし", "紙", "アプリ", "外部サービス"]),
        ],
        "現場・製造": [
            ("店舗数", "number"),
            ("発注方式", ["手動", "自動発注", "本部一括"]),
            ("棚卸頻度", ["月次", "四半期", "年次"]),
        ],
        "common_issues": [
            "在庫の過不足が発生する",
            "顧客データの活用ができていない",
            "実店舗とECの連携ができていない",
            "人手不足・シフト管理が大変",
            "売れ筋・死に筋の分析ができていない",
            "店舗間の情報共有ができていない",
            "キャッシュレス対応の管理が煩雑",
        ]
    },
    "福祉・介護": {
        "all": [
            ("介護ソフト", "text"),
            ("記録方式", ["紙", "PC入力", "タブレット"]),
            ("国保連請求", ["手作業", "ソフト連携"]),
            ("シフト作成", ["紙", "Excel", "専用ソフト"]),
            ("サービス種別", ["特養", "老健", "グループホーム", "デイサービス", "訪問介護", "居宅介護支援", "障害福祉", "その他"]),
            ("利用者数", "number"),
        ],
        "common_issues": [
            "記録業務の負担が大きい",
            "請求ミス・返戻が多い",
            "シフト作成に時間がかかる",
            "職員の定着率が低い",
            "情報共有（申し送り）が不十分",
            "ケアプラン作成の負荷が高い",
            "監査対応の書類準備が大変",
            "ヒヤリハット・事故報告の管理",
        ]
    },
    "金融・保険": {
        "all": [
            ("コンプライアンス体制", ["整備中", "運用中", "外部監査済"]),
            ("顧客対応記録", ["紙", "Excel", "CRM"]),
            ("本人確認", ["紙", "eKYC対応"]),
        ],
        "営業": [
            ("提案書作成", ["手作業", "Excel", "専用ツール"]),
            ("契約管理", ["紙", "Excel", "システム"]),
        ],
        "common_issues": [
            "顧客対応履歴の管理が煩雑",
            "コンプライアンス対応の負荷が高い",
            "帳票・報告書作成の手間",
            "顧客情報のセキュリティ管理",
            "複数システムへの二重入力",
            "高齢顧客への対応",
            "契約更新・満期管理",
        ]
    },
    "公共・団体": {
        "all": [
            ("組合員/会員管理", ["紙", "Excel", "専用システム"]),
            ("組合員/会員数", "number"),
            ("事業部門数", "number"),
        ],
        "経理": [
            ("共済管理", ["紙", "Excel", "専用システム"]),
            ("融資管理", ["紙", "Excel", "専用システム"]),
            ("購買事業", ["紙", "Excel", "専用システム"]),
        ],
        "common_issues": [
            "組合員/会員情報の一元管理ができていない",
            "複数事業のシステムが分散している",
            "報告書・統計作成の負荷が高い",
            "高齢化による人手不足",
            "若手職員の育成",
            "組合員/会員向けサービスのデジタル化",
            "本所・支所間の情報共有",
        ]
    }
}

# デフォルトの部門別質問（業種に関係なく共通）
DEFAULT_DEPARTMENT_QUESTIONS = {
    "経理": {
        "organization": [
            ("経理担当者数", "number"),
            ("決算期", "number"),
        ],
        "workflow": [
            ("請求書処理", ["紙", "Excel", "システム", "電子"]),
            ("経費精算", ["紙", "Excel", "システム"]),
            ("支払処理", ["手動", "半自動", "自動"]),
            ("月次決算所要日数", "number"),
        ],
        "issues": [
            "請求書の手入力が多い",
            "経費精算の承認フローが煩雑",
            "月次決算に時間がかかる",
            "部門間のデータ連携ができていない",
            "紙の証憑管理が大変",
            "インボイス制度対応",
            "電子帳簿保存法対応",
        ]
    },
    "人事・総務": {
        "organization": [
            ("人事担当者数", "number"),
            ("年間採用数", "number"),
            ("年間退職数", "number"),
        ],
        "workflow": [
            ("勤怠管理", ["紙", "Excel", "タイムカード", "システム"]),
            ("給与計算", ["手計算", "Excel", "ソフト", "外注"]),
            ("入退社手続", ["紙中心", "半デジタル", "デジタル"]),
            ("評価制度", ["なし", "紙", "Excel", "システム"]),
        ],
        "issues": [
            "勤怠集計に時間がかかる",
            "有給残日数の管理が煩雑",
            "入退社時の手続きが属人化",
            "採用管理ができていない",
            "人事評価の運用が形骸化",
            "社会保険手続きが大変",
            "年末調整が紙で大変",
        ]
    },
    "営業": {
        "organization": [
            ("営業担当者数", "number"),
            ("顧客数", "number"),
        ],
        "workflow": [
            ("顧客管理", ["名刺", "Excel", "CRM", "その他"]),
            ("見積作成", ["手書き", "Excel", "システム"]),
            ("受注管理", ["紙", "Excel", "システム"]),
            ("日報・週報", ["なし", "紙", "Excel", "システム"]),
        ],
        "issues": [
            "顧客情報が属人化している",
            "見積作成に時間がかかる",
            "商談履歴が共有できていない",
            "売上予測が立てられない",
            "日報が形骸化している",
            "営業ノウハウが共有できていない",
        ]
    },
    "現場・製造": {
        "organization": [
            ("現場作業者数", "number"),
            ("シフト体制", ["日勤のみ", "2交代", "3交代"]),
        ],
        "workflow": [
            ("作業指示", ["口頭", "紙", "ホワイトボード", "システム"]),
            ("実績入力", ["紙", "Excel", "システム"]),
            ("在庫管理", ["目視", "紙", "Excel", "システム"]),
            ("品質管理", ["紙", "Excel", "システム"]),
        ],
        "issues": [
            "作業指示の伝達ミスがある",
            "実績入力が二重になっている",
            "在庫数が合わない",
            "設備の稼働状況が見えない",
            "品質トラブルの原因追跡が難しい",
            "ベテランのノウハウが継承できない",
        ]
    },
    "情報システム": {
        "organization": [
            ("IT担当者数（専任）", "number"),
            ("IT担当者数（兼任）", "number"),
        ],
        "workflow": [
            ("PC管理", ["台帳なし", "Excel", "資産管理ツール"]),
            ("アカウント管理", ["手動", "半自動", "ID管理システム"]),
            ("ヘルプデスク", ["口頭", "メール", "チケット管理"]),
            ("バックアップ", ["なし", "手動", "自動"]),
        ],
        "issues": [
            "問い合わせ対応に追われている",
            "IT資産の把握ができていない",
            "入退社時のアカウント管理が煩雑",
            "セキュリティ対策が不十分",
            "システム間の連携ができていない",
            "属人化で担当者不在時に困る",
        ]
    }
}


# =============================================================================
# ヘルパー関数
# =============================================================================

def apply_style(cell, font=None, fill=None, border=None, alignment=None):
    """セルにスタイルを適用"""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment


def add_section_header(ws, row, text, col_start=1, col_end=6):
    """セクションヘッダーを追加"""
    cell = ws.cell(row=row, column=col_start, value=f"■ {text}")
    apply_style(cell, font=STYLES['section_font'])
    ws.row_dimensions[row].height = 25
    return row + 1


def add_input_row(ws, row, label, col_start=1, input_width=3):
    """ラベル + 入力欄の行を追加"""
    # ラベル
    label_cell = ws.cell(row=row, column=col_start, value=label)
    apply_style(label_cell, font=STYLES['label_font'])
    
    # 入力欄
    input_cell = ws.cell(row=row, column=col_start + 1)
    apply_style(input_cell, 
                font=STYLES['input_font'],
                fill=STYLES['input_fill'],
                border=STYLES['box_border'])
    
    # 入力欄をマージ
    if input_width > 1:
        ws.merge_cells(
            start_row=row, start_column=col_start + 1,
            end_row=row, end_column=col_start + input_width
        )
    
    ws.row_dimensions[row].height = 22
    return row + 1


def add_checkbox_row(ws, row, options, col_start=1):
    """チェックボックス選択肢の行を追加"""
    checkbox_text = "  ".join([f"□ {opt}" for opt in options])
    cell = ws.cell(row=row, column=col_start, value=checkbox_text)
    apply_style(cell, font=STYLES['label_font'])
    ws.row_dimensions[row].height = 22
    return row + 1


def add_labeled_checkbox_row(ws, row, label, options, col_start=1):
    """ラベル + チェックボックスの行を追加"""
    # ラベル
    label_cell = ws.cell(row=row, column=col_start, value=label)
    apply_style(label_cell, font=STYLES['label_font'])
    
    # チェックボックス
    checkbox_text = "  ".join([f"□ {opt}" for opt in options])
    checkbox_cell = ws.cell(row=row, column=col_start + 1, value=checkbox_text)
    apply_style(checkbox_cell, font=STYLES['label_font'])
    
    ws.row_dimensions[row].height = 22
    return row + 1


def add_issue_checkboxes(ws, row, issues, col_start=1):
    """課題チェックボックス群を追加"""
    for issue in issues:
        cell = ws.cell(row=row, column=col_start, value=f"□ {issue}")
        apply_style(cell, font=STYLES['label_font'])
        ws.row_dimensions[row].height = 22
        row += 1
    
    # その他
    cell = ws.cell(row=row, column=col_start, value="□ その他")
    apply_style(cell, font=STYLES['label_font'])
    other_cell = ws.cell(row=row, column=col_start + 1)
    apply_style(other_cell, 
                font=STYLES['input_font'],
                fill=STYLES['input_fill'],
                border=STYLES['box_border'])
    ws.merge_cells(start_row=row, start_column=col_start + 1, end_row=row, end_column=col_start + 3)
    ws.row_dimensions[row].height = 22
    
    return row + 1


def add_free_text_area(ws, row, label, num_lines=3, col_start=1, col_end=5):
    """自由記述エリアを追加"""
    # ラベル
    label_cell = ws.cell(row=row, column=col_start, value=f"● {label}")
    apply_style(label_cell, font=STYLES['label_font'])
    row += 1
    
    # 入力エリア
    for _ in range(num_lines):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, 
                        font=STYLES['input_font'],
                        fill=STYLES['input_fill'],
                        border=STYLES['box_border'])
        ws.row_dimensions[row].height = 25
        row += 1
    
    return row


# =============================================================================
# シート生成関数
# =============================================================================

def create_basic_info_sheet(wb, company_name, industry, hearing_date):
    """基本情報シートを作成"""
    ws = wb.active
    ws.title = "基本情報"
    
    # 列幅設定
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    row = 1
    
    # タイトル
    title_cell = ws.cell(row=row, column=1, value="ヒアリングシート")
    apply_style(title_cell, font=STYLES['title_font'])
    
    date_cell = ws.cell(row=row, column=5, value=f"日付: {hearing_date}")
    apply_style(date_cell, font=STYLES['label_font'], 
                alignment=Alignment(horizontal='right'))
    row += 1
    
    # 会社名
    company_cell = ws.cell(row=row, column=1, value=f"{company_name} 様")
    apply_style(company_cell, font=Font(name='メイリオ', size=14, bold=True))
    
    industry_cell = ws.cell(row=row, column=5, value=f"業種: {industry}")
    apply_style(industry_cell, font=STYLES['label_font'],
                alignment=Alignment(horizontal='right'))
    row += 2
    
    # 企業情報セクション
    row = add_section_header(ws, row, "企業情報")
    row = add_input_row(ws, row, "会社名")
    row = add_input_row(ws, row, "所在地")
    
    # 設立年・従業員数（横並び）
    ws.cell(row=row, column=1, value="設立年")
    apply_style(ws.cell(row=row, column=1), font=STYLES['label_font'])
    apply_style(ws.cell(row=row, column=2), 
                font=STYLES['input_font'],
                fill=STYLES['input_fill'],
                border=STYLES['box_border'])
    
    ws.cell(row=row, column=3, value="従業員数")
    apply_style(ws.cell(row=row, column=3), font=STYLES['label_font'])
    apply_style(ws.cell(row=row, column=4), 
                font=STYLES['input_font'],
                fill=STYLES['input_fill'],
                border=STYLES['box_border'])
    ws.cell(row=row, column=5, value="名")
    apply_style(ws.cell(row=row, column=5), font=STYLES['label_font'])
    ws.row_dimensions[row].height = 22
    row += 1
    
    row = add_input_row(ws, row, "事業内容")
    row = add_input_row(ws, row, "年商規模")
    row += 1
    
    # 担当者情報セクション
    row = add_section_header(ws, row, "担当者情報")
    row = add_input_row(ws, row, "部署・役職")
    row = add_input_row(ws, row, "氏名")
    row = add_input_row(ws, row, "連絡先")
    row += 1
    
    # ヒアリング目的セクション
    row = add_section_header(ws, row, "ヒアリング目的")
    purposes = [
        "初回商談・現状把握",
        "特定課題の深掘り",
        "DX/AI導入検討",
    ]
    for purpose in purposes:
        cell = ws.cell(row=row, column=1, value=f"□ {purpose}")
        apply_style(cell, font=STYLES['label_font'])
        ws.row_dimensions[row].height = 22
        row += 1
    
    # その他
    cell = ws.cell(row=row, column=1, value="□ その他")
    apply_style(cell, font=STYLES['label_font'])
    apply_style(ws.cell(row=row, column=2), 
                font=STYLES['input_font'],
                fill=STYLES['input_fill'],
                border=STYLES['box_border'])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    
    # グリッド線を非表示
    ws.sheet_view.showGridLines = False
    
    return ws


def create_it_environment_sheet(wb):
    """IT環境シートを作成"""
    ws = wb.create_sheet("IT環境")
    
    # 列幅設定
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    
    row = 1
    
    # 使用ツール・システムセクション
    row = add_section_header(ws, row, "現在の使用ツール・システム")
    
    # ヘッダー
    headers = ["カテゴリ", "ツール名", "導入年", "満足度(1-5)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, font=STYLES['label_font'], fill=STYLES['header_fill'])
    ws.row_dimensions[row].height = 22
    row += 1
    
    # ツールカテゴリ
    categories = [
        "会計ソフト",
        "勤怠管理",
        "顧客管理（CRM）",
        "グループウェア",
        "ファイル共有",
        "コミュニケーション",
        "基幹システム",
        "その他",
    ]
    
    for category in categories:
        ws.cell(row=row, column=1, value=category)
        apply_style(ws.cell(row=row, column=1), font=STYLES['label_font'])
        for col in range(2, 5):
            apply_style(ws.cell(row=row, column=col),
                        font=STYLES['input_font'],
                        fill=STYLES['input_fill'],
                        border=STYLES['box_border'])
        ws.row_dimensions[row].height = 22
        row += 1
    
    row += 1
    
    # IT環境セクション
    row = add_section_header(ws, row, "IT環境")
    
    # PC台数
    ws.cell(row=row, column=1, value="社内PC台数")
    apply_style(ws.cell(row=row, column=1), font=STYLES['label_font'])
    apply_style(ws.cell(row=row, column=2),
                font=STYLES['input_font'],
                fill=STYLES['input_fill'],
                border=STYLES['box_border'])
    ws.cell(row=row, column=3, value="台")
    apply_style(ws.cell(row=row, column=3), font=STYLES['label_font'])
    
    ws.cell(row=row, column=4, value="ノートPC比率")
    apply_style(ws.cell(row=row, column=4), font=STYLES['label_font'])
    apply_style(ws.cell(row=row, column=5),
                font=STYLES['input_font'],
                fill=STYLES['input_fill'],
                border=STYLES['box_border'])
    ws.row_dimensions[row].height = 22
    row += 1
    
    row = add_labeled_checkbox_row(ws, row, "サーバー", 
                                    ["オンプレ", "クラウド", "ハイブリッド"])
    row = add_labeled_checkbox_row(ws, row, "ネットワーク",
                                    ["有線のみ", "Wi-Fi完備", "VPN利用"])
    row += 1
    
    # IT担当者セクション
    row = add_section_header(ws, row, "IT担当者")
    row = add_labeled_checkbox_row(ws, row, "専任担当",
                                    ["いる（    名）", "いない（兼任）"])
    
    ws.cell(row=row, column=1, value="外部委託")
    apply_style(ws.cell(row=row, column=1), font=STYLES['label_font'])
    ws.cell(row=row, column=2, value="□ あり")
    apply_style(ws.cell(row=row, column=2), font=STYLES['label_font'])
    apply_style(ws.cell(row=row, column=3),
                font=STYLES['input_font'],
                fill=STYLES['input_fill'],
                border=STYLES['box_border'])
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 22
    row += 1
    
    ws.cell(row=row, column=2, value="□ なし")
    apply_style(ws.cell(row=row, column=2), font=STYLES['label_font'])
    
    # グリッド線を非表示
    ws.sheet_view.showGridLines = False
    
    return ws


def create_department_sheet(wb, department, industry):
    """部門別ヒアリングシートを作成"""
    ws = wb.create_sheet(department)
    
    # 列幅設定
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    row = 1
    
    # タイトル
    title_cell = ws.cell(row=row, column=1, value=f"■ {department}部門 ヒアリング")
    apply_style(title_cell, font=STYLES['title_font'])
    row += 2
    
    # デフォルトの部門質問を取得
    dept_questions = DEFAULT_DEPARTMENT_QUESTIONS.get(department, {})
    
    # 組織体制
    if 'organization' in dept_questions:
        row = add_section_header(ws, row, "組織体制")
        for label, input_type in dept_questions['organization']:
            row = add_input_row(ws, row, label)
        row += 1
    
    # 現在の業務フロー
    if 'workflow' in dept_questions:
        row = add_section_header(ws, row, "現在の業務フロー")
        for label, options in dept_questions['workflow']:
            if isinstance(options, list):
                row = add_labeled_checkbox_row(ws, row, label, options)
            else:
                row = add_input_row(ws, row, label)
        row += 1
    
    # 業種別追加質問
    industry_q = INDUSTRY_QUESTIONS.get(industry, {})
    
    # 全部門共通の業種別質問
    if 'all' in industry_q:
        row = add_section_header(ws, row, f"【{industry}】追加項目")
        for label, options in industry_q['all']:
            if isinstance(options, list):
                row = add_labeled_checkbox_row(ws, row, label, options)
            else:
                row = add_input_row(ws, row, label)
        row += 1
    
    # 部門固有の業種別質問
    if department in industry_q:
        for label, options in industry_q[department]:
            if isinstance(options, list):
                row = add_labeled_checkbox_row(ws, row, label, options)
            else:
                row = add_input_row(ws, row, label)
        row += 1
    
    # 困っていること
    row = add_section_header(ws, row, "困っていること（複数選択可）")
    
    # デフォルトの課題
    issues = dept_questions.get('issues', [])
    
    # 業種別の課題を追加
    if 'common_issues' in industry_q:
        issues = issues + [i for i in industry_q['common_issues'] if i not in issues]
    
    row = add_issue_checkboxes(ws, row, issues)
    row += 1
    
    # 具体的なエピソード
    row = add_free_text_area(ws, row, "具体的なエピソード・数値", num_lines=3)
    
    # グリッド線を非表示
    ws.sheet_view.showGridLines = False
    
    return ws


def create_priority_sheet(wb):
    """優先度・ゴール整理シートを作成"""
    ws = wb.create_sheet("優先度・ゴール整理")
    
    # 列幅設定
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    row = 1
    
    # 課題の優先順位
    row = add_section_header(ws, row, "課題の優先順位（上位3つ）")
    for i in range(1, 4):
        ws.cell(row=row, column=1, value=f"{i}.")
        apply_style(ws.cell(row=row, column=1), font=STYLES['label_font'])
        apply_style(ws.cell(row=row, column=2),
                    font=STYLES['input_font'],
                    fill=STYLES['input_fill'],
                    border=STYLES['box_border'])
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 25
        row += 1
    row += 1
    
    # 理想の状態
    row = add_free_text_area(ws, row, "理想の状態（1年後にどうなっていたいか）", num_lines=3)
    row += 1
    
    # 予算感
    row = add_section_header(ws, row, "予算感")
    budgets = ["〜50万円", "50〜100万円", "100〜300万円", 
               "300〜500万円", "500万円以上", "未定"]
    row = add_checkbox_row(ws, row, budgets)
    row += 1
    
    # 希望スケジュール
    row = add_section_header(ws, row, "希望スケジュール")
    row = add_labeled_checkbox_row(ws, row, "着手希望時期",
                                    ["すぐ", "3ヶ月以内", "半年以内", "未定"])
    row = add_input_row(ws, row, "完了希望時期")
    row += 1
    
    # 意思決定プロセス
    row = add_section_header(ws, row, "意思決定プロセス")
    row = add_input_row(ws, row, "決裁者")
    row = add_labeled_checkbox_row(ws, row, "検討体制",
                                    ["担当者判断", "部長決裁", "役員決裁"])
    row += 1
    
    # 次のステップ
    row = add_section_header(ws, row, "次のステップ")
    next_steps = [
        "詳細ヒアリング（部門別）",
        "現地調査・業務観察",
        "提案書作成",
        "見積依頼",
    ]
    for step in next_steps:
        cell = ws.cell(row=row, column=1, value=f"□ {step}")
        apply_style(cell, font=STYLES['label_font'])
        ws.row_dimensions[row].height = 22
        row += 1
    
    # その他
    cell = ws.cell(row=row, column=1, value="□ その他")
    apply_style(cell, font=STYLES['label_font'])
    apply_style(ws.cell(row=row, column=2),
                font=STYLES['input_font'],
                fill=STYLES['input_fill'],
                border=STYLES['box_border'])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    
    # グリッド線を非表示
    ws.sheet_view.showGridLines = False
    
    return ws


# =============================================================================
# メイン関数
# =============================================================================

def create_hearing_sheet(
    company_name: str,
    industry: str,
    departments: list,
    hearing_date: str = None,
    output_path: str = None
):
    """
    ヒアリングシートを生成
    
    Args:
        company_name: ヒアリング先企業名
        industry: 業種
        departments: 対象部門リスト
        hearing_date: ヒアリング予定日（省略時は今日）
        output_path: 出力パス（省略時は自動生成）
    
    Returns:
        生成されたファイルパス
    """
    # 日付のデフォルト
    if hearing_date is None:
        hearing_date = datetime.now().strftime("%Y-%m-%d")
    
    # 出力パスのデフォルト
    if output_path is None:
        date_str = datetime.now().strftime("%Y%m%d")
        output_path = f"/mnt/user-data/outputs/ヒアリングシート_{company_name}_{date_str}.xlsx"
    
    # Workbook作成
    wb = Workbook()
    
    # シート1: 基本情報
    create_basic_info_sheet(wb, company_name, industry, hearing_date)
    
    # シート2: IT環境
    create_it_environment_sheet(wb)
    
    # シート3〜: 部門別
    for dept in departments:
        if dept in DEPARTMENTS:
            create_department_sheet(wb, dept, industry)
    
    # 最終シート: 優先度・ゴール整理
    create_priority_sheet(wb)
    
    # 保存
    output_dir = Path(output_path).parent
    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ ヒアリングシートを生成しました: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description='ヒアリングシート生成')
    parser.add_argument('--company', required=True, help='ヒアリング先企業名')
    parser.add_argument('--industry', required=True, choices=INDUSTRIES, help='業種')
    parser.add_argument('--departments', nargs='+', required=True, help='対象部門')
    parser.add_argument('--date', help='ヒアリング予定日 (YYYY-MM-DD)')
    parser.add_argument('--output', help='出力ファイルパス')
    
    args = parser.parse_args()
    
    create_hearing_sheet(
        company_name=args.company,
        industry=args.industry,
        departments=args.departments,
        hearing_date=args.date,
        output_path=args.output
    )


if __name__ == '__main__':
    main()
