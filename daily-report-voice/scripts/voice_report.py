#!/usr/bin/env python3
"""
音声入力業務日報作成スクリプト
音声テキストを解析し、5列形式の日報Excelファイルを生成
"""

import re
import sys
import os
from datetime import datetime, date
from openpyxl import load_workbook
from pathlib import Path

class VoiceReportProcessor:
    def __init__(self):
        self.required_fields = ['作業内容', '課題', '次回予定']
        self.extracted_data = {
            '日付': datetime.now().strftime('%Y-%m-%d'),
            '報告者': 'Ken',
            '作業内容': '',
            '課題': '',
            '次回予定': ''
        }
        
    def analyze_voice_text(self, voice_text):
        """音声テキストから項目を抽出"""
        # 作業内容の抽出パターン
        work_patterns = [
            r'今日は(.+?)(?:しました|やりました|実施|完了|作成)',
            r'午前中?(.+?)(?:の作業|を実施|をやった)',
            r'(.+?)(?:に|で|を)(\d+時間?|半日|一日)',
            r'(.+?)(?:システム|開発|設計|会議|打ち合わせ|作業)',
        ]
        
        # 課題の抽出パターン
        issue_patterns = [
            r'(?:課題|問題)(?:が|として|は)(.+?)(?:です|ます|。)',
            r'(.+?)(?:で問題|が課題|に困って)',
            r'(?:困った|難しい|うまくいかない)(.+)',
        ]
        
        # 予定の抽出パターン  
        plan_patterns = [
            r'明日(?:は|の予定は?)(.+?)(?:です|ます|。)',
            r'来週(.+?)(?:予定|する予定)',
            r'(?:次回|今度)(.+?)(?:です|ます|。)',
        ]
        
        # 作業内容を抽出
        for pattern in work_patterns:
            match = re.search(pattern, voice_text, re.IGNORECASE)
            if match:
                work_content = match.group(1).strip()
                self.extracted_data['作業内容'] = work_content
                break
                
        # 課題を抽出
        for pattern in issue_patterns:
            match = re.search(pattern, voice_text, re.IGNORECASE)
            if match:
                issue_content = match.group(1).strip()
                self.extracted_data['課題'] = issue_content
                break
                
        # 予定を抽出
        for pattern in plan_patterns:
            match = re.search(pattern, voice_text, re.IGNORECASE)
            if match:
                plan_content = match.group(1).strip()
                self.extracted_data['次回予定'] = plan_content
                break
        
        return self.extracted_data
    
    def check_missing_fields(self):
        """不足項目をチェック"""
        missing = []
        for field in self.required_fields:
            if not self.extracted_data.get(field, '').strip():
                missing.append(field)
        return missing
    
    def generate_followup_questions(self, missing_fields):
        """不足項目に対する質問を生成"""
        questions = {
            '作業内容': "今日はどのような作業をされましたか？",
            '課題': "特に課題や問題はありませんでしたか？",
            '次回予定': "明日以降の予定はありますか？"
        }
        return [questions[field] for field in missing_fields if field in questions]
    
    def create_excel_report(self, template_path, output_path=None):
        """Excelファイルを作成"""
        if not output_path:
            today = datetime.now().strftime('%Y-%m-%d')
            output_path = f"MIRAICHI_日報_{today}.xlsx"
        
        # テンプレートを読み込み
        wb = load_workbook(template_path)
        sheet = wb.active
        
        # 今日の日付の行を探す
        today_date = datetime.now().strftime('%Y-%m-%d')
        target_row = None
        
        for row_num in range(2, 33):  # Row 2-32 (31 days)
            cell_value = sheet.cell(row=row_num, column=1).value
            if cell_value and str(cell_value)[:10] == today_date:
                target_row = row_num
                break
        
        if not target_row:
            # 空いている行を探す
            for row_num in range(2, 33):
                if not sheet.cell(row=row_num, column=2).value:  # Reporter column empty
                    target_row = row_num
                    sheet.cell(row=row_num, column=1, value=today_date)  # Set date
                    break
        
        if target_row:
            # データを挿入
            sheet.cell(row=target_row, column=2, value=self.extracted_data['報告者'])
            sheet.cell(row=target_row, column=3, value=self.extracted_data['作業内容'])
            sheet.cell(row=target_row, column=4, value=self.extracted_data['課題'])
            sheet.cell(row=target_row, column=5, value=self.extracted_data['次回予定'])
            
            # ファイル保存
            wb.save(output_path)
            return output_path
        else:
            raise Exception("適切な行が見つかりません。テンプレートを確認してください。")

def main():
    processor = VoiceReportProcessor()
    
    if len(sys.argv) < 2:
        print("使用方法: python voice_report.py '音声テキスト' [テンプレートパス] [出力パス]")
        sys.exit(1)
    
    voice_text = sys.argv[1]
    template_path = sys.argv[2] if len(sys.argv) > 2 else "../assets/template.xlsx"
    output_path = sys.argv[3] if len(sys.argv) > 3 else None
    
    # 音声解析
    extracted = processor.analyze_voice_text(voice_text)
    print("抽出されたデータ:")
    for key, value in extracted.items():
        print(f"  {key}: {value}")
    
    # 不足項目チェック
    missing = processor.check_missing_fields()
    if missing:
        print("\n不足している項目:")
        questions = processor.generate_followup_questions(missing)
        for question in questions:
            print(f"  - {question}")
        return
    
    # Excelファイル生成
    try:
        output_file = processor.create_excel_report(template_path, output_path)
        print(f"\n✅ 日報ファイルが作成されました: {output_file}")
    except Exception as e:
        print(f"❌ エラー: {e}")

if __name__ == "__main__":
    main()
