#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
見積書自動生成スクリプト（Jony Iveデザイン）
極限までシンプルに、美しく、機能的に。
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

# 見積書情報の設定
QUOTATION_DATA = {
    # 発行者情報
    "issuer_name": "株式会社けんけん",
    "issuer_address": "〒760-0000 香川県高松市○○町1-2-3",
    "issuer_tel": "087-123-4567",
    "issuer_email": "info@kenken.co.jp",
    
    # 見積先情報（デフォルト）
    "client_name": "合同会社大吉",
    "client_address": "〒000-0000 ○○県○○市○○町1-1-1",
    
    # 見積内容（デフォルト）
    "description": "生成AIアドバイザリー業務",
    "amount": 200000,  # 税抜金額
    
    # その他
    "tax_rate": 0.10,  # 消費税率10%
    "validity_days": 30,  # 見積有効期限（日数）
    "notes": [
        "お支払条件：月末締め翌月末払い",
        "契約期間：1年間（自動更新）",
        "恐れ入りますが、振込手数料はご負担ください。"
    ]
}

def create_quotation(client_name=None, description=None, amount=None, 
                    quotation_date=None, quotation_number=None, items=None):
    """
    見積書を作成する（Jony Iveデザイン）
    
    Args:
        client_name: 見積先名
        description: 業務内容（単一項目の場合）
        amount: 税抜金額（単一項目の場合）
        quotation_date: 見積日
        quotation_number: 見積書番号
        items: 複数明細の場合 [{"description": "", "qty": 1, "price": 100000}, ...]
    """
    # デフォルト値の設定
    if client_name is None:
        client_name = QUOTATION_DATA["client_name"]
    if quotation_date is None:
        quotation_date = datetime.now()
    if quotation_number is None:
        quotation_number = f"QUO-{quotation_date.strftime('%Y%m')}-001"
    
    # 明細の設定
    if items is None:
        if description is None:
            description = QUOTATION_DATA["description"]
        if amount is None:
            amount = QUOTATION_DATA["amount"]
        items = [{"description": description, "qty": 1, "price": amount}]
    
    # 金額計算
    subtotal = sum(item["qty"] * item["price"] for item in items)
    tax_amount = int(subtotal * QUOTATION_DATA["tax_rate"])
    total_amount = subtotal + tax_amount
    
    # 有効期限
    validity_date = quotation_date + timedelta(days=QUOTATION_DATA["validity_days"])
    
    # Excelワークブック作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "御見積書"
    
    # === Jony Iveデザイン哲学 ===
    # 請求書と完全に統一されたデザイン
    
    # フォント定義 - すべてメイリオで統一
    font_title = Font(name="メイリオ", size=28, bold=True, color="000000")
    font_large_amount = Font(name="メイリオ", size=36, bold=True, color="000000")
    font_section_header = Font(name="メイリオ", size=11, bold=True, color="666666")
    font_body = Font(name="メイリオ", size=10, color="333333")
    font_body_light = Font(name="メイリオ", size=9, color="999999")
    font_amount = Font(name="メイリオ", size=11, color="000000")
    font_amount_total = Font(name="メイリオ", size=14, bold=True, color="000000")
    
    # 罫線定義
    thin_line = Border(bottom=Side(style='thin', color='E0E0E0'))
    
    # 列幅設定
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 2
    
    # === ヘッダーセクション ===
    ws.row_dimensions[1].height = 15
    ws.row_dimensions[2].height = 20
    
    # タイトル
    ws.merge_cells('B3:E3')
    ws['B3'] = "御見積書"
    ws['B3'].font = font_title
    ws['B3'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 35
    
    # 見積書番号と日付
    ws['B5'] = quotation_number
    ws['B5'].font = font_body_light
    ws['E5'] = quotation_date.strftime('%Y.%m.%d')
    ws['E5'].font = font_body_light
    ws['E5'].alignment = Alignment(horizontal="right")
    
    # セパレーター
    ws['B6'].border = thin_line
    ws['C6'].border = thin_line
    ws['D6'].border = thin_line
    ws['E6'].border = thin_line
    
    ws.row_dimensions[7].height = 12
    
    # === 見積先セクション ===
    ws['B8'] = client_name
    ws['B8'].font = Font(name="メイリオ", size=12, color="000000")
    
    ws.row_dimensions[9].height = 20
    
    # === 見積金額 ===
    ws.merge_cells('B10:E10')
    ws['B10'] = f"¥{total_amount:,}"
    ws['B10'].font = font_large_amount
    ws['B10'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[10].height = 54  # 最適化: 36pt × 1.5 = 54 (見切れ防止)
    
    ws.row_dimensions[11].height = 8
    
    # 有効期限
    ws['B12'] = f"お見積有効期限  {validity_date.strftime('%Y年%m月%d日')}"
    ws['B12'].font = font_body_light
    
    ws.row_dimensions[13].height = 20
    
    # === 明細セクション ===
    row = 14
    ws['B14'] = "明細"
    ws['B14'].font = font_section_header
    
    # セパレーター
    ws['B15'].border = thin_line
    ws['C15'].border = thin_line
    ws['D15'].border = thin_line
    ws['E15'].border = thin_line
    
    row = 16
    
    # 明細行
    for item in items:
        ws[f'B{row}'] = item["description"]
        ws[f'B{row}'].font = font_body
        
        if item["qty"] > 1 or len(items) > 1:
            # 複数項目または数量が2以上の場合は数量と単価を表示
            ws[f'C{row}'] = item["qty"]
            ws[f'C{row}'].font = font_body
            ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="center")
            
            ws[f'D{row}'] = item["price"]
            ws[f'D{row}'].font = font_amount
            ws[f'D{row}'].number_format = '¥#,##0'
            ws[f'D{row}'].alignment = Alignment(horizontal="right", vertical="center")
        
        ws[f'E{row}'] = item["qty"] * item["price"]
        ws[f'E{row}'].font = font_amount
        ws[f'E{row}'].number_format = '¥#,##0'
        ws[f'E{row}'].alignment = Alignment(horizontal="right", vertical="center")
        
        ws.row_dimensions[row].height = 22
        row += 1
    
    # 空行
    ws.row_dimensions[row].height = 4
    row += 1
    
    # 小計
    ws[f'D{row}'] = "小計"
    ws[f'D{row}'].font = font_body_light
    ws[f'D{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws[f'E{row}'] = subtotal
    ws[f'E{row}'].font = font_amount
    ws[f'E{row}'].number_format = '¥#,##0'
    ws[f'E{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1
    
    # 消費税
    ws[f'D{row}'] = "消費税（10%）"
    ws[f'D{row}'].font = font_body_light
    ws[f'D{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws[f'E{row}'] = tax_amount
    ws[f'E{row}'].font = font_amount
    ws[f'E{row}'].number_format = '¥#,##0'
    ws[f'E{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1
    
    # セパレーター
    ws[f'D{row}'].border = thin_line
    ws[f'E{row}'].border = thin_line
    row += 1
    
    # 合計
    ws[f'D{row}'] = "合計"
    ws[f'D{row}'].font = font_amount_total
    ws[f'D{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws[f'E{row}'] = total_amount
    ws[f'E{row}'].font = font_amount_total
    ws[f'E{row}'].number_format = '¥#,##0'
    ws[f'E{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 25
    row += 1
    
    ws.row_dimensions[row].height = 25
    row += 1
    
    # === 備考セクション ===
    ws[f'B{row}'] = "備考"
    ws[f'B{row}'].font = font_section_header
    row += 1
    
    # セパレーター
    ws[f'B{row}'].border = thin_line
    ws[f'C{row}'].border = thin_line
    ws[f'D{row}'].border = thin_line
    ws[f'E{row}'].border = thin_line
    row += 1
    
    # 備考内容
    for note in QUOTATION_DATA["notes"]:
        ws[f'B{row}'] = note
        ws[f'B{row}'].font = font_body
        ws.row_dimensions[row].height = 18
        row += 1
    
    ws.row_dimensions[row].height = 25
    row += 1
    
    # === 発行者情報 ===
    ws[f'B{row}'] = QUOTATION_DATA["issuer_name"]
    ws[f'B{row}'].font = font_body_light
    ws.row_dimensions[row].height = 16
    row += 1
    
    ws[f'B{row}'] = QUOTATION_DATA["issuer_address"]
    ws[f'B{row}'].font = font_body_light
    ws.row_dimensions[row].height = 14
    row += 1
    
    ws[f'B{row}'] = f"{QUOTATION_DATA['issuer_tel']}  {QUOTATION_DATA['issuer_email']}"
    ws[f'B{row}'].font = font_body_light
    ws.row_dimensions[row].height = 14
    
    # 下部余白
    ws.row_dimensions[row+1].height = 15
    
    # グリッド線を非表示
    ws.sheet_view.showGridLines = False
    
    # Excelファイル保存
    excel_filename = f"/mnt/user-data/outputs/御見積書_{quotation_number}.xlsx"
    wb.save(excel_filename)
    print(f"Excelファイルを作成しました: {excel_filename}")
    
    return excel_filename

if __name__ == "__main__":
    # デフォルト設定で見積書を作成
    create_quotation()
