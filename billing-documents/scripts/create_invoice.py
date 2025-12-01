#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
請求書自動生成スクリプト（Jony Iveデザイン）
極限までシンプルに、美しく、機能的に。
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

# 請求書情報の設定
INVOICE_DATA = {
    # 発行者情報
    "issuer_name": "株式会社けんけん",
    "issuer_address": "〒760-0000 香川県高松市○○町1-2-3",
    "issuer_tel": "087-123-4567",
    "issuer_email": "info@kenken.co.jp",
    
    # 振込先情報
    "bank_name": "○○銀行",
    "branch_name": "高松支店",
    "account_type": "普通",
    "account_number": "1234567",
    "account_holder": "カ)ケンケン",
    
    # 請求先情報（デフォルト）
    "client_name": "合同会社大吉",
    "client_address": "〒000-0000 ○○県○○市○○町1-1-1",
    
    # 請求内容（デフォルト）
    "description": "生成AIアドバイザリー業務",
    "amount": 200000,  # 税抜金額
    
    # その他
    "tax_rate": 0.10,  # 消費税率10%
    "payment_note": "恐れ入りますが、振込手数料はご負担ください。"
}

def create_invoice(client_name=None, description=None, amount=None, invoice_date=None, invoice_number=None):
    """
    請求書を作成する（Jony Iveデザイン）
    
    極限までシンプルに。不要な装飾を排除し、余白と整列で美しさを表現。
    """
    # デフォルト値の設定
    if client_name is None:
        client_name = INVOICE_DATA["client_name"]
    if description is None:
        description = INVOICE_DATA["description"]
    if amount is None:
        amount = INVOICE_DATA["amount"]
    if invoice_date is None:
        invoice_date = datetime.now()
    if invoice_number is None:
        invoice_number = f"INV-{invoice_date.strftime('%Y%m')}-001"
    
    # 金額計算
    tax_amount = int(amount * INVOICE_DATA["tax_rate"])
    total_amount = amount + tax_amount
    
    # 支払期限（請求日の月末）
    last_day = (invoice_date.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    payment_deadline = last_day
    
    # 対象期間（請求日の月）
    period = invoice_date.strftime("%Y年%m月分")
    
    # Excelワークブック作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求書"
    
    # === Jony Iveデザイン哲学 ===
    # 1. 極限までシンプルに - 不要な装飾を排除
    # 2. 余白を贅沢に使う
    # 3. 完璧な整列
    # 4. 控えめな色使い（グレーとアクセントカラーのみ）
    # 5. フォント階層の明確化
    
    # フォント定義 - すべてメイリオで統一
    font_title = Font(name="メイリオ", size=28, bold=True, color="000000")
    font_large_amount = Font(name="メイリオ", size=36, bold=True, color="000000")
    font_section_header = Font(name="メイリオ", size=11, bold=True, color="666666")
    font_body = Font(name="メイリオ", size=10, color="333333")
    font_body_light = Font(name="メイリオ", size=9, color="999999")
    font_amount = Font(name="メイリオ", size=11, color="000000")
    font_amount_total = Font(name="メイリオ", size=14, bold=True, color="000000")
    
    # 罫線定義 - 最小限の使用
    thin_line = Border(bottom=Side(style='thin', color='E0E0E0'))
    no_border = Border()
    
    # 背景色 - ごく微妙なグレーのみ
    subtle_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    
    # 列幅設定 - 余白を贅沢に
    ws.column_dimensions['A'].width = 2   # 左余白
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 2   # 右余白
    
    # === ヘッダーセクション ===
    # 上部余白
    ws.row_dimensions[1].height = 15
    ws.row_dimensions[2].height = 20
    
    # タイトル - 大胆にシンプルに
    ws.merge_cells('B3:E3')
    ws['B3'] = "請求書"
    ws['B3'].font = font_title
    ws['B3'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 35
    
    # 請求書番号と日付 - 小さく控えめに
    ws['B5'] = invoice_number
    ws['B5'].font = font_body_light
    ws['E5'] = invoice_date.strftime('%Y.%m.%d')
    ws['E5'].font = font_body_light
    ws['E5'].alignment = Alignment(horizontal="right")
    
    # セパレーター（細いライン）
    ws['B6'].border = thin_line
    ws['C6'].border = thin_line
    ws['D6'].border = thin_line
    ws['E6'].border = thin_line
    
    ws.row_dimensions[7].height = 12
    
    # === 請求先セクション ===
    ws['B8'] = client_name
    ws['B8'].font = Font(name="メイリオ", size=12, color="000000")
    
    ws.row_dimensions[9].height = 20
    
    # === 請求金額 - 最も目立たせる ===
    ws.merge_cells('B10:E10')
    ws['B10'] = f"¥{total_amount:,}"
    ws['B10'].font = font_large_amount
    ws['B10'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[10].height = 54  # 最適化: 36pt × 1.5 = 54 (見切れ防止)
    
    ws.row_dimensions[11].height = 8
    
    # 支払期限 - 控えめに
    ws['B12'] = f"お支払期限  {payment_deadline.strftime('%Y年%m月%d日')}"
    ws['B12'].font = font_body_light
    
    ws.row_dimensions[13].height = 20
    
    # === 明細セクション - 罫線を最小限に ===
    # ヘッダー
    row = 14
    ws['B14'] = "明細"
    ws['B14'].font = font_section_header
    
    # セパレーター
    ws['B15'].border = thin_line
    ws['C15'].border = thin_line
    ws['D15'].border = thin_line
    ws['E15'].border = thin_line
    
    row = 16
    
    # 業務内容
    ws[f'B{row}'] = f"{description}（{period}）"
    ws[f'B{row}'].font = font_body
    ws[f'E{row}'] = amount
    ws[f'E{row}'].font = font_amount
    ws[f'E{row}'].number_format = '¥#,##0'
    ws[f'E{row}'].alignment = Alignment(horizontal="right", vertical="center")
    
    ws.row_dimensions[row].height = 22
    row += 1
    
    # 空行
    ws.row_dimensions[row].height = 4
    row += 1
    
    # 小計
    ws[f'D{row}'] = "小計"
    ws[f'D{row}'].font = font_body_light
    ws[f'D{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws[f'E{row}'] = amount
    ws[f'E{row}'].font = font_amount
    ws[f'E{row}'].number_format = '¥#,##0'
    ws[f'E{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1
    
    # 消費税
    ws[f'D{row}'] = "消費税（10%）"
    ws[f'D{row}'].font = font_body_light
    ws[f'D{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws[f'E{row}'] = tax_amount
    ws[f'E{row}'].font = font_amount
    ws[f'E{row}'].number_format = '¥#,##0'
    ws[f'E{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1
    
    # セパレーター
    ws[f'D{row}'].border = thin_line
    ws[f'E{row}'].border = thin_line
    row += 1
    
    # 合計
    ws[f'D{row}'] = "合計"
    ws[f'D{row}'].font = font_amount_total
    ws[f'D{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws[f'E{row}'] = total_amount
    ws[f'E{row}'].font = font_amount_total
    ws[f'E{row}'].number_format = '¥#,##0'
    ws[f'E{row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 25
    row += 1
    
    ws.row_dimensions[row].height = 25
    row += 1
    
    # === 振込先情報セクション ===
    ws[f'B{row}'] = "振込先"
    ws[f'B{row}'].font = font_section_header
    row += 1
    
    # セパレーター
    ws[f'B{row}'].border = thin_line
    ws[f'C{row}'].border = thin_line
    ws[f'D{row}'].border = thin_line
    ws[f'E{row}'].border = thin_line
    row += 1
    
    ws[f'B{row}'] = f"{INVOICE_DATA['bank_name']} {INVOICE_DATA['branch_name']}"
    ws[f'B{row}'].font = font_body
    ws.row_dimensions[row].height = 18
    row += 1
    
    ws[f'B{row}'] = f"{INVOICE_DATA['account_type']} {INVOICE_DATA['account_number']}"
    ws[f'B{row}'].font = font_body
    ws.row_dimensions[row].height = 18
    row += 1
    
    ws[f'B{row}'] = f"{INVOICE_DATA['account_holder']}"
    ws[f'B{row}'].font = font_body
    ws.row_dimensions[row].height = 18
    row += 1
    
    ws.row_dimensions[row].height = 15
    row += 1
    
    # 備考 - 最も控えめに
    ws[f'B{row}'] = INVOICE_DATA["payment_note"]
    ws[f'B{row}'].font = font_body_light
    ws.row_dimensions[row].height = 18
    row += 1
    
    ws.row_dimensions[row].height = 25
    row += 1
    
    # === 発行者情報 - 最下部に控えめに ===
    ws[f'B{row}'] = INVOICE_DATA["issuer_name"]
    ws[f'B{row}'].font = font_body_light
    ws.row_dimensions[row].height = 16
    row += 1
    
    ws[f'B{row}'] = INVOICE_DATA["issuer_address"]
    ws[f'B{row}'].font = font_body_light
    ws.row_dimensions[row].height = 14
    row += 1
    
    ws[f'B{row}'] = f"{INVOICE_DATA['issuer_tel']}  {INVOICE_DATA['issuer_email']}"
    ws[f'B{row}'].font = font_body_light
    ws.row_dimensions[row].height = 14
    
    # 下部余白
    ws.row_dimensions[row+1].height = 15
    
    # ページ全体の背景を白に保つ（Appleのミニマリズム）
    # グリッド線を非表示に
    ws.sheet_view.showGridLines = False
    
    # Excelファイル保存
    excel_filename = f"/mnt/user-data/outputs/請求書_{invoice_number}.xlsx"
    wb.save(excel_filename)
    print(f"Excelファイルを作成しました: {excel_filename}")
    
    return excel_filename

if __name__ == "__main__":
    # デフォルト設定で請求書を作成
    create_invoice()
