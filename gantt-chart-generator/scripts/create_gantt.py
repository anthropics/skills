#!/usr/bin/env python3
"""
Gantt Chart Generator
プロジェクトのガントチャートをExcel形式で生成
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import os

# ========================================
# プロジェクトデータ（ここを編集）
# ========================================
PROJECT_DATA = {
    "project_name": "DXコンサルティングプロジェクト",
    "start_date": "2025-01-06",
    "responsible": "プロジェクトマネージャー",
    "description": "クライアント企業のDX推進支援",
    "phases": [
        {
            "name": "フェーズ1: 現状分析",
            "tasks": [
                {"name": "キックオフミーティング", "assignee": "全員", "days": 1},
                {"name": "ヒアリング実施", "assignee": "田中", "days": 5},
                {"name": "業務フロー分析", "assignee": "佐藤", "days": 3},
                {"name": "課題整理", "assignee": "田中", "days": 2},
            ],
            "milestone": {"name": "現状分析完了", "days_from_phase_start": 11}
        },
        {
            "name": "フェーズ2: 施策立案",
            "tasks": [
                {"name": "改善案検討", "assignee": "全員", "days": 3},
                {"name": "提案書作成", "assignee": "田中", "days": 5},
                {"name": "レビュー・修正", "assignee": "佐藤", "days": 2},
            ],
            "milestone": {"name": "提案書提出", "days_from_phase_start": 10}
        },
        {
            "name": "フェーズ3: 実装支援",
            "tasks": [
                {"name": "実装計画策定", "assignee": "田中", "days": 3},
                {"name": "ツール導入支援", "assignee": "佐藤", "days": 10},
                {"name": "運用マニュアル作成", "assignee": "田中", "days": 5},
            ],
            "milestone": {"name": "導入完了", "days_from_phase_start": 18}
        },
        {
            "name": "フェーズ4: 定着化",
            "tasks": [
                {"name": "トレーニング実施", "assignee": "全員", "days": 3},
                {"name": "運用フォロー", "assignee": "田中", "days": 10},
                {"name": "最終レポート作成", "assignee": "佐藤", "days": 3},
            ],
            "milestone": {"name": "プロジェクト完了", "days_from_phase_start": 16}
        }
    ]
}

# ========================================
# デザイン定数（Jony Ive哲学 - グレースケール＋1アクセント）
# ========================================
COLORS = {
    # 基本グレースケール
    "black": "000000",
    "dark_gray": "333333",
    "medium_gray": "666666",
    "light_gray": "999999",
    "pale_gray": "CCCCCC",
    "separator": "E0E0E0",
    "bg_gray": "F5F5F5",
    "white": "FFFFFF",
    
    # アクセント（1色のみ）
    "accent": "5B7B94",
    
    # ステータス（グレー濃淡）
    "status_pending": "CCCCCC",    # 未着手（薄い＝控えめ）
    "status_progress": "666666",   # 進行中（中間＝注目）
    "status_done": "333333",       # 完了（濃い＝確定）
    "status_hold": "FFFFFF",       # 保留（白＝空）
    "milestone": "5B7B94",         # マイルストーン（唯一のアクセント）
    
    # 構造
    "header_bg": "333333",         # ヘッダー背景
    "phase_bg": "F5F5F5",          # フェーズ背景
}

FONT_NAME = "メイリオ"

def create_gantt_chart(project_data=None, output_dir="/mnt/user-data/outputs"):
    """ガントチャートを生成"""
    if project_data is None:
        project_data = PROJECT_DATA
    
    wb = Workbook()
    
    # シート1: プロジェクト概要
    ws_overview = wb.active
    ws_overview.title = "プロジェクト概要"
    create_overview_sheet(ws_overview, project_data)
    
    # シート2: タスク（ガントチャート統合）
    ws_tasks = wb.create_sheet("タスク")
    create_task_sheet(ws_tasks, project_data)
    
    # 出力
    os.makedirs(output_dir, exist_ok=True)
    safe_name = project_data["project_name"].replace(" ", "_").replace("/", "_")
    filename = f"ガントチャート_{safe_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    print(f"✅ ガントチャートを作成しました: {filepath}")
    return filepath


def create_overview_sheet(ws, data):
    """プロジェクト概要シートを作成"""
    ws.sheet_view.showGridLines = False
    
    # 列幅設定
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 50
    
    row = 2
    
    # タイトル
    ws.cell(row=row, column=2, value=data["project_name"])
    ws.cell(row=row, column=2).font = Font(name=FONT_NAME, size=16, bold=True, color=COLORS["black"])
    row += 2
    
    # 基本情報
    start_date = datetime.strptime(data["start_date"], "%Y-%m-%d")
    end_date = calculate_end_date(data, start_date)
    
    info_items = [
        ("期間", f"{start_date.strftime('%Y年%m月%d日')} 〜 {end_date.strftime('%Y年%m月%d日')}"),
        ("責任者", data["responsible"]),
        ("概要", data.get("description", "")),
    ]
    
    for label, value in info_items:
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=2).font = Font(name=FONT_NAME, size=10, bold=True, color=COLORS["medium_gray"])
        ws.cell(row=row, column=3, value=value)
        ws.cell(row=row, column=3).font = Font(name=FONT_NAME, size=10, color=COLORS["dark_gray"])
        row += 1
    
    row += 1
    
    # 主要フェーズ
    ws.cell(row=row, column=2, value="主要フェーズ")
    ws.cell(row=row, column=2).font = Font(name=FONT_NAME, size=11, bold=True, color=COLORS["dark_gray"])
    row += 1
    
    for i, phase in enumerate(data["phases"], 1):
        ws.cell(row=row, column=2, value=f"フェーズ{i}")
        ws.cell(row=row, column=2).font = Font(name=FONT_NAME, size=10, color=COLORS["medium_gray"])
        ws.cell(row=row, column=3, value=phase["name"])
        ws.cell(row=row, column=3).font = Font(name=FONT_NAME, size=10, color=COLORS["dark_gray"])
        row += 1
    
    row += 1
    
    # マイルストーン
    ws.cell(row=row, column=2, value="マイルストーン")
    ws.cell(row=row, column=2).font = Font(name=FONT_NAME, size=11, bold=True, color=COLORS["dark_gray"])
    row += 1
    
    phase_start = start_date
    ms_count = 1
    for phase in data["phases"]:
        if "milestone" in phase and phase["milestone"]:
            ms = phase["milestone"]
            ms_date = phase_start + timedelta(days=ms["days_from_phase_start"])
            ws.cell(row=row, column=2, value=f"MS{ms_count}")
            ws.cell(row=row, column=2).font = Font(name=FONT_NAME, size=10, color=COLORS["milestone"])
            ws.cell(row=row, column=3, value=f"◆ {ms['name']} ({ms_date.strftime('%m/%d')})")
            ws.cell(row=row, column=3).font = Font(name=FONT_NAME, size=10, color=COLORS["dark_gray"])
            row += 1
            ms_count += 1
        
        # 次のフェーズの開始日を計算
        phase_days = sum(t["days"] for t in phase["tasks"])
        phase_start = phase_start + timedelta(days=phase_days)


def create_task_sheet(ws, data):
    """タスクシート（ガントチャート統合）を作成 - 動的数式版"""
    ws.sheet_view.showGridLines = False
    
    start_date = datetime.strptime(data["start_date"], "%Y-%m-%d")
    end_date = calculate_end_date(data, start_date)
    total_weeks = ((end_date - start_date).days // 7) + 2
    total_weeks = min(total_weeks, 24)
    
    # ヘッダー行
    headers = ["WBS", "タスク名", "担当", "開始日", "終了日", "日数", "ステータス"]
    week_headers = [f"W{i+1}" for i in range(total_weeks)]
    all_headers = headers + week_headers
    
    # 列幅設定
    col_widths = [8, 35, 12, 14, 14, 6, 10] + [5] * total_weeks
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # タイトル行
    row = 1
    ws.cell(row=row, column=1, value=f"{data['project_name']} - ガントチャート")
    ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=14, bold=True, color=COLORS["black"])
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.row_dimensions[row].height = 25
    
    # プロジェクト開始日を隠しセルに格納（数式参照用）
    ws.cell(row=1, column=100, value=start_date)  # CV1に開始日を格納
    ws.cell(row=1, column=100).number_format = 'YYYY/MM/DD'
    
    # 開始日表示
    row = 2
    ws.cell(row=row, column=1, value=f"開始日: {start_date.strftime('%Y年%m月%d日')}")
    ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=9, color=COLORS["medium_gray"])
    
    # 行3: 各週の開始日（数式参照用、非表示）
    row = 3
    for week in range(total_weeks):
        col = 8 + week
        if week == 0:
            ws.cell(row=row, column=col, value=f"=$CV$1")
        else:
            ws.cell(row=row, column=col, value=f"=$CV$1+{week*7}")
        ws.cell(row=row, column=col).number_format = 'YYYY/MM/DD'
        ws.cell(row=row, column=col).font = Font(name=FONT_NAME, size=8, color=COLORS["light_gray"])
    ws.row_dimensions[row].height = 12
    ws.row_dimensions[row].hidden = True  # 行を非表示
    
    row = 4
    
    # ヘッダー
    header_fill = PatternFill(start_color=COLORS["header_bg"], fill_type="solid")
    header_font = Font(name=FONT_NAME, size=10, bold=True, color=COLORS["white"])
    thin_border = Border(
        bottom=Side(style='thin', color=COLORS["separator"])
    )
    
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[row].height = 20
    
    # ステータスのドロップダウン設定
    status_validation = DataValidation(
        type="list",
        formula1='"未着手,進行中,完了,保留"',
        allow_blank=True
    )
    ws.add_data_validation(status_validation)
    
    # データ行
    row = 5
    current_date = start_date
    phase_num = 0
    ms_num = 0
    task_rows = []  # タスク行を記録（条件付き書式用）
    
    for phase in data["phases"]:
        phase_num += 1
        phase_start_date = current_date
        
        # フェーズ行
        phase_fill = PatternFill(start_color=COLORS["phase_bg"], fill_type="solid")
        phase_font = Font(name=FONT_NAME, size=10, bold=True, color=COLORS["dark_gray"])
        
        phase_days = sum(t["days"] for t in phase["tasks"])
        phase_end_date = phase_start_date + timedelta(days=phase_days - 1)
        
        ws.cell(row=row, column=1, value=str(phase_num))
        ws.cell(row=row, column=2, value=phase["name"])
        ws.cell(row=row, column=4, value=phase_start_date)
        ws.cell(row=row, column=4).number_format = 'YYYY/MM/DD'
        ws.cell(row=row, column=5, value=phase_end_date)
        ws.cell(row=row, column=5).number_format = 'YYYY/MM/DD'
        ws.cell(row=row, column=6, value=phase_days)
        
        for col in range(1, len(all_headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = phase_fill
            cell.font = phase_font
            cell.alignment = Alignment(vertical='center')
        
        ws.row_dimensions[row].height = 22
        row += 1
        
        # タスク行
        task_num = 0
        for task in phase["tasks"]:
            task_num += 1
            wbs = f"{phase_num}.{task_num}"
            task_end_date = current_date + timedelta(days=task["days"] - 1)
            
            ws.cell(row=row, column=1, value=wbs)
            ws.cell(row=row, column=2, value=task["name"])
            ws.cell(row=row, column=3, value=task["assignee"])
            
            # 開始日・終了日を日付形式で入力
            ws.cell(row=row, column=4, value=current_date)
            ws.cell(row=row, column=4).number_format = 'YYYY/MM/DD'
            ws.cell(row=row, column=5, value=task_end_date)
            ws.cell(row=row, column=5).number_format = 'YYYY/MM/DD'
            
            # 日数は数式
            ws.cell(row=row, column=6, value=f'=E{row}-D{row}+1')
            ws.cell(row=row, column=7, value="未着手")
            
            # スタイル適用
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(name=FONT_NAME, size=10, color=COLORS["dark_gray"])
                cell.alignment = Alignment(vertical='center')
                cell.border = thin_border
            
            # ステータス列にドロップダウン追加
            status_validation.add(ws.cell(row=row, column=7))
            
            # ガントバー（■）を動的数式で配置
            for week in range(total_weeks):
                col = 8 + week
                week_col_letter = get_column_letter(col)
                # 数式: タスク期間と週が重なるかチェック
                # 週開始日は行3に格納、週終了日は週開始日+6
                formula = f'=IF(AND($D{row}<={week_col_letter}$3+6,$E{row}>={week_col_letter}$3),"■","")'
                ws.cell(row=row, column=col, value=formula)
                ws.cell(row=row, column=col).font = Font(name=FONT_NAME, size=9, color=COLORS["status_pending"])
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            task_rows.append(row)
            ws.row_dimensions[row].height = 18
            current_date = task_end_date + timedelta(days=1)
            row += 1
        
        # マイルストーン行
        if "milestone" in phase and phase["milestone"]:
            ms_num += 1
            ms = phase["milestone"]
            ms_date = phase_start_date + timedelta(days=ms["days_from_phase_start"])
            
            ws.cell(row=row, column=1, value=f"MS{ms_num}")
            ws.cell(row=row, column=2, value=f"◆ {ms['name']}")
            ws.cell(row=row, column=4, value=ms_date)
            ws.cell(row=row, column=4).number_format = 'YYYY/MM/DD'
            ws.cell(row=row, column=5, value=ms_date)
            ws.cell(row=row, column=5).number_format = 'YYYY/MM/DD'
            ws.cell(row=row, column=6, value=1)
            
            ms_font = Font(name=FONT_NAME, size=10, bold=True, color=COLORS["milestone"])
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = ms_font
                cell.alignment = Alignment(vertical='center')
            
            # マイルストーンマーカー（動的数式）
            for week in range(total_weeks):
                col = 8 + week
                week_col_letter = get_column_letter(col)
                formula = f'=IF(AND($D{row}<={week_col_letter}$3+6,$E{row}>={week_col_letter}$3),"◆","")'
                ws.cell(row=row, column=col, value=formula)
                ws.cell(row=row, column=col).font = Font(name=FONT_NAME, size=10, color=COLORS["milestone"])
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            ws.row_dimensions[row].height = 18
            row += 1
    
    # 条件付き書式（ステータス連動）
    data_start_row = 5
    data_end_row = row - 1
    week_start_col = 8
    week_end_col = 7 + total_weeks
    
    # ガントバー範囲
    gantt_range = f"{get_column_letter(week_start_col)}{data_start_row}:{get_column_letter(week_end_col)}{data_end_row}"
    
    # ステータス「進行中」→ ミディアムグレー（注目）
    progress_fill = PatternFill(start_color=COLORS["status_progress"], fill_type="solid")
    progress_font = Font(color=COLORS["white"])
    ws.conditional_formatting.add(
        gantt_range,
        FormulaRule(
            formula=[f'AND($G{data_start_row}="進行中", {get_column_letter(week_start_col)}{data_start_row}="■")'],
            fill=progress_fill,
            font=progress_font
        )
    )
    
    # ステータス「完了」→ ダークグレー（確定）
    done_fill = PatternFill(start_color=COLORS["status_done"], fill_type="solid")
    done_font = Font(color=COLORS["white"])
    ws.conditional_formatting.add(
        gantt_range,
        FormulaRule(
            formula=[f'AND($G{data_start_row}="完了", {get_column_letter(week_start_col)}{data_start_row}="■")'],
            fill=done_fill,
            font=done_font
        )
    )
    
    # ステータス「保留」→ 白背景、ダークグレー文字（空の状態）
    hold_fill = PatternFill(start_color=COLORS["status_hold"], fill_type="solid")
    hold_font = Font(color=COLORS["dark_gray"])
    ws.conditional_formatting.add(
        gantt_range,
        FormulaRule(
            formula=[f'AND($G{data_start_row}="保留", {get_column_letter(week_start_col)}{data_start_row}="■")'],
            fill=hold_fill,
            font=hold_font
        )
    )
    
    # マイルストーン◆の条件付き書式（アクセントブルー背景）
    ms_fill = PatternFill(start_color=COLORS["milestone"], fill_type="solid")
    ms_font_white = Font(color=COLORS["white"])
    ws.conditional_formatting.add(
        gantt_range,
        FormulaRule(
            formula=[f'{get_column_letter(week_start_col)}{data_start_row}="◆"'],
            fill=ms_fill,
            font=ms_font_white
        )
    )
    
    # === 凡例を追加 ===
    row += 2  # 空行
    legend_row = row
    
    ws.cell(row=legend_row, column=1, value="【凡例】")
    ws.cell(row=legend_row, column=1).font = Font(name=FONT_NAME, size=10, bold=True, color=COLORS["dark_gray"])
    
    # 凡例項目（背景色, 記号, ラベル, 文字色, 枠線有無）
    legend_items = [
        (COLORS["status_pending"], "■", "未着手", COLORS["dark_gray"], False),
        (COLORS["status_progress"], "■", "進行中", COLORS["white"], False),
        (COLORS["status_done"], "■", "完了", COLORS["white"], False),
        (COLORS["status_hold"], "□", "保留", COLORS["dark_gray"], True),  # 枠線付き
        (COLORS["milestone"], "◆", "マイルストーン", COLORS["white"], False),
        (COLORS["phase_bg"], "―", "フェーズ区分", COLORS["dark_gray"], False),
    ]
    
    legend_row += 1
    for bg_color, symbol, label, font_color, has_border in legend_items:
        # 色セル（A列）に記号を表示
        color_cell = ws.cell(row=legend_row, column=1, value=symbol)
        color_cell.fill = PatternFill(start_color=bg_color, fill_type="solid")
        color_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=font_color)
        color_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 保留は枠線を追加
        if has_border:
            color_cell.border = Border(
                left=Side(style='thin', color=COLORS["dark_gray"]),
                right=Side(style='thin', color=COLORS["dark_gray"]),
                top=Side(style='thin', color=COLORS["dark_gray"]),
                bottom=Side(style='thin', color=COLORS["dark_gray"])
            )
        
        # ラベルセル（B列）
        label_cell = ws.cell(row=legend_row, column=2, value=label)
        label_cell.font = Font(name=FONT_NAME, size=9, color=COLORS["dark_gray"])
        label_cell.alignment = Alignment(vertical='center')
        
        ws.row_dimensions[legend_row].height = 16
        legend_row += 1
    
    # ウィンドウ枠の固定
    ws.freeze_panes = "H5"


def calculate_end_date(data, start_date):
    """プロジェクト終了日を計算"""
    total_days = 0
    for phase in data["phases"]:
        phase_days = sum(t["days"] for t in phase["tasks"])
        total_days += phase_days
    return start_date + timedelta(days=total_days - 1)


if __name__ == "__main__":
    create_gantt_chart()
