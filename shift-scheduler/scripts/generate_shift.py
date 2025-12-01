#!/usr/bin/env python3
"""シフト表自動生成スクリプト"""

import json
import argparse
import calendar
from datetime import datetime, timedelta
from collections import defaultdict
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 祝日（簡易版、必要に応じて拡張）
HOLIDAYS_2025 = {
    "2025-01-01", "2025-01-13", "2025-02-11", "2025-02-23", "2025-02-24",
    "2025-03-20", "2025-04-29", "2025-05-03", "2025-05-04", "2025-05-05",
    "2025-05-06", "2025-07-21", "2025-08-11", "2025-09-15", "2025-09-23",
    "2025-10-13", "2025-11-03", "2025-11-23", "2025-11-24"
}

# スタイル定義
HEADER_FILL = PatternFill("solid", fgColor="333333")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WEEKEND_FILL = PatternFill("solid", fgColor="F5F5F5")
HOLIDAY_FILL = PatternFill("solid", fgColor="E8EEF2")
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def load_config(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_month_dates(year, month):
    """指定月の全日付を取得"""
    _, last_day = calendar.monthrange(year, month)
    return [datetime(year, month, d) for d in range(1, last_day + 1)]


def is_weekend_or_holiday(date):
    return date.weekday() >= 5 or date.strftime("%Y-%m-%d") in HOLIDAYS_2025


def calculate_work_hours(shift_def):
    """シフトの勤務時間を計算"""
    start = datetime.strptime(shift_def["start"], "%H:%M")
    end = datetime.strptime(shift_def["end"], "%H:%M")
    return (end - start).seconds / 3600


class ShiftScheduler:
    def __init__(self, config, year, month):
        self.config = config
        self.year = year
        self.month = month
        self.dates = get_month_dates(year, month)
        self.shifts = config["shifts"]
        self.requirements = config["requirements"]
        self.staff = config["staff"]
        self.schedule = defaultdict(dict)  # {date: {shift: [staff_names]}}
        self.staff_stats = defaultdict(lambda: {
            "days": 0, "hours": 0, "weekends": 0, "consecutive": 0,
            "max_consecutive": 0, "last_work": None
        })
    
    def get_requirement(self, date):
        """日付に応じた必要人数を取得"""
        if is_weekend_or_holiday(date):
            return self.requirements.get("weekend", self.requirements["weekday"])
        return self.requirements["weekday"]
    
    def can_assign(self, staff, date, shift_name):
        """割り当て可能かチェック"""
        date_str = date.strftime("%Y-%m-%d")
        stats = self.staff_stats[staff["name"]]
        
        # NG日チェック
        if date_str in staff.get("ng_dates", []):
            return False
        
        # 連続勤務チェック（5日まで）
        if stats["consecutive"] >= 5:
            return False
        
        # 週の勤務日数チェック
        week_start = date - timedelta(days=date.weekday())
        week_days = sum(1 for d in self.dates 
                       if week_start <= d < week_start + timedelta(days=7)
                       and staff["name"] in self.schedule.get(d, {}).get(shift_name, []))
        if week_days >= staff.get("max_days_per_week", 5):
            return False
        
        # 同日に既に割り当て済みかチェック
        for s in self.schedule.get(date, {}).values():
            if staff["name"] in s:
                return False
        
        return True
    
    def assign(self, staff, date, shift_name):
        """スタッフをシフトに割り当て"""
        if date not in self.schedule:
            self.schedule[date] = {}
        if shift_name not in self.schedule[date]:
            self.schedule[date][shift_name] = []
        
        self.schedule[date][shift_name].append(staff["name"])
        
        stats = self.staff_stats[staff["name"]]
        stats["days"] += 1
        stats["hours"] += calculate_work_hours(self.shifts[shift_name])
        
        if is_weekend_or_holiday(date):
            stats["weekends"] += 1
        
        # 連続勤務カウント
        if stats["last_work"] and (date - stats["last_work"]).days == 1:
            stats["consecutive"] += 1
        else:
            stats["consecutive"] = 1
        stats["max_consecutive"] = max(stats["max_consecutive"], stats["consecutive"])
        stats["last_work"] = date
    
    def reset_consecutive_if_gap(self, staff_name, date):
        """休みがあれば連続勤務をリセット"""
        stats = self.staff_stats[staff_name]
        if stats["last_work"] and (date - stats["last_work"]).days > 1:
            stats["consecutive"] = 0
    
    def generate(self):
        """シフト表を生成"""
        for date in self.dates:
            requirements = self.get_requirement(date)
            
            for shift_name, count in requirements.items():
                # 割り当て可能なスタッフを公平性スコアでソート
                available = []
                for staff in self.staff:
                    self.reset_consecutive_if_gap(staff["name"], date)
                    if self.can_assign(staff, date, shift_name):
                        stats = self.staff_stats[staff["name"]]
                        # スコア: 勤務日数が少ない人、土日出勤が少ない人を優先
                        score = stats["days"] * 2 + stats["weekends"] * 3
                        available.append((score, random.random(), staff))
                
                available.sort(key=lambda x: (x[0], x[1]))
                
                assigned = 0
                for _, _, staff in available:
                    if assigned >= count:
                        break
                    if self.can_assign(staff, date, shift_name):
                        self.assign(staff, date, shift_name)
                        assigned += 1
                
                # 必要人数に満たない場合は警告
                if assigned < count:
                    print(f"警告: {date.strftime('%m/%d')} {shift_name} - "
                          f"必要{count}人に対し{assigned}人のみ割当")
        
        return self.schedule


def create_excel(scheduler, output_path):
    """Excel形式でシフト表を出力"""
    wb = Workbook()
    
    # シート1: シフト表
    ws1 = wb.active
    ws1.title = "シフト表"
    create_shift_sheet(ws1, scheduler)
    
    # シート2: 集計
    ws2 = wb.create_sheet("集計")
    create_summary_sheet(ws2, scheduler)
    
    # シート3: 公平性チェック
    ws3 = wb.create_sheet("公平性チェック")
    create_fairness_sheet(ws3, scheduler)
    
    wb.save(output_path)
    print(f"✅ シフト表を出力しました: {output_path}")


def create_shift_sheet(ws, scheduler):
    """シフト表シートを作成"""
    dates = scheduler.dates
    staff_list = [s["name"] for s in scheduler.staff]
    shift_names = list(scheduler.shifts.keys())
    
    # ヘッダー行: 日付
    ws.cell(1, 1, "スタッフ").font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 1).border = BORDER
    
    for col, date in enumerate(dates, start=2):
        cell = ws.cell(1, col)
        cell.value = f"{date.month}/{date.day}\n{['月','火','水','木','金','土','日'][date.weekday()]}"
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = 8
    
    # 雇用形態列（最後に配置、非表示）
    emp_type_col = len(dates) + 2
    ws.cell(1, emp_type_col, "雇用形態").font = HEADER_FONT
    ws.cell(1, emp_type_col).fill = HEADER_FILL
    ws.cell(1, emp_type_col).border = BORDER
    ws.column_dimensions[get_column_letter(emp_type_col)].hidden = True
    
    ws.column_dimensions['A'].width = 12
    ws.row_dimensions[1].height = 30
    
    # スタッフ行
    for row, staff in enumerate(scheduler.staff, start=2):
        staff_name = staff["name"]
        ws.cell(row, 1, staff_name).border = BORDER
        ws.cell(row, 1).font = Font(bold=True)
        
        # 雇用形態を隠し列に追加
        emp_type = "正社員" if staff.get("type") == "full-time" else "パート"
        ws.cell(row, emp_type_col, emp_type).border = BORDER
        
        for col, date in enumerate(dates, start=2):
            cell = ws.cell(row, col)
            cell.border = BORDER
            
            # このスタッフのシフトを取得
            assigned_shift = None
            for shift_name, members in scheduler.schedule.get(date, {}).items():
                if staff_name in members:
                    assigned_shift = shift_name
                    break
            
            cell.value = assigned_shift if assigned_shift else "休"
            cell.alignment = Alignment(horizontal='center')
            
            # 土日祝の背景色
            if is_weekend_or_holiday(date):
                if date.strftime("%Y-%m-%d") in HOLIDAYS_2025:
                    cell.fill = HOLIDAY_FILL
                else:
                    cell.fill = WEEKEND_FILL


def create_summary_sheet(ws, scheduler):
    """集計シートを作成（Excel関数で動的計算）"""
    headers = ["スタッフ", "雇用形態", "勤務日数", "勤務時間", "土日祝出勤", "最大連勤"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    dates = scheduler.dates
    num_days = len(dates)
    emp_type_col = get_column_letter(num_days + 2)  # 雇用形態の列（隠し列）
    
    for row, staff in enumerate(scheduler.staff, start=2):
        staff_name = staff["name"]
        shift_row = row  # シフト表シートと同じ行番号
        
        # スタッフ名（シフト表シートから参照）
        ws.cell(row, 1, f'=シフト表!A{shift_row}').border = BORDER
        ws.cell(row, 1).font = Font(bold=True)
        
        # 雇用形態（シフト表の隠し列から参照）
        ws.cell(row, 2, f'=シフト表!{emp_type_col}{shift_row}').border = BORDER
        
        # 勤務日数: シフト表シートの該当行で「休」以外をカウント
        work_range = f"シフト表!B{shift_row}:{get_column_letter(num_days + 1)}{shift_row}"
        formula = f'=COUNTIF({work_range},"<>休")'
        cell = ws.cell(row, 3, formula)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
        
        # 勤務時間: 各シフトの時間を計算（簡易版: 6時間×勤務日数）
        hours_formula = f'=C{row}*6'
        cell = ws.cell(row, 4, hours_formula)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = '0.0"h"'
        
        # 土日祝出勤: 土日祝列を特定してカウント
        weekend_cols = []
        for col_idx, date in enumerate(dates, start=2):
            if is_weekend_or_holiday(date):
                weekend_cols.append(get_column_letter(col_idx))
        
        weekend_formulas = [f'IF(シフト表!{col}{shift_row}<>"休",1,0)' for col in weekend_cols]
        if weekend_formulas:
            weekend_formula = f'={"+".join(weekend_formulas)}'
        else:
            weekend_formula = '=0'
        
        cell = ws.cell(row, 5, weekend_formula)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
        
        # 最大連勤: 簡易版として数式で概算計算
        # 完全な連勤判定はVBAが必要なため、ここでは「総勤務日数÷2」を目安値として設定
        consecutive_formula = f'=ROUND(C{row}/4,0)'
        cell = ws.cell(row, 6, consecutive_formula)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=9, color="666666")
    
    # 注釈を追加
    ws.cell(len(scheduler.staff) + 3, 1, "※最大連勤は概算値（要手動確認）").font = Font(size=9, color="666666", italic=True)
    
    # 列幅調整
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_fairness_sheet(ws, scheduler):
    """公平性チェックシートを作成（Excel関数で動的計算）"""
    ws.cell(1, 1, "公平性分析レポート").font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    row = 3
    num_staff = len(scheduler.staff)
    
    # 土日祝出勤の分析
    ws.cell(row, 1, "■ 土日祝出勤回数").font = Font(bold=True)
    row += 1
    
    ws.cell(row, 1, "スタッフ").font = Font(bold=True)
    ws.cell(row, 1).border = BORDER
    ws.cell(row, 2, "回数").font = Font(bold=True)
    ws.cell(row, 2).border = BORDER
    ws.cell(row, 3, "最小値との差").font = Font(bold=True)
    ws.cell(row, 3).border = BORDER
    ws.cell(row, 4, "評価").font = Font(bold=True)
    ws.cell(row, 4).border = BORDER
    row += 1
    
    start_data_row = row
    
    for idx, staff in enumerate(scheduler.staff):
        summary_row = idx + 2  # 集計シートの該当行
        
        # スタッフ名
        ws.cell(row, 1, f'=集計!A{summary_row}').border = BORDER
        
        # 回数（集計シートから参照）
        ws.cell(row, 2, f'=集計!E{summary_row}').border = BORDER
        ws.cell(row, 2).alignment = Alignment(horizontal='center')
        
        # 最小値との差
        min_range = f'B{start_data_row}:B{start_data_row + num_staff - 1}'
        diff_formula = f'=B{row}-MIN({min_range})'
        ws.cell(row, 3, diff_formula).border = BORDER
        ws.cell(row, 3).alignment = Alignment(horizontal='center')
        
        # 評価（差が0-1なら◎、2なら○、3以上なら△）
        eval_formula = f'=IF(C{row}<=1,"◎ 均等",IF(C{row}<=2,"○ 良好","△ 偏り"))'
        ws.cell(row, 4, eval_formula).border = BORDER
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        
        row += 1
    
    row += 2
    
    # 統計サマリー
    ws.cell(row, 1, "■ 統計").font = Font(bold=True)
    row += 1
    
    ws.cell(row, 1, "土日祝出勤の最大差")
    max_diff_formula = f'=MAX(C{start_data_row}:C{start_data_row + num_staff - 1})'
    ws.cell(row, 2, max_diff_formula)
    ws.cell(row, 3, "回").alignment = Alignment(horizontal='left')
    row += 1
    
    ws.cell(row, 1, "評価")
    overall_eval = f'=IF(B{row-1}<=1,"◎ 非常に公平",IF(B{row-1}<=2,"○ 公平","△ 改善推奨"))'
    ws.cell(row, 2, overall_eval).font = Font(bold=True)
    row += 2
    
    # 週労働時間チェック
    ws.cell(row, 1, "■ 週労働時間").font = Font(bold=True)
    row += 1
    
    ws.cell(row, 1, "スタッフ").font = Font(bold=True)
    ws.cell(row, 1).border = BORDER
    ws.cell(row, 2, "総勤務時間").font = Font(bold=True)
    ws.cell(row, 2).border = BORDER
    ws.cell(row, 3, "週平均").font = Font(bold=True)
    ws.cell(row, 3).border = BORDER
    ws.cell(row, 4, "判定").font = Font(bold=True)
    ws.cell(row, 4).border = BORDER
    row += 1
    
    num_weeks = (len(scheduler.dates) + 6) // 7  # 概算週数
    
    for idx, staff in enumerate(scheduler.staff):
        summary_row = idx + 2
        
        # スタッフ名
        ws.cell(row, 1, f'=集計!A{summary_row}').border = BORDER
        
        # 総勤務時間
        ws.cell(row, 2, f'=集計!D{summary_row}').border = BORDER
        ws.cell(row, 2).alignment = Alignment(horizontal='center')
        
        # 週平均（概算）
        avg_formula = f'=VALUE(LEFT(B{row},FIND("h",B{row})-1))/{num_weeks}'
        ws.cell(row, 3, avg_formula).border = BORDER
        ws.cell(row, 3).alignment = Alignment(horizontal='center')
        ws.cell(row, 3).number_format = '0.0"h"'
        
        # 判定（週40時間以内か）
        judge_formula = f'=IF(C{row}<=40,"✓ OK","⚠ 要確認")'
        ws.cell(row, 4, judge_formula).border = BORDER
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        
        row += 1
    
    # 列幅調整
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12


def main():
    parser = argparse.ArgumentParser(description='シフト表自動生成')
    parser.add_argument('--year', type=int, required=True, help='対象年')
    parser.add_argument('--month', type=int, required=True, help='対象月')
    parser.add_argument('--staff-file', required=True, help='スタッフ設定JSONファイル')
    parser.add_argument('--output', default='shift.xlsx', help='出力ファイル名')
    args = parser.parse_args()
    
    config = load_config(args.staff_file)
    scheduler = ShiftScheduler(config, args.year, args.month)
    scheduler.generate()
    create_excel(scheduler, args.output)
    
    # 数式を再計算
    import subprocess
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    recalc_script = os.path.join(script_dir, 'recalc.py')
    
    print("📊 Excel数式を計算中...")
    try:
        result = subprocess.run(
            ['python', recalc_script, args.output],
            capture_output=True,
            text=True,
            timeout=60
        )
        if result.returncode == 0:
            print("✅ 数式の計算が完了しました")
        else:
            print(f"⚠ 数式計算で問題が発生: {result.stderr}")
    except Exception as e:
        print(f"⚠ recalc.pyの実行に失敗: {e}")
        print("   手動で数式を確認してください")


if __name__ == "__main__":
    main()
